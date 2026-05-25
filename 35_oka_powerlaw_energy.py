"""
35_oka_powerlaw_energy.py
======================================================================
CPU DirectPDK vs OKA device-aware PDK simulation with power-law
photoresponse and CDS/read-pulse-based energy accounting.

Main model
----------
I_photo = I_light - I_dark = R(I_state) * P^alpha
R(I_state) is fitted from experimental programmed-state-dependent
photoresponse data and used as a compact empirical model.

OKA signed kernel representation
--------------------------------
effective weight ∝ R(I_state) - R_ref
R_ref is a median or weight-aware reference response. This emulates signed
kernel operation by bias/reference subtraction while each device response is
positive.

Read energy
-----------
CDS mode:
  E_read = V_read * t_read * sum(I_dark + I_light)
Light-only mode:
  E_read = V_read * t_read * sum(I_light)

Run
---
python 35_oka_powerlaw_energy.py
"""

import argparse
import csv
import glob
import os
import random
from dataclasses import dataclass

import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
from PIL import Image
from scipy.optimize import curve_fit
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl

SEED = 42
torch.manual_seed(SEED)
np.random.seed(SEED)
random.seed(SEED)


def parse_float_list(s):
    return [float(x.strip()) for x in s.split(",") if x.strip()]


def energy_v_nA_us_to_pJ(v_read, current_nA, time_us):
    # V * nA * us = 1e-15 J = 0.001 pJ
    return v_read * current_nA * time_us * 1e-3


@dataclass
class PowerLawPhotoModel:
    alpha: float
    I_states: np.ndarray
    R_states: np.ndarray
    R_fit_mode: str = "exp"
    A: float = None
    I0: float = None

    def __post_init__(self):
        self.I_states = np.asarray(self.I_states, dtype=float)
        self.R_states = np.asarray(self.R_states, dtype=float)
        order = np.argsort(self.I_states)
        self.I_states = self.I_states[order]
        self.R_states = self.R_states[order]
        if self.R_fit_mode == "exp":
            self._fit_exp()

    def _fit_exp(self):
        def exp_model(I, A, I0):
            return A * np.exp(-np.asarray(I) / I0)
        A0 = max(float(np.max(self.R_states)), 1e-12)
        I00 = max(float(np.mean(self.I_states)), 1.0)
        try:
            popt, _ = curve_fit(
                exp_model, self.I_states, self.R_states,
                p0=[A0, I00], bounds=([1e-12, 1e-9], [1e6, 1e6]),
                maxfev=20000
            )
            self.A, self.I0 = float(popt[0]), float(popt[1])
        except Exception:
            y = np.log(np.clip(self.R_states, 1e-12, None))
            m, b = np.polyfit(self.I_states, y, 1)
            self.I0 = float(max(-1.0 / m, 1e-9)) if m < 0 else I00
            self.A = float(np.exp(b))

    def R(self, I_state):
        I_state = np.asarray(I_state, dtype=float)
        if self.R_fit_mode == "exp":
            return self.A * np.exp(-I_state / self.I0)
        return np.interp(
            I_state, self.I_states, self.R_states,
            left=self.R_states[0], right=self.R_states[-1]
        )

    def I_photo(self, P, I_state):
        return self.R(I_state) * np.power(np.asarray(P, dtype=float), self.alpha)

    @classmethod
    def fit_from_xlsx(cls, path, R_fit_mode="exp", alpha_bounds=(0.05, 2.0)):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))

        I_states = []
        for s in rows[1][1:]:
            if s is None:
                continue
            if isinstance(s, str):
                I_states.append(float(s.replace("nA", "").strip()))
            else:
                I_states.append(float(s))
        I_states = np.asarray(I_states, dtype=float)

        P_vals, Iphoto = [], []
        for r in rows[2:]:
            if r[0] is None:
                continue
            P_vals.append(float(r[0]))
            Iphoto.append([float(v) for v in r[1:1+len(I_states)]])
        P_vals = np.asarray(P_vals, dtype=float)
        Iphoto = np.asarray(Iphoto, dtype=float)

        n_state = len(I_states)
        P_grid, S_grid = np.meshgrid(P_vals, np.arange(n_state), indexing="ij")

        def model(X, *params):
            P_in, S_in = X
            log_Rs = np.asarray(params[:-1])
            alpha = params[-1]
            R_s = np.exp(log_Rs[S_in.astype(int)])
            return R_s * np.power(P_in, alpha)

        R_init = np.mean(Iphoto / np.clip(P_vals[:, None], 1e-12, None), axis=0)
        p0 = list(np.log(np.clip(R_init, 1e-12, None))) + [1.0]
        lower = [-30.0] * n_state + [alpha_bounds[0]]
        upper = [30.0] * n_state + [alpha_bounds[1]]
        popt, _ = curve_fit(
            model, (P_grid.flatten(), S_grid.flatten()), Iphoto.flatten(),
            p0=p0, bounds=(lower, upper), maxfev=50000
        )
        R_states = np.exp(np.asarray(popt[:-1]))
        alpha = float(popt[-1])
        pred = model((P_grid.flatten(), S_grid.flatten()), *popt)
        truth = Iphoto.flatten()
        rmse = float(np.sqrt(np.mean((truth - pred) ** 2)))
        rel = float(np.mean(np.abs((truth - pred) / np.clip(truth, 1e-9, None))) * 100)

        print("[PR power-law fit]")
        print(f"  I_photo = R(I_state) * P^alpha")
        print(f"  alpha={alpha:.4f}")
        print(f"  I_states={np.round(I_states, 3)} nA")
        print(f"  R_states={np.round(R_states, 6)}")
        print(f"  RMSE={rmse:.4f} nA, mean rel err={rel:.2f}%")

        obj = cls(alpha, I_states, R_states, R_fit_mode=R_fit_mode)
        if R_fit_mode == "exp":
            print(f"[R fit] R(I)=A exp(-I/I0), A={obj.A:.6g}, I0={obj.I0:.3f} nA")
        else:
            print("[R fit] interpolation mode")
        return obj, (P_vals, I_states, Iphoto)


@dataclass
class LTDModel:
    V_levels: np.ndarray
    params: dict
    I_init_avg: float
    I_min: float
    I_max: float

    @staticmethod
    def _double_exp(N, a1, t1, a2, t2, Iinf):
        return a1*np.exp(-N/np.abs(t1)) + a2*np.exp(-N/np.abs(t2)) + Iinf

    @classmethod
    def fit_from_xlsx(cls, path, v_assignment=None):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        data = np.array([list(r) for r in ws.iter_rows(values_only=True)], dtype=float)
        pulse = data[:, 0].astype(int)
        curves_nA = data[:, 1:] * 1e9

        ratio = curves_nA[0] / np.maximum(curves_nA[-1], 1e-9)
        control_idx = int(np.argmin(np.abs(ratio - 1.0)))
        delta = curves_nA[0] - curves_nA[-1]
        order = np.argsort(-delta)
        valid_idx = [i for i in order if i != control_idx]

        if v_assignment is None:
            v_assignment = [10.0, 9.5, 9.0, 8.5, 8.0, 7.5, 7.0, 6.5, 6.0]
        if len(v_assignment) < len(valid_idx):
            raise ValueError("v_assignment is shorter than the number of fitted LTD curves")
        col_to_V = dict(zip(valid_idx, v_assignment[:len(valid_idx)]))

        params = {}
        for i in valid_idx:
            V = float(col_to_V[i])
            y = curves_nA[:, i]
            try:
                p, _ = curve_fit(
                    cls._double_exp, pulse, y,
                    p0=[max(y[0]*0.7, 1e-9), 5, max(y[0]*0.3, 1e-9), 50, max(y[-1], 1e-9)],
                    bounds=([0, 0.1, 0, 0.1, 0], [1e4, 1e5, 1e4, 1e6, 1e4]),
                    maxfev=50000
                )
                params[V] = tuple(p)
            except Exception as e:
                print(f"[Warning] LTD fit failed at V={V}: {e}")

        V_levels = np.array(sorted(params.keys()), dtype=float)
        I_init_avg = float(curves_nA[0].mean())
        I_max = I_init_avg
        I_min = float(min(cls._double_exp(120, *params[V]) for V in V_levels))
        print("[LTD fit]")
        print(f"  V_levels={list(V_levels)}")
        print(f"  I range={I_min:.3f} ~ {I_max:.3f} nA")
        return cls(V_levels, params, I_init_avg, I_min, I_max)

    def current_after(self, V, user_N):
        if user_N <= 0 or V <= 0:
            return float(self.I_init_avg)
        return float(self._double_exp(user_N + 1, *self.params[float(V)]))

    def best_V_N(self, target_I, N_max=3):
        candidates = [(0.0, 0, float(self.I_init_avg), abs(float(self.I_init_avg) - target_I))]
        for V in self.V_levels:
            for N in range(1, int(N_max) + 1):
                I = self.current_after(V, N)
                candidates.append((float(V), int(N), I, abs(I - target_I)))
        candidates.sort(key=lambda x: (x[3], x[1], x[0]))
        return candidates[0]


def simulate_discrete_programming(I_target_map, ltd_model, N_budget):
    H, W, K = I_target_map.shape
    I_actual = np.zeros_like(I_target_map, dtype=float)
    V_map = np.zeros_like(I_target_map, dtype=float)
    N_map = np.zeros_like(I_target_map, dtype=np.int32)
    err_map = np.zeros_like(I_target_map, dtype=float)
    for y in range(H):
        for x in range(W):
            for k in range(K):
                V, N, I, err = ltd_model.best_V_N(float(I_target_map[y, x, k]), N_max=N_budget)
                V_map[y, x, k] = V
                N_map[y, x, k] = N
                I_actual[y, x, k] = I
                err_map[y, x, k] = err
    n_total = H * W * K
    print("[Discrete PGM]")
    print(f"  cells={n_total:,}, avg pulse/cell={N_map.mean():.3f}, max={int(N_map.max())}")
    print(f"  no-pulse={(N_map==0).sum():,} ({100*(N_map==0).sum()/n_total:.2f}%)")
    print(f"  pgm err avg={err_map.mean():.4f} nA, max={err_map.max():.4f} nA")
    return I_actual, V_map, N_map, err_map


# Image processing

def radial_map(H, W, device="cpu"):
    ys = torch.linspace(-1, 1, H, device=device)
    xs = torch.linspace(-1, 1, W, device=device)
    gy, gx = torch.meshgrid(ys, xs, indexing="ij")
    r = (gy**2 + gx**2).sqrt() / (2**0.5)
    return r, gy, gx


def spatially_varying_conv(x, kernels):
    B, C, H, W = x.shape
    patches = F.unfold(x, kernel_size=3, padding=1).view(B, C, 9, H*W)
    k = kernels.view(H*W, 9).T.unsqueeze(0).unsqueeze(0)
    return (patches * k).sum(dim=2).view(B, C, H, W).clamp(0, 1)


def degrade_coma_vig(x, sigma0=0.40, alpha_psf=2.0, coma_k=0.60, alpha_vig=4.0):
    r, gy, gx = radial_map(*x.shape[2:], x.device)
    r2 = r**2
    rs = r.clamp(1e-6)
    sigma = (sigma0 + alpha_psf*r2).clamp(0.1, 2.0)
    shift_y = coma_k*r2*(gy/rs)
    shift_x = coma_k*r2*(gx/rs)
    coords = torch.tensor([[-1.,-1.],[-1.,0.],[-1.,1.],[0.,-1.],[0.,0.],[0.,1.],[1.,-1.],[1.,0.],[1.,1.]], device=x.device)
    d2 = (coords**2).sum(-1)
    s = sigma.unsqueeze(-1)
    g1 = torch.exp(-d2/(2*s**2)); g1 = g1/g1.sum(-1, keepdim=True)
    sy = shift_y.unsqueeze(-1); sx = shift_x.unsqueeze(-1)
    dy = coords[:,0].unsqueeze(0).unsqueeze(0) - sy
    dx = coords[:,1].unsqueeze(0).unsqueeze(0) - sx
    g2 = torch.exp(-(dy**2+dx**2)/(2*s**2)); g2 = g2/g2.sum(-1, keepdim=True)
    w = (r*0.9).clamp(0, 0.9).unsqueeze(-1)
    k = (1-w)*g1 + w*g2
    k = k/k.sum(-1, keepdim=True)
    x = spatially_varying_conv(x, k)
    V = 1.0/(1.0 + alpha_vig*r2)
    return (x*V.unsqueeze(0).unsqueeze(0)).clamp(0, 1)


class DirectPDK(nn.Module):
    def __init__(self, H, W):
        super().__init__()
        init = torch.zeros(H, W, 9)
        init[:, :, 4] = 1.0
        self.kernels = nn.Parameter(init)
    def forward(self, x_eff):
        return spatially_varying_conv(x_eff, self.kernels)


def loss_fn(out, clean):
    l1 = F.l1_loss(out, clean)
    ox = out[:,:,:,1:] - out[:,:,:,:-1]
    oy = out[:,:,1:,:] - out[:,:,:-1,:]
    cx = clean[:,:,:,1:] - clean[:,:,:,:-1]
    cy = clean[:,:,1:,:] - clean[:,:,:-1,:]
    return l1 + 0.1*(F.l1_loss(ox, cx) + F.l1_loss(oy, cy))


def train_pdk(model, train_imgs, degrade_fn, alpha_signal, n_iter, lr, batch):
    opt = torch.optim.Adam(model.parameters(), lr=lr)
    sched = torch.optim.lr_scheduler.CosineAnnealingLR(opt, T_max=n_iter)
    X = torch.cat(train_imgs, dim=0)
    n = X.shape[0]
    for i in range(n_iter):
        idx = torch.randperm(n)[:batch]
        clean = X[idx]
        deg = degrade_fn(clean)
        deg_eff = deg.clamp(0, 1).pow(alpha_signal)
        opt.zero_grad()
        loss = loss_fn(model(deg_eff), clean)
        loss.backward(); opt.step(); sched.step()
        if (i+1) % 200 == 0:
            print(f"  iter {i+1}/{n_iter}, loss={loss.item():.5f}")


# Weight mapping

def map_weights_to_I_state(weights_t, pr_model, ltd_model, N_budget=3, ref_mode="weight_aware", ref_percentile=50.0):
    weights = weights_t.detach().cpu().numpy()
    I_high = float(ltd_model.I_init_avg)
    strongest_V = float(np.max(ltd_model.V_levels))
    I_low = float(ltd_model.current_after(strongest_V, N_budget))
    R1, R2 = float(pr_model.R(I_high)), float(pr_model.R(I_low))
    R_min, R_max = min(R1, R2), max(R1, R2)
    w_pos = max(0.0, float(weights.max()))
    w_neg = max(0.0, float(-weights.min()))
    if ref_mode == "median":
        R_ref = float(np.percentile([R_min, R_max], ref_percentile))
    else:
        R_ref = (w_pos * R_min + w_neg * R_max) / (w_pos + w_neg + 1e-12)
    R_ref = float(np.clip(R_ref, R_min + 1e-12, R_max - 1e-12))
    scale_pos = w_pos / max(R_max - R_ref, 1e-12)
    scale_neg = w_neg / max(R_ref - R_min, 1e-12)
    scale = max(scale_pos, scale_neg, 1e-12)
    R_target = np.clip(R_ref + weights/scale, R_min + 1e-12, R_max - 1e-12)
    if pr_model.R_fit_mode == "exp":
        I_target = -pr_model.I0 * np.log(R_target / pr_model.A)
    else:
        I_grid = np.linspace(min(I_low, I_high), max(I_low, I_high), 2000)
        R_grid = pr_model.R(I_grid)
        order = np.argsort(R_grid)
        I_target = np.interp(R_target, R_grid[order], I_grid[order])
    I_target = np.clip(I_target, min(I_low, I_high), max(I_low, I_high))
    info = dict(I_low=I_low, I_high=I_high, R_min=R_min, R_max=R_max, R_ref=R_ref, scale=float(scale), ref_mode=ref_mode, N_budget=N_budget)
    print("[Weight -> OKA]")
    print(f"  weight range={weights.min():+.4f} ~ {weights.max():+.4f}")
    print(f"  R range={R_min:.6g} ~ {R_max:.6g}, R_ref={R_ref:.6g}, scale={scale:.6g}")
    print(f"  I_target range={I_target.min():.3f} ~ {I_target.max():.3f} nA")
    return I_target, float(scale), info


def unfold_patches_np(img_hw):
    H, W = img_hw.shape
    pad = np.pad(img_hw, 1, mode="edge")
    patches = np.zeros((H, W, 9), dtype=float)
    coords = [(-1,-1),(-1,0),(-1,1),(0,-1),(0,0),(0,1),(1,-1),(1,0),(1,1)]
    for k, (dy, dx) in enumerate(coords):
        patches[:, :, k] = pad[1+dy:1+dy+H, 1+dx:1+dx+W]
    return patches


def oka_forward_image(deg_t, I_actual_map, pr_model, mapping_info, P_scale):
    deg = deg_t.squeeze().detach().cpu().numpy().astype(float)
    X_patches = unfold_patches_np(deg)
    P_patches = P_scale * X_patches
    R_actual = pr_model.R(I_actual_map)
    R_ref = float(mapping_info["R_ref"])
    scale = float(mapping_info["scale"])
    signal_nA = ((R_actual - R_ref) * np.power(P_patches, pr_model.alpha)).sum(axis=2)
    out = signal_nA * scale / (P_scale ** pr_model.alpha)
    return torch.from_numpy(np.clip(out, 0, 1)).float().unsqueeze(0).unsqueeze(0)


def oka_read_energy_pJ(deg_t, I_actual_map, pr_model, P_scale, V_read=1.0, t_read_us=1.0, read_mode="cds", include_ref_branch=False, R_ref=None):
    deg = deg_t.squeeze().detach().cpu().numpy().astype(float)
    P_patches = P_scale * unfold_patches_np(deg)
    I_dark = np.asarray(I_actual_map, dtype=float)
    I_photo = pr_model.R(I_dark) * np.power(P_patches, pr_model.alpha)
    I_light = I_dark + I_photo
    if read_mode == "cds":
        current_sum_nA = float((I_dark + I_light).sum())
    elif read_mode == "light_only":
        current_sum_nA = float(I_light.sum())
    else:
        raise ValueError("read_mode must be cds or light_only")
    if include_ref_branch:
        if R_ref is None:
            raise ValueError("R_ref required for reference branch")
        I_grid = np.linspace(np.min(I_dark), np.max(I_dark), 2000)
        I_ref = float(I_grid[np.argmin(np.abs(pr_model.R(I_grid) - R_ref))])
        I_ref_dark = np.full_like(I_dark, I_ref)
        I_ref_photo = R_ref * np.power(P_patches, pr_model.alpha)
        I_ref_light = I_ref_dark + I_ref_photo
        current_sum_nA += float((I_ref_dark + I_ref_light).sum() if read_mode == "cds" else I_ref_light.sum())
    return energy_v_nA_us_to_pJ(V_read, current_sum_nA, t_read_us), current_sum_nA


def oka_programming_energy_pJ(V_map, N_map, C_gate_fF=345.46):
    return float((0.5 * C_gate_fF * (V_map**2) * N_map * 1e-3).sum())


ENERGY = dict(cpu_mac_pj=100.0, dram_read_pj_bit=5.0, dram_write_pj_bit=5.0, sram_read_pj_bit=0.05, sram_write_pj_bit=0.05, adc_pj_sample=5.0, sensor_pj_pix=0.5)
LATENCY = dict(cpu_mac_ns=1.0, dram_ns=50.0, adc_ns=100.0, sensor_ns=1000.0, sram_ns=1.0)


def cpu_pdk_energy_latency(H, W, k=9, n_bit=8, mode="pdk_map_dram"):
    n_pix = H * W
    e_sensor = n_pix * ENERGY["sensor_pj_pix"]
    e_adc = n_pix * ENERGY["adc_pj_sample"]
    e_mac = n_pix * k * ENERGY["cpu_mac_pj"]
    if mode == "pdk_map_dram":
        e_raw_write = n_pix*n_bit*ENERGY["dram_write_pj_bit"]
        e_input_read = n_pix*k*n_bit*ENERGY["dram_read_pj_bit"]
        e_weight_read = n_pix*k*n_bit*ENERGY["dram_read_pj_bit"]
        e_out_write = n_pix*n_bit*ENERGY["dram_write_pj_bit"]
    elif mode == "pdk_map_sram":
        e_raw_write = n_pix*n_bit*ENERGY["dram_write_pj_bit"]
        e_input_read = n_pix*k*n_bit*ENERGY["sram_read_pj_bit"]
        e_weight_read = n_pix*k*n_bit*ENERGY["sram_read_pj_bit"]
        e_out_write = n_pix*n_bit*ENERGY["dram_write_pj_bit"]
    elif mode == "mac_only":
        e_raw_write = 0.0; e_input_read = 0.0; e_weight_read = 0.0
        e_out_write = n_pix*n_bit*ENERGY["sram_write_pj_bit"]
    else:
        raise ValueError("unknown cpu mode")
    total = e_sensor + e_adc + e_raw_write + e_input_read + e_weight_read + e_mac + e_out_write
    cores_simd = 40.0
    t_sensor = LATENCY["sensor_ns"]
    t_adc = H * LATENCY["adc_ns"]
    t_mac = n_pix*k*LATENCY["cpu_mac_ns"]/cores_simd
    if mode == "pdk_map_dram":
        t_mem = n_pix*(2*k+2)*LATENCY["dram_ns"]/cores_simd
    elif mode == "pdk_map_sram":
        t_mem = n_pix*2*LATENCY["dram_ns"]/cores_simd + n_pix*(2*k)*LATENCY["sram_ns"]/cores_simd
    else:
        t_mem = 0.0
    return total, t_sensor+t_adc+t_mac+t_mem, dict(sensor=e_sensor, adc=e_adc, raw_write=e_raw_write, input_read=e_input_read, weight_read=e_weight_read, mac=e_mac, out_write=e_out_write)


# IO and metrics

def load_train_images(data_dir, n, size):
    exts = ("*.png", "*.jpg", "*.jpeg", "*.PNG", "*.JPG", "*.JPEG")
    paths = []
    for ext in exts:
        paths += glob.glob(os.path.join(data_dir, "**", ext), recursive=True)
        paths += glob.glob(os.path.join(data_dir, ext))
    paths = sorted(set(paths)); random.shuffle(paths)
    imgs = []
    for p in paths:
        try:
            img = Image.open(p).convert("L")
            w, h = img.size; s = min(w, h)
            img = img.crop(((w-s)//2, (h-s)//2, (w+s)//2, (h+s)//2)).resize((size, size), Image.BILINEAR)
            imgs.append(torch.from_numpy(np.array(img, dtype=np.float32)/255.).unsqueeze(0).unsqueeze(0))
        except Exception:
            continue
        if len(imgs) == n:
            break
    print(f"[Load] {len(imgs)} train images from {data_dir}")
    return imgs


def find_image(data_dir, name):
    exts = (".png", ".jpg", ".jpeg", ".PNG", ".JPG", ".JPEG")
    for ext in exts:
        c = os.path.join(data_dir, name+ext)
        if os.path.exists(c): return c
        hits = glob.glob(os.path.join(data_dir, "**", name+ext), recursive=True)
        if hits: return hits[0]
    return None


def load_image(path, size):
    img = Image.open(path).convert("L")
    w, h = img.size; s = min(w, h)
    img = img.crop(((w-s)//2, (h-s)//2, (w+s)//2, (h+s)//2)).resize((size, size), Image.BILINEAR)
    return torch.from_numpy(np.array(img, dtype=np.float32)/255.).unsqueeze(0).unsqueeze(0)


def psnr(a, b):
    mse = F.mse_loss(a.float(), b.float()).item()
    return 99.9 if mse < 1e-10 else 10*np.log10(1.0/mse)


def ssim_simple(a, b):
    a, b = a.float(), b.float()
    mu_a = F.avg_pool2d(a, 11, stride=1, padding=5)
    mu_b = F.avg_pool2d(b, 11, stride=1, padding=5)
    mu_a2, mu_b2, mu_ab = mu_a**2, mu_b**2, mu_a*mu_b
    sa2 = F.avg_pool2d(a**2, 11, stride=1, padding=5) - mu_a2
    sb2 = F.avg_pool2d(b**2, 11, stride=1, padding=5) - mu_b2
    sab = F.avg_pool2d(a*b, 11, stride=1, padding=5) - mu_ab
    c1, c2 = 0.01**2, 0.03**2
    return (((2*mu_ab+c1)*(2*sab+c2))/((mu_a2+mu_b2+c1)*(sa2+sb2+c2))).mean().item()


def save_img(t, path):
    arr = np.clip(t.squeeze().detach().cpu().numpy(), 0, 1)
    fig, ax = plt.subplots(1, 1, figsize=(arr.shape[1]/100, arr.shape[0]/100), dpi=300)
    ax.imshow(arr, cmap="gray", vmin=0, vmax=1); ax.axis("off")
    plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
    fig.savefig(path, dpi=300, bbox_inches="tight", pad_inches=0)
    plt.close(fig)


def save_compare(clean, deg, cpu, oka, metrics, path):
    fig, axes = plt.subplots(1, 4, figsize=(16, 4.2))
    titles = ["Input", f"Degraded\nPSNR={metrics['psnr_deg']:.2f}", f"CPU DirectPDK\nPSNR={metrics['psnr_cpu']:.2f}", f"OKA device-aware\nPSNR={metrics['psnr_oka']:.2f}"]
    for ax, title, img in zip(axes, titles, [clean, deg, cpu, oka]):
        ax.imshow(img.squeeze().detach().cpu().numpy(), cmap="gray", vmin=0, vmax=1)
        ax.set_title(title, fontsize=10, fontweight="bold"); ax.axis("off")
    plt.tight_layout(); fig.savefig(path, dpi=150, bbox_inches="tight"); plt.close(fig)


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--pr-data", default="./expdata/photoresponsivity.xlsx")
    p.add_argument("--ltd-data", default="./expdata/depression_data.xlsx")
    p.add_argument("--train-dir", default="./data/BSD300/images/train")
    p.add_argument("--inf-dir", default="./data/BSD300/images/test")
    p.add_argument("--inf-name", default="102061")
    p.add_argument("--img-size", type=int, default=256)
    p.add_argument("--R-fit-mode", choices=["exp", "interp"], default="exp")
    p.add_argument("--alpha-min", type=float, default=0.05)
    p.add_argument("--alpha-max", type=float, default=2.0)
    p.add_argument("--n-train", type=int, default=200)
    p.add_argument("--pdk-iter", type=int, default=2000)
    p.add_argument("--pdk-lr", type=float, default=0.01)
    p.add_argument("--batch", type=int, default=32)
    p.add_argument("--sigma0", type=float, default=0.40)
    p.add_argument("--alpha-psf", type=float, default=2.0)
    p.add_argument("--coma-k", type=float, default=0.60)
    p.add_argument("--alpha-vig", type=float, default=4.0)
    p.add_argument("--N-budget", type=int, default=3)
    p.add_argument("--ref-mode", choices=["weight_aware", "median"], default="weight_aware")
    p.add_argument("--ref-percentile", type=float, default=50.0)
    p.add_argument("--v-assignment", default="10,9.5,9,8.5,8,7.5,7,6.5,6")
    p.add_argument("--P-scale", type=float, default=1000.0)
    p.add_argument("--V-read", type=float, default=1.0)
    p.add_argument("--read-mode", choices=["cds", "light_only"], default="cds")
    p.add_argument("--include-ref-branch", action="store_true")
    p.add_argument("--lab-read-ms", type=float, default=10.0)
    p.add_argument("--fast-read-us", type=float, default=1.0)
    p.add_argument("--cpu-mode", choices=["pdk_map_dram", "pdk_map_sram", "mac_only"], default="pdk_map_dram")
    p.add_argument("--C-gate-fF", type=float, default=345.46)
    p.add_argument("--n-reuse", type=int, default=1000)
    p.add_argument("--res-dir", default="./res/oka35_powerlaw_energy")
    return p.parse_args()


def main():
    args = parse_args()
    os.makedirs(args.res_dir, exist_ok=True)
    ind_dir = os.path.join(args.res_dir, "individual"); os.makedirs(ind_dir, exist_ok=True)
    H = W = args.img_size

    print("="*70); print("[1] Fit PR and LTD models"); print("="*70)
    pr_model, _ = PowerLawPhotoModel.fit_from_xlsx(args.pr_data, R_fit_mode=args.R_fit_mode, alpha_bounds=(args.alpha_min, args.alpha_max))
    ltd_model = LTDModel.fit_from_xlsx(args.ltd_data, v_assignment=parse_float_list(args.v_assignment))

    print("\n" + "="*70); print("[2] Train CPU DirectPDK"); print("="*70)
    train_imgs = load_train_images(args.train_dir, args.n_train, H)
    if len(train_imgs) == 0:
        raise RuntimeError(f"No training images found in {args.train_dir}")
    degrade = lambda x: degrade_coma_vig(x, args.sigma0, args.alpha_psf, args.coma_k, args.alpha_vig)
    pdk_cpu = DirectPDK(H, W)
    train_pdk(pdk_cpu, train_imgs, degrade, pr_model.alpha, args.pdk_iter, args.pdk_lr, args.batch)
    weights = pdk_cpu.kernels.detach()
    torch.save({"kernels": weights, "alpha_signal": pr_model.alpha}, os.path.join(args.res_dir, "cpu_directpdk.pt"))

    print("\n" + "="*70); print("[3] Map weights to OKA and simulate discrete PGM"); print("="*70)
    I_target_map, scale_oka, mapping_info = map_weights_to_I_state(weights, pr_model, ltd_model, args.N_budget, args.ref_mode, args.ref_percentile)
    I_actual_map, V_map, N_map, err_map = simulate_discrete_programming(I_target_map, ltd_model, args.N_budget)
    torch.save({"I_target": torch.from_numpy(I_target_map).float(), "I_actual": torch.from_numpy(I_actual_map).float(), "V_map": V_map, "N_map": N_map, "err_map": err_map, "mapping_info": mapping_info, "scale_oka": scale_oka, "alpha": pr_model.alpha}, os.path.join(args.res_dir, "oka_mapped_device_states.pt"))

    print("\n" + "="*70); print("[4] Inference"); print("="*70)
    src = find_image(args.inf_dir, args.inf_name)
    if src is None:
        raise RuntimeError(f"Could not find {args.inf_name} in {args.inf_dir}")
    clean = load_image(src, H)
    with torch.no_grad():
        deg = degrade(clean)
        deg_eff = deg.clamp(0, 1).pow(pr_model.alpha)
        out_cpu = pdk_cpu(deg_eff).clamp(0, 1)
        out_oka = oka_forward_image(deg, I_actual_map, pr_model, mapping_info, args.P_scale).clamp(0, 1)
    metrics = dict(psnr_deg=psnr(clean, deg), ssim_deg=ssim_simple(clean, deg), psnr_cpu=psnr(clean, out_cpu), ssim_cpu=ssim_simple(clean, out_cpu), psnr_oka=psnr(clean, out_oka), ssim_oka=ssim_simple(clean, out_oka), psnr_cpu_oka=psnr(out_cpu, out_oka), ssim_cpu_oka=ssim_simple(out_cpu, out_oka))
    for k, v in metrics.items(): print(f"  {k}: {v:.4f}")
    save_compare(clean, deg, out_cpu, out_oka, metrics, os.path.join(args.res_dir, "inference_compare.png"))
    for name, img in [("input", clean), ("degraded", deg), ("cpu", out_cpu), ("oka", out_oka)]:
        save_img(img, os.path.join(ind_dir, f"{args.inf_name}_{name}.png"))

    print("\n" + "="*70); print("[5] Energy / latency"); print("="*70)
    e_cpu_pJ, t_cpu_ns, br_cpu = cpu_pdk_energy_latency(H, W, mode=args.cpu_mode)
    e_pgm_total_pJ = oka_programming_energy_pJ(V_map, N_map, args.C_gate_fF)
    e_pgm_amort_pJ = e_pgm_total_pJ / max(args.n_reuse, 1)
    lab_read_us = args.lab_read_ms * 1000.0
    e_read_lab_pJ, cur_lab = oka_read_energy_pJ(deg, I_actual_map, pr_model, args.P_scale, args.V_read, lab_read_us, args.read_mode, args.include_ref_branch, mapping_info["R_ref"])
    e_read_fast_pJ, cur_fast = oka_read_energy_pJ(deg, I_actual_map, pr_model, args.P_scale, args.V_read, args.fast_read_us, args.read_mode, args.include_ref_branch, mapping_info["R_ref"])
    n_pix = H * W
    e_oka_adc_pJ = n_pix * ENERGY["adc_pj_sample"]
    e_oka_sram_pJ = n_pix * 8 * ENERGY["sram_write_pj_bit"]
    e_oka_lab_pJ = e_pgm_amort_pJ + e_read_lab_pJ + e_oka_adc_pJ + e_oka_sram_pJ
    e_oka_fast_pJ = e_pgm_amort_pJ + e_read_fast_pJ + e_oka_adc_pJ + e_oka_sram_pJ
    t_oka_lab_ns = lab_read_us * 1000.0 + H * LATENCY["adc_ns"]
    t_oka_fast_ns = args.fast_read_us * 1000.0 + H * LATENCY["adc_ns"]

    print(f"[CPU] mode={args.cpu_mode}, energy={e_cpu_pJ/1e6:.6f} uJ, latency={t_cpu_ns/1e6:.6f} ms")
    for k, v in br_cpu.items(): print(f"  {k:<12}: {v/1e6:.6f} uJ ({100*v/e_cpu_pJ:.2f}%)")
    def print_oka(label, e_total, e_read, t_ns, cur):
        print(f"\n[OKA-{label}] read_mode={args.read_mode}, ref_branch={args.include_ref_branch}")
        print(f"  total energy={e_total/1e6:.6f} uJ, latency={t_ns/1e6:.6f} ms")
        print(f"  PGM total={e_pgm_total_pJ/1e6:.6f} uJ, amortized={e_pgm_amort_pJ/1e6:.6f} uJ/frame (reuse={args.n_reuse})")
        print(f"  read={e_read/1e6:.6f} uJ, ADC={e_oka_adc_pJ/1e6:.6f} uJ, SRAM={e_oka_sram_pJ/1e6:.6f} uJ")
        print(f"  current sum={cur:.3f} nA, CPU/OKA energy={e_cpu_pJ/e_total:.3f}x, latency={t_cpu_ns/t_ns:.3f}x")
    print_oka(f"lab_{args.lab_read_ms}ms", e_oka_lab_pJ, e_read_lab_pJ, t_oka_lab_ns, cur_lab)
    print_oka(f"fast_{args.fast_read_us}us", e_oka_fast_pJ, e_read_fast_pJ, t_oka_fast_ns, cur_fast)

    csv_path = os.path.join(args.res_dir, "35_summary.csv")
    with open(csv_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["category", "metric", "value"])
        for k, v in metrics.items(): w.writerow(["quality", k, f"{v:.6f}"])
        w.writerow(["fit", "alpha", f"{pr_model.alpha:.6f}"])
        if pr_model.R_fit_mode == "exp":
            w.writerow(["fit", "R_A", f"{pr_model.A:.10g}"])
            w.writerow(["fit", "R_I0_nA", f"{pr_model.I0:.6f}"])
        for k, v in mapping_info.items(): w.writerow(["mapping", k, v])
        for k, v in br_cpu.items(): w.writerow(["cpu_breakdown_uJ", k, f"{v/1e6:.6f}"])
        w.writerow(["energy", "cpu_uJ", f"{e_cpu_pJ/1e6:.6f}"])
        w.writerow(["energy", "oka_lab_uJ", f"{e_oka_lab_pJ/1e6:.6f}"])
        w.writerow(["energy", "oka_fast_uJ", f"{e_oka_fast_pJ/1e6:.6f}"])
        w.writerow(["energy", "ratio_cpu_oka_lab", f"{e_cpu_pJ/e_oka_lab_pJ:.6f}"])
        w.writerow(["energy", "ratio_cpu_oka_fast", f"{e_cpu_pJ/e_oka_fast_pJ:.6f}"])
        w.writerow(["latency", "cpu_ms", f"{t_cpu_ns/1e6:.6f}"])
        w.writerow(["latency", "oka_lab_ms", f"{t_oka_lab_ns/1e6:.6f}"])
        w.writerow(["latency", "oka_fast_ms", f"{t_oka_fast_ns/1e6:.6f}"])
        w.writerow(["programming", "avg_pulse_per_cell", f"{N_map.mean():.6f}"])
        w.writerow(["programming", "max_pulse", f"{int(N_map.max())}"])
        w.writerow(["programming", "avg_pgm_err_nA", f"{err_map.mean():.6f}"])
        w.writerow(["programming", "max_pgm_err_nA", f"{err_map.max():.6f}"])
    print(f"\n[Saved] {csv_path}")
    print(f"[Saved] {os.path.join(args.res_dir, 'inference_compare.png')}")
    print(f"[Done] {os.path.abspath(args.res_dir)}")


if __name__ == "__main__":
    main()
