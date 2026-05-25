
"""
35i_oka_deltaR_powerlaw_fast_ispp_variation.py
======================================================================
CPU DirectPDK vs OKA DeltaR-PDK simulation with power-law photoresponse,
continuous-ISPP programming, fixed responsivity variation, CDS/read-pulse energy, energy/latency plots, and
CPU-vs-OKA comparison-scheme table.

Core idea
---------
Instead of directly mapping a pre-trained digital kernel weight w to
photoresponsivity R, the OKA-side PDK variable is trained as:

    DeltaR = R - R_ref

under the measured power-law device forward model:

    I_photo = I_light - I_dark = R(I_state) * P^alpha

Thus, the OKA PDK is trained as a responsivity-difference kernel:

    Y_OKA = norm * sum_k DeltaR_k(x,y) * P_k(x,y)^alpha

After training:
    DeltaR -> R_target = R_ref + DeltaR
             -> I_target through inverse R(I_state)
             -> I_actual through continuous ISPP or measured discrete LTD programming states

Timing note
-----------
A 10 ms single read pulse corresponds to 100 fps if the read pulse alone
sets the frame period. In CDS mode, two 10 ms read pulses correspond to
20 ms per frame, i.e., 50 fps, before ADC and optional read gaps.


The default CPU exposure/readout latency is 33.333 ms, corresponding to the
30 fps CMOS image-sensor baseline used in RedEye. Set --cpu-exposure-ms 0
only when comparing pure post-capture digital processing latency.

Energy
------
CDS read energy is computed from actual read-pulse current, not from only
the subtracted photocurrent signal:

    CDS        : E = V_read * t_read * sum(I_dark + I_light)
    light_only : E = V_read * t_read * sum(I_light)

where:
    I_light = I_dark + R(I_state) * P^alpha

Outputs
-------
- inference_compare.png
- energy_latency_compare.png
- comparison_scheme.csv
- comparison_scheme.png
- individual/*.png
- 35i_summary.csv
- cpu_directpdk.pt
- oka_deltaR_model.pt
- oka_programmed_states.pt
"""

import argparse
import csv
import glob
import os
import random
from dataclasses import dataclass

import numpy as np
import openpyxl
import torch
import torch.nn as nn
import torch.nn.functional as F
from PIL import Image
from scipy.optimize import curve_fit

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt


SEED = 42
torch.manual_seed(SEED)
np.random.seed(SEED)
random.seed(SEED)


# =============================================================================
# Photoresponse model
# =============================================================================

@dataclass
class PowerLawPhotoModel:
    alpha: float
    I_states: np.ndarray
    R_states: np.ndarray
    R_fit_mode: str = "exp"
    A: float = None
    I0: float = None

    def __post_init__(self):
        self.I_states = np.asarray(self.I_states, dtype=float)
        self.R_states = np.asarray(self.R_states, dtype=float)
        order = np.argsort(self.I_states)
        self.I_states = self.I_states[order]
        self.R_states = self.R_states[order]
        if self.R_fit_mode == "exp":
            self._fit_exp()

    def _fit_exp(self):
        def model(I, A, I0):
            return A * np.exp(-np.asarray(I) / I0)

        A0 = max(float(np.max(self.R_states)), 1e-15)
        I00 = max(float(np.mean(np.abs(self.I_states))), 1.0)

        try:
            popt, _ = curve_fit(
                model, self.I_states, self.R_states,
                p0=[A0, I00],
                bounds=([1e-20, 1e-9], [1e12, 1e12]),
                maxfev=50000,
            )
            self.A = float(popt[0])
            self.I0 = float(popt[1])
        except Exception:
            y = np.log(np.clip(self.R_states, 1e-20, None))
            x = self.I_states
            m, b = np.polyfit(x, y, 1)
            self.I0 = float(max(-1.0 / m, 1e-9)) if m < 0 else I00
            self.A = float(np.exp(b))

    def R(self, I_state):
        I_state = np.asarray(I_state, dtype=float)
        if self.R_fit_mode == "exp":
            return self.A * np.exp(-I_state / self.I0)
        return np.interp(
            I_state,
            self.I_states,
            self.R_states,
            left=self.R_states[0],
            right=self.R_states[-1],
        )

    def I_photo(self, P, I_state):
        return self.R(I_state) * np.power(np.asarray(P, dtype=float), self.alpha)

    @classmethod
    def fit_from_xlsx(cls, path, R_fit_mode="exp", alpha_bounds=(0.05, 2.0)):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))

        I_states = []
        for s in rows[1][1:]:
            if s is None:
                continue
            if isinstance(s, str):
                I_states.append(float(s.replace("nA", "").strip()))
            else:
                I_states.append(float(s))
        I_states = np.asarray(I_states, dtype=float)

        P_list, Iph_list = [], []
        for r in rows[2:]:
            if r[0] is None:
                continue
            P_list.append(float(r[0]))
            Iph_list.append([float(v) for v in r[1:1 + len(I_states)]])

        P_vals = np.asarray(P_list, dtype=float)
        Iph = np.asarray(Iph_list, dtype=float)
        n_state = len(I_states)

        P_grid, S_grid = np.meshgrid(P_vals, np.arange(n_state), indexing="ij")

        def fit_model(X, *params):
            P_in, S_in = X
            log_Rs = np.asarray(params[:-1])
            alpha = params[-1]
            R_s = np.exp(log_Rs[S_in.astype(int)])
            return R_s * np.power(P_in, alpha)

        R_init = np.mean(Iph / np.clip(P_vals[:, None], 1e-15, None), axis=0)
        p0 = list(np.log(np.clip(R_init, 1e-20, None))) + [1.0]
        lower = [-50.0] * n_state + [alpha_bounds[0]]
        upper = [50.0] * n_state + [alpha_bounds[1]]

        popt, _ = curve_fit(
            fit_model,
            (P_grid.flatten(), S_grid.flatten()),
            Iph.flatten(),
            p0=p0,
            bounds=(lower, upper),
            maxfev=50000,
        )

        R_states = np.exp(np.asarray(popt[:-1]))
        alpha = float(popt[-1])
        pred = fit_model((P_grid.flatten(), S_grid.flatten()), *popt)
        truth = Iph.flatten()
        rmse = float(np.sqrt(np.mean((truth - pred) ** 2)))
        rel = float(np.mean(np.abs((truth - pred) / np.clip(truth, 1e-12, None))) * 100)

        obj = cls(alpha, I_states, R_states, R_fit_mode=R_fit_mode)

        print("[Photoresponse fit]")
        print("  I_photo = R(I_state) * P^alpha")
        print(f"  alpha = {alpha:.6f}")
        print(f"  I_states[nA] = {np.round(I_states, 4)}")
        print(f"  R_states = {np.round(R_states, 10)}")
        print(f"  fit rmse = {rmse:.6f} nA, mean rel err = {rel:.3f}%")
        if R_fit_mode == "exp":
            print(f"  R(I_state) = A exp(-I/I0), A={obj.A:.8g}, I0={obj.I0:.6f} nA")
        else:
            print("  R(I_state) = linear interpolation over extracted R states")

        return obj, (P_vals, I_states, Iph)


# =============================================================================
# LTD / programming model
# =============================================================================

@dataclass
class LTDModel:
    V_levels: np.ndarray
    params: dict
    I_init_avg: float
    I_min: float
    I_max: float
    N_meas_max: int

    @staticmethod
    def _double_exp(N, a1, t1, a2, t2, Iinf):
        return a1 * np.exp(-N / np.abs(t1)) + a2 * np.exp(-N / np.abs(t2)) + Iinf

    @classmethod
    def fit_from_xlsx(cls, path, v_assignment=None):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        data = np.array([list(r) for r in ws.iter_rows(values_only=True)], dtype=float)

        pulse = data[:, 0].astype(int)
        curves_nA = data[:, 1:] * 1e9

        ratio = curves_nA[0] / np.maximum(curves_nA[-1], 1e-12)
        control_idx = int(np.argmin(np.abs(ratio - 1.0)))
        delta = curves_nA[0] - curves_nA[-1]
        order = np.argsort(-delta)
        valid_idx = [i for i in order if i != control_idx]

        if v_assignment is None:
            v_assignment = [10.0, 9.5, 9.0, 8.5, 8.0, 7.5, 7.0, 6.5, 6.0]
        if len(v_assignment) < len(valid_idx):
            raise ValueError("v_assignment is shorter than the detected LTD curves.")

        col_to_V = dict(zip(valid_idx, v_assignment[:len(valid_idx)]))
        params = {}

        for i in valid_idx:
            V = float(col_to_V[i])
            y = curves_nA[:, i]
            try:
                p, _ = curve_fit(
                    cls._double_exp,
                    pulse,
                    y,
                    p0=[max(y[0] * 0.7, 1e-9), 5.0, max(y[0] * 0.3, 1e-9), 50.0, max(y[-1], 1e-9)],
                    bounds=([0, 0.1, 0, 0.1, 0], [1e5, 1e6, 1e5, 1e7, 1e5]),
                    maxfev=50000,
                )
                params[V] = tuple(p)
            except Exception as e:
                print(f"[Warning] LTD fit failed for V={V}: {e}")

        V_levels = np.array(sorted(params.keys()), dtype=float)
        I_init_avg = float(curves_nA[0].mean())
        I_max = I_init_avg
        I_min = float(min(cls._double_exp(120, *params[V]) for V in V_levels))

        print("[LTD fit]")
        print(f"  V levels = {list(V_levels)}")
        print(f"  I range = {I_min:.6f} ~ {I_max:.6f} nA")
        for V in sorted(V_levels, reverse=True):
            p = params[V]
            print(f"  V={V:.2f}: a1={p[0]:.4f}, t1={p[1]:.4f}, a2={p[2]:.4f}, t2={p[3]:.4f}, Iinf={p[4]:.4f}")

        N_meas_max = int(np.max(pulse))
        print(f"  measured max pulse = {N_meas_max}")
        return cls(V_levels, params, I_init_avg, I_min, I_max, N_meas_max)

    def current_after(self, V, user_N):
        if user_N <= 0 or V <= 0:
            return float(self.I_init_avg)
        return float(self._double_exp(user_N + 1, *self.params[float(V)]))

    def best_V_N(self, target_I, N_max=0):
        # N_max <= 0 means no artificial pulse budget:
        # use the full measured pulse range from the LTD dataset.
        if N_max is None or int(N_max) <= 0:
            N_eff = int(self.N_meas_max)
        else:
            N_eff = int(N_max)

        candidates = [(0.0, 0, float(self.I_init_avg), abs(float(self.I_init_avg) - target_I))]
        for V in self.V_levels:
            for N in range(1, N_eff + 1):
                I = self.current_after(float(V), N)
                candidates.append((float(V), int(N), I, abs(I - target_I)))
        candidates.sort(key=lambda x: (x[3], x[1], x[0]))
        return candidates[0]



def build_programming_lut(ltd_model, N_budget):
    """
    Build a small measured-trajectory LUT once:
        target current -> nearest (V, N, achieved current)

    The LUT size is only ~1 + (#V levels * #pulses), e.g. ~1000 entries,
    so vectorized search over all H*W*9 OKA cells is much faster than
    nested Python loops.
    """
    if N_budget is None or int(N_budget) <= 0:
        N_eff = int(ltd_model.N_meas_max)
    else:
        N_eff = int(N_budget)

    V_all = [0.0]
    N_all = [0]
    I_all = [float(ltd_model.I_init_avg)]

    N_arr = np.arange(1, N_eff + 1, dtype=np.int32)
    for V in ltd_model.V_levels:
        V = float(V)
        I_arr = LTDModel._double_exp(N_arr + 1, *ltd_model.params[V]).astype(float)
        V_all.extend([V] * len(N_arr))
        N_all.extend(N_arr.tolist())
        I_all.extend(I_arr.tolist())

    V_all = np.asarray(V_all, dtype=float)
    N_all = np.asarray(N_all, dtype=np.int32)
    I_all = np.asarray(I_all, dtype=float)

    # Sort by achieved current for searchsorted-based nearest lookup.
    # Current is the primary key; N and V make the order stable for near-duplicates.
    order = np.lexsort((V_all, N_all, I_all))
    lut = dict(
        I=I_all[order],
        V=V_all[order],
        N=N_all[order],
        N_eff=N_eff,
    )

    print(f"[Programming LUT] entries = {len(I_all):,}, pulse range = 0 ~ {N_eff}")
    print(f"  achieved-current LUT range = {lut['I'].min():.6f} ~ {lut['I'].max():.6f} nA")
    return lut


def lookup_nearest_programming_vectorized(I_target, ltd_model, N_budget):
    """
    Vectorized nearest lookup for all target currents.

    Returns:
        achieved_map, V_map, N_map, err_map
    """
    lut = build_programming_lut(ltd_model, N_budget)
    flat = np.asarray(I_target, dtype=float).reshape(-1)

    I_lut = lut["I"]
    V_lut = lut["V"]
    N_lut = lut["N"]

    idx_right = np.searchsorted(I_lut, flat, side="left")
    idx_right = np.clip(idx_right, 0, len(I_lut) - 1)
    idx_left = np.clip(idx_right - 1, 0, len(I_lut) - 1)

    err_right = np.abs(I_lut[idx_right] - flat)
    err_left = np.abs(I_lut[idx_left] - flat)

    # Choose the lower-error candidate. For nearly identical errors, prefer
    # fewer pulses, then lower voltage. This roughly matches best_V_N tie logic.
    choose_right = err_right < err_left
    tie = np.isclose(err_right, err_left, rtol=0.0, atol=1e-12)

    choose_right = np.where(tie & (N_lut[idx_right] < N_lut[idx_left]), True, choose_right)
    choose_right = np.where(tie & (N_lut[idx_right] > N_lut[idx_left]), False, choose_right)

    tie2 = tie & (N_lut[idx_right] == N_lut[idx_left])
    choose_right = np.where(tie2 & (V_lut[idx_right] < V_lut[idx_left]), True, choose_right)
    choose_right = np.where(tie2 & (V_lut[idx_right] >= V_lut[idx_left]), False, choose_right)

    idx = np.where(choose_right, idx_right, idx_left)

    achieved = I_lut[idx].reshape(I_target.shape)
    V_map = V_lut[idx].reshape(I_target.shape)
    N_map = N_lut[idx].reshape(I_target.shape)
    err_map = np.abs(achieved - I_target)

    return achieved, V_map, N_map, err_map


def print_programming_summary(title, I_actual, V_map, N_map, err_map, energy_estimate=False):
    H, W, K = I_actual.shape
    n_total = H * W * K
    n_zero = int((N_map == 0).sum())

    print(title)
    print(f"  total cells = {n_total:,}")
    print(f"  avg pulses/cell = {N_map.mean():.6f}, max = {int(N_map.max())}")
    print(f"  no-pulse cells = {n_zero:,} ({100*n_zero/n_total:.3f}%)")
    if energy_estimate:
        print("  V,N are used as an energy-estimation trace; state is verified to target.")
    print(f"  PGM error = avg {err_map.mean():.6f} nA, max {err_map.max():.6f} nA")
    print(f"  I_actual range = {I_actual.min():.6f} ~ {I_actual.max():.6f} nA")

    active = V_map[N_map > 0]
    if active.size > 0:
        for V in sorted(set(active.flatten().tolist())):
            count = int(((V_map == V) & (N_map > 0)).sum())
            if energy_estimate:
                print(f"  V={V:.2f} V used for energy estimate: {count:,} cells ({100*count/n_total:.3f}%)")
            else:
                print(f"  V={V:.2f} V: {count:,} cells ({100*count/n_total:.3f}%)")


def simulate_discrete_programming(I_target, ltd_model, N_budget):
    """
    Conservative lower-bound mode:
    each target is quantized to the nearest measured (V,N) programming state.
    """
    I_actual, V_map, N_map, err_map = lookup_nearest_programming_vectorized(I_target, ltd_model, N_budget)
    print_programming_summary("[Discrete programming - vectorized LUT]",
                              I_actual, V_map, N_map, err_map, energy_estimate=False)
    return I_actual, V_map, N_map, err_map


def estimate_continuous_ispp_programming(I_target, ltd_model, N_budget=0, mode="target_exact"):
    """
    Continuous ISPP / program-and-verify approximation.

    I_actual is set to I_target, because the verify loop is assumed to reach
    the desired state. A vectorized measured-trajectory LUT is still used to
    estimate the programming V,N trace for capacitive PGM energy.
    """
    if mode != "target_exact":
        raise ValueError("Only target_exact is supported for continuous ISPP mode.")

    achieved_for_energy, V_map, N_map, _ = lookup_nearest_programming_vectorized(
        I_target, ltd_model, N_budget
    )

    I_actual = np.asarray(I_target, dtype=float).copy()
    err_map = np.zeros_like(I_actual, dtype=float)

    print_programming_summary("[Continuous ISPP programming - vectorized LUT]",
                              I_actual, V_map, N_map, err_map, energy_estimate=True)

    return I_actual, V_map, N_map, err_map


def apply_responsivity_variation(I_actual, pr_model, range_info,
                                 variation_pct=10.0,
                                 variation_mode="deltaR_gain",
                                 variation_seed=42):
    """
    Apply fixed device-to-device responsivity variation.

    This is a spatial/device mismatch map, not frame-to-frame temporal noise.
    The same H*W*9 multiplier map is used for the whole inference.

    variation_pct is interpreted as coefficient of variation [%].
    A lognormal multiplier is used so the multiplier remains positive and
    has approximately unit mean.
    """
    R_nominal = pr_model.R(I_actual)
    delta_R_nominal = R_nominal - range_info["R_ref"]

    if variation_mode == "none" or variation_pct <= 0:
        mult = np.ones_like(R_nominal, dtype=float)
        R_actual = R_nominal.copy()
        delta_R_actual = delta_R_nominal.copy()
    else:
        cv = float(variation_pct) / 100.0
        sigma_ln = np.sqrt(np.log1p(cv ** 2))
        rng = np.random.default_rng(int(variation_seed))
        mult = np.exp(rng.normal(0.0, sigma_ln, size=R_nominal.shape) - 0.5 * sigma_ln ** 2)

        if variation_mode == "deltaR_gain":
            delta_R_actual = delta_R_nominal * mult
            R_actual = range_info["R_ref"] + delta_R_actual
        elif variation_mode == "absolute_R":
            R_actual = R_nominal * mult
            delta_R_actual = R_actual - range_info["R_ref"]
        else:
            raise ValueError("variation_mode must be none, deltaR_gain, or absolute_R.")

        # Keep the varied responsivity inside the physically allowed model range.
        R_actual = np.clip(R_actual, range_info["R_min"], range_info["R_max"])
        delta_R_actual = R_actual - range_info["R_ref"]

    print("[Responsivity variation]")
    print(f"  mode = {variation_mode}, CV = {variation_pct:.3f}%, seed = {variation_seed}")
    print(f"  multiplier mean/std = {mult.mean():.6f} / {mult.std():.6f}")
    print(f"  R nominal range = {R_nominal.min():.8g} ~ {R_nominal.max():.8g}")
    print(f"  R actual  range = {R_actual.min():.8g} ~ {R_actual.max():.8g}")
    print(f"  DeltaR actual range = {delta_R_actual.min():.8g} ~ {delta_R_actual.max():.8g}")

    stats = dict(
        variation_mode=variation_mode,
        variation_pct=float(variation_pct),
        variation_seed=int(variation_seed),
        multiplier_mean=float(mult.mean()),
        multiplier_std=float(mult.std()),
        R_nominal_min=float(R_nominal.min()),
        R_nominal_max=float(R_nominal.max()),
        R_actual_min=float(R_actual.min()),
        R_actual_max=float(R_actual.max()),
        delta_R_actual_min=float(delta_R_actual.min()),
        delta_R_actual_max=float(delta_R_actual.max()),
    )

    return R_actual, delta_R_actual, mult, stats



# =============================================================================
# Image processing and PDK models
# =============================================================================

def radial_map(H, W, device="cpu"):
    ys = torch.linspace(-1, 1, H, device=device)
    xs = torch.linspace(-1, 1, W, device=device)
    gy, gx = torch.meshgrid(ys, xs, indexing="ij")
    r = (gy**2 + gx**2).sqrt() / (2**0.5)
    return r, gy, gx


def spatially_varying_conv(x, kernels):
    B, C, H, W = x.shape
    patches = F.unfold(x, kernel_size=3, padding=1).view(B, C, 9, H * W)
    k = kernels.view(H * W, 9).T.unsqueeze(0).unsqueeze(0)
    return (patches * k).sum(dim=2).view(B, C, H, W).clamp(0, 1)


def degrade_coma_vig(x, sigma0=0.40, alpha_psf=2.0, coma_k=0.60, alpha_vig=4.0):
    r, gy, gx = radial_map(*x.shape[2:], x.device)
    r2 = r**2
    rs = r.clamp(1e-6)

    sigma = (sigma0 + alpha_psf * r2).clamp(0.1, 2.0)
    shift_y = coma_k * r2 * (gy / rs)
    shift_x = coma_k * r2 * (gx / rs)

    coords = torch.tensor([
        [-1., -1.], [-1., 0.], [-1., 1.],
        [0., -1.], [0., 0.], [0., 1.],
        [1., -1.], [1., 0.], [1., 1.]
    ], device=x.device)

    d2 = (coords**2).sum(-1)
    s = sigma.unsqueeze(-1)

    g1 = torch.exp(-d2 / (2 * s**2))
    g1 = g1 / g1.sum(-1, keepdim=True)

    sy = shift_y.unsqueeze(-1)
    sx = shift_x.unsqueeze(-1)
    dy = coords[:, 0].unsqueeze(0).unsqueeze(0) - sy
    dx = coords[:, 1].unsqueeze(0).unsqueeze(0) - sx
    g2 = torch.exp(-(dy**2 + dx**2) / (2 * s**2))
    g2 = g2 / g2.sum(-1, keepdim=True)

    w = (r * 0.9).clamp(0, 0.9).unsqueeze(-1)
    k = (1 - w) * g1 + w * g2
    k = k / k.sum(-1, keepdim=True)

    x_blur = spatially_varying_conv(x, k)
    V = 1.0 / (1.0 + alpha_vig * r2)
    return (x_blur * V.unsqueeze(0).unsqueeze(0)).clamp(0, 1)


def loss_fn(out, clean):
    l1 = F.l1_loss(out, clean)

    def gmap(t):
        gx = t[:, :, :, 1:] - t[:, :, :, :-1]
        gy = t[:, :, 1:, :] - t[:, :, :-1, :]
        return gx, gy

    ox, oy = gmap(out)
    cx, cy = gmap(clean)
    return l1 + 0.1 * (F.l1_loss(ox, cx) + F.l1_loss(oy, cy))


class DirectPDK(nn.Module):
    """
    CPU ideal DirectPDK:
        Y_CPU = sum_k w_k(x,y) X_k
    """
    def __init__(self, H, W):
        super().__init__()
        init = torch.zeros(H, W, 9)
        init[:, :, 4] = 1.0
        self.kernels = nn.Parameter(init)

    def forward(self, x):
        return spatially_varying_conv(x, self.kernels)


class DeltaROKA_PDK(nn.Module):
    """
    OKA PDK with DeltaR as trainable variable.

    Forward:
        P = P_scale * X
        signal = sum_k DeltaR_k * P_k^alpha
        output = signal * norm_gain

    DeltaR is constrained to the available positive/negative R range by tanh.
    """
    def __init__(self, H, W, alpha, P_scale, delta_R_limit, init_center=0.25):
        super().__init__()
        self.H = H
        self.W = W
        self.alpha = float(alpha)
        self.P_scale = float(P_scale)
        self.delta_R_limit = float(delta_R_limit)

        init_delta = torch.zeros(H, W, 9)
        init_delta[:, :, 4] = init_center * self.delta_R_limit

        # raw_delta is initialized as atanh(init_delta / limit)
        safe = torch.clamp(init_delta / max(self.delta_R_limit, 1e-20), -0.95, 0.95)
        self.raw_delta = nn.Parameter(torch.atanh(safe))

        self.norm_gain = 1.0 / max(self.delta_R_limit * (self.P_scale ** self.alpha), 1e-20)

    def delta_R(self):
        return self.delta_R_limit * torch.tanh(self.raw_delta)

    def forward(self, x):
        B, C, H, W = x.shape
        P_alpha = (x.clamp(0, 1) * self.P_scale).pow(self.alpha)
        patches = F.unfold(P_alpha, kernel_size=3, padding=1).view(B, C, 9, H * W)
        dR = self.delta_R().view(H * W, 9).T.unsqueeze(0).unsqueeze(0)
        signal = (patches * dR).sum(dim=2).view(B, C, H, W)
        return (signal * self.norm_gain).clamp(0, 1)


def train_model(model, train_imgs, degrade_fn, n_iter, lr, batch, label):
    opt = torch.optim.Adam(model.parameters(), lr=lr)
    sched = torch.optim.lr_scheduler.CosineAnnealingLR(opt, T_max=n_iter)
    X = torch.cat(train_imgs, dim=0)
    n = X.shape[0]

    print(f"[Training] {label}")
    for i in range(n_iter):
        idx = torch.randperm(n)[:batch]
        clean = X[idx]
        deg = degrade_fn(clean)

        opt.zero_grad()
        out = model(deg)
        loss = loss_fn(out, clean)
        loss.backward()
        opt.step()
        sched.step()

        if (i + 1) % 200 == 0:
            print(f"  {label}: iter {i+1}/{n_iter}, loss={loss.item():.6f}")


# =============================================================================
# DeltaR -> R -> I_state mapping
# =============================================================================

def choose_R_ref_and_range(pr_model, ltd_model, N_budget, ref_mode="median"):
    # N_budget <= 0 means no artificial max-pulse constraint.
    N_eff = int(ltd_model.N_meas_max) if (N_budget is None or int(N_budget) <= 0) else int(N_budget)

    I_high = float(ltd_model.I_init_avg)
    strongest_V = float(np.max(ltd_model.V_levels))
    I_low = float(ltd_model.current_after(strongest_V, N_eff))

    R_a = float(pr_model.R(I_high))
    R_b = float(pr_model.R(I_low))
    R_min = min(R_a, R_b)
    R_max = max(R_a, R_b)

    if ref_mode == "median":
        R_ref = 0.5 * (R_min + R_max)
    elif ref_mode == "low_bias":
        # OKA-favorable if learned DeltaR tends positive.
        R_ref = R_min + 0.35 * (R_max - R_min)
    else:
        raise ValueError("ref_mode must be median or low_bias.")

    delta_R_limit = min(R_ref - R_min, R_max - R_ref)
    if delta_R_limit <= 0:
        raise ValueError("Invalid R reference / range.")

    info = dict(
        I_low=I_low,
        I_high=I_high,
        R_min=R_min,
        R_max=R_max,
        R_ref=R_ref,
        delta_R_limit=delta_R_limit,
        strongest_V=strongest_V,
        N_budget=N_eff,
        N_budget_requested=N_budget,
        ref_mode=ref_mode,
    )

    print("[OKA DeltaR range]")
    print(f"  I reachable = {min(I_low, I_high):.6f} ~ {max(I_low, I_high):.6f} nA")
    print(f"  R reachable = {R_min:.8g} ~ {R_max:.8g}")
    print(f"  R_ref({ref_mode}) = {R_ref:.8g}")
    print(f"  programming pulse range used = 0 ~ {N_eff} pulses")
    print(f"  DeltaR limit = +/- {delta_R_limit:.8g}")

    return info


def deltaR_to_I_target(delta_R_map, pr_model, range_info):
    R_target = range_info["R_ref"] + delta_R_map
    R_target = np.clip(R_target, range_info["R_min"] + 1e-20, range_info["R_max"] - 1e-20)

    if pr_model.R_fit_mode == "exp":
        I_target = -pr_model.I0 * np.log(R_target / pr_model.A)
    else:
        I_grid = np.linspace(min(range_info["I_low"], range_info["I_high"]),
                             max(range_info["I_low"], range_info["I_high"]), 5000)
        R_grid = pr_model.R(I_grid)
        order = np.argsort(R_grid)
        I_target = np.interp(R_target, R_grid[order], I_grid[order])

    I_target = np.clip(
        I_target,
        min(range_info["I_low"], range_info["I_high"]),
        max(range_info["I_low"], range_info["I_high"]),
    )
    return I_target, R_target


def unfold_patches_np(img_hw):
    H, W = img_hw.shape
    pad = np.pad(img_hw, 1, mode="edge")
    patches = np.zeros((H, W, 9), dtype=float)
    coords = [
        (-1, -1), (-1, 0), (-1, 1),
        (0, -1), (0, 0), (0, 1),
        (1, -1), (1, 0), (1, 1),
    ]
    for k, (dy, dx) in enumerate(coords):
        patches[:, :, k] = pad[1 + dy:1 + dy + H, 1 + dx:1 + dx + W]
    return patches


def oka_forward_programmed(deg_t, I_actual, pr_model, range_info, P_scale, norm_gain,
                            R_actual_map=None):
    deg = deg_t.squeeze().detach().cpu().numpy().astype(float)
    P_alpha = np.power(P_scale * unfold_patches_np(deg), pr_model.alpha)

    if R_actual_map is None:
        R_actual = pr_model.R(I_actual)
    else:
        R_actual = np.asarray(R_actual_map, dtype=float)
    delta_R_actual = R_actual - range_info["R_ref"]

    signal = (delta_R_actual * P_alpha).sum(axis=2)
    out = np.clip(signal * norm_gain, 0, 1)
    return torch.from_numpy(out).float().unsqueeze(0).unsqueeze(0)


# =============================================================================
# Energy and latency
# =============================================================================

ENERGY = dict(
    cpu_mac_pj=100.0,
    dram_read_pj_bit=5.0,
    dram_write_pj_bit=5.0,
    sram_read_pj_bit=0.05,
    sram_write_pj_bit=0.05,
    adc_pj_sample=5.0,
    sensor_pj_pix=0.5,
)

LATENCY = dict(
    cpu_mac_ns=1.0,
    dram_ns=50.0,
    sram_ns=1.0,
    adc_ns=100.0,
    sensor_ns=1000.0,
)


def cpu_pdk_energy_latency(H, W, k=9, n_bit=8, mode="pdk_map_dram"):
    n_pix = H * W
    e_sensor = n_pix * ENERGY["sensor_pj_pix"]
    e_adc = n_pix * ENERGY["adc_pj_sample"]
    e_mac = n_pix * k * ENERGY["cpu_mac_pj"]

    if mode == "pdk_map_dram":
        e_raw_write = n_pix * n_bit * ENERGY["dram_write_pj_bit"]
        e_input_read = n_pix * k * n_bit * ENERGY["dram_read_pj_bit"]
        e_weight_read = n_pix * k * n_bit * ENERGY["dram_read_pj_bit"]
        e_out_write = n_pix * n_bit * ENERGY["dram_write_pj_bit"]
    elif mode == "pdk_map_sram":
        e_raw_write = n_pix * n_bit * ENERGY["dram_write_pj_bit"]
        e_input_read = n_pix * k * n_bit * ENERGY["sram_read_pj_bit"]
        e_weight_read = n_pix * k * n_bit * ENERGY["sram_read_pj_bit"]
        e_out_write = n_pix * n_bit * ENERGY["dram_write_pj_bit"]
    elif mode == "mac_only":
        e_raw_write = 0.0
        e_input_read = 0.0
        e_weight_read = 0.0
        e_out_write = n_pix * n_bit * ENERGY["sram_write_pj_bit"]
    else:
        raise ValueError("unknown CPU mode")

    total = e_sensor + e_adc + e_raw_write + e_input_read + e_weight_read + e_mac + e_out_write

    cores_simd = 40.0
    t_sensor = LATENCY["sensor_ns"]
    t_adc = H * LATENCY["adc_ns"]
    t_mac = n_pix * k * LATENCY["cpu_mac_ns"] / cores_simd

    if mode == "pdk_map_dram":
        t_mem = n_pix * (2 * k + 2) * LATENCY["dram_ns"] / cores_simd
    elif mode == "pdk_map_sram":
        t_mem = n_pix * 2 * LATENCY["dram_ns"] / cores_simd + n_pix * 2 * k * LATENCY["sram_ns"] / cores_simd
    else:
        t_mem = 0.0

    latency = t_sensor + t_adc + t_mac + t_mem
    br = dict(
        sensor=e_sensor, adc=e_adc, raw_write=e_raw_write,
        input_read=e_input_read, weight_read=e_weight_read,
        mac=e_mac, out_write=e_out_write,
    )
    return total, latency, br


def oka_programming_energy_pJ(V_map, N_map, C_gate_fF):
    return float((0.5 * C_gate_fF * (V_map ** 2) * N_map * 1e-3).sum())


def oka_read_energy_pJ(deg_t, I_actual, pr_model, P_scale, V_read, t_read_us,
                       read_mode="cds", include_ref_branch=False, R_ref=None,
                       R_actual_map=None):
    deg = deg_t.squeeze().detach().cpu().numpy().astype(float)
    P_patches = P_scale * unfold_patches_np(deg)

    I_dark = np.asarray(I_actual, dtype=float)
    if R_actual_map is None:
        R_actual = pr_model.R(I_dark)
    else:
        R_actual = np.asarray(R_actual_map, dtype=float)
    I_photo = R_actual * np.power(P_patches, pr_model.alpha)
    I_light = I_dark + I_photo

    if read_mode == "cds":
        current_sum = float((I_dark + I_light).sum())
    elif read_mode == "light_only":
        current_sum = float(I_light.sum())
    else:
        raise ValueError("read_mode must be cds or light_only")

    if include_ref_branch:
        if R_ref is None:
            raise ValueError("R_ref required for reference branch")
        # Approximate reference state current by nearest R on actual range.
        I_grid = np.linspace(float(I_dark.min()), float(I_dark.max()), 5000)
        R_grid = pr_model.R(I_grid)
        I_ref = float(I_grid[np.argmin(np.abs(R_grid - R_ref))])

        I_ref_dark = np.full_like(I_dark, I_ref)
        I_ref_photo = R_ref * np.power(P_patches, pr_model.alpha)
        I_ref_light = I_ref_dark + I_ref_photo

        if read_mode == "cds":
            current_sum += float((I_ref_dark + I_ref_light).sum())
        else:
            current_sum += float(I_ref_light.sum())

    # V * nA * us = 0.001 pJ
    return V_read * current_sum * t_read_us * 1e-3, current_sum


# =============================================================================
# IO, metrics, plotting, scheme table
# =============================================================================

def parse_float_list(s):
    return [float(x.strip()) for x in s.split(",") if x.strip()]


def load_train_images(data_dir, n, size):
    exts = ("*.png", "*.jpg", "*.jpeg", "*.PNG", "*.JPG", "*.JPEG")
    paths = []
    for ext in exts:
        paths += glob.glob(os.path.join(data_dir, "**", ext), recursive=True)
        paths += glob.glob(os.path.join(data_dir, ext))
    paths = sorted(set(paths))
    random.shuffle(paths)

    imgs = []
    for p in paths:
        try:
            img = Image.open(p).convert("L")
            w, h = img.size
            s = min(w, h)
            img = img.crop(((w - s)//2, (h - s)//2, (w + s)//2, (h + s)//2))
            img = img.resize((size, size), Image.BILINEAR)
            arr = np.array(img, dtype=np.float32) / 255.0
            imgs.append(torch.from_numpy(arr).unsqueeze(0).unsqueeze(0))
        except Exception:
            continue
        if len(imgs) == n:
            break
    print(f"[Load] {len(imgs)} train images from {data_dir}")
    return imgs


def find_image(data_dir, name):
    exts = (".png", ".jpg", ".jpeg", ".PNG", ".JPG", ".JPEG")
    for ext in exts:
        p = os.path.join(data_dir, name + ext)
        if os.path.exists(p):
            return p
        hits = glob.glob(os.path.join(data_dir, "**", name + ext), recursive=True)
        if hits:
            return hits[0]
    return None


def load_image(path, size):
    img = Image.open(path).convert("L")
    w, h = img.size
    s = min(w, h)
    img = img.crop(((w - s)//2, (h - s)//2, (w + s)//2, (h + s)//2))
    img = img.resize((size, size), Image.BILINEAR)
    arr = np.array(img, dtype=np.float32) / 255.0
    return torch.from_numpy(arr).unsqueeze(0).unsqueeze(0)


def psnr(a, b):
    mse = F.mse_loss(a.float(), b.float()).item()
    return 99.9 if mse < 1e-12 else 10 * np.log10(1.0 / mse)


def ssim_simple(a, b):
    a = a.float()
    b = b.float()
    mu_a = F.avg_pool2d(a, 11, stride=1, padding=5)
    mu_b = F.avg_pool2d(b, 11, stride=1, padding=5)
    mu_a2 = mu_a ** 2
    mu_b2 = mu_b ** 2
    mu_ab = mu_a * mu_b
    sig_a2 = F.avg_pool2d(a ** 2, 11, stride=1, padding=5) - mu_a2
    sig_b2 = F.avg_pool2d(b ** 2, 11, stride=1, padding=5) - mu_b2
    sig_ab = F.avg_pool2d(a * b, 11, stride=1, padding=5) - mu_ab
    c1, c2 = 0.01**2, 0.03**2
    return (((2*mu_ab + c1) * (2*sig_ab + c2)) /
            ((mu_a2 + mu_b2 + c1) * (sig_a2 + sig_b2 + c2))).mean().item()


def save_img(t, path):
    arr = np.clip(t.squeeze().detach().cpu().numpy(), 0, 1)
    fig, ax = plt.subplots(1, 1, figsize=(arr.shape[1]/100, arr.shape[0]/100), dpi=300)
    ax.imshow(arr, cmap="gray", vmin=0, vmax=1)
    ax.axis("off")
    plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
    fig.savefig(path, dpi=300, bbox_inches="tight", pad_inches=0)
    plt.close(fig)


def save_inference_compare(clean, deg, cpu, oka, metrics, path):
    fig, axes = plt.subplots(1, 4, figsize=(16, 4.2))
    panels = [
        ("Input", clean),
        (f"Degraded\nPSNR={metrics['psnr_deg']:.2f}, SSIM={metrics['ssim_deg']:.3f}", deg),
        (f"CPU DirectPDK\nPSNR={metrics['psnr_cpu']:.2f}, SSIM={metrics['ssim_cpu']:.3f}", cpu),
        (f"OKA DeltaR-PDK\nPSNR={metrics['psnr_oka']:.2f}, SSIM={metrics['ssim_oka']:.3f}", oka),
    ]
    for ax, (title, img) in zip(axes, panels):
        ax.imshow(img.squeeze().detach().cpu().numpy(), cmap="gray", vmin=0, vmax=1)
        ax.set_title(title, fontsize=10, fontweight="bold")
        ax.axis("off")
    plt.tight_layout()
    fig.savefig(path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def save_energy_latency_plot(cpu_energy_uJ, lab_energy_uJ, fast_energy_uJ,
                             cpu_latency_ms, lab_latency_ms, fast_latency_ms,
                             path):
    labels = ["CPU", "OKA-lab", "OKA-fast"]
    energies = [cpu_energy_uJ, lab_energy_uJ, fast_energy_uJ]
    latencies = [cpu_latency_ms, lab_latency_ms, fast_latency_ms]

    fig, axes = plt.subplots(1, 2, figsize=(10, 4.2))

    axes[0].bar(labels, energies)
    axes[0].set_ylabel("Energy per frame (µJ)")
    axes[0].set_yscale("log")
    axes[0].set_title("Energy comparison")
    for i, v in enumerate(energies):
        axes[0].text(i, v, f"{v:.3g}", ha="center", va="bottom", fontsize=8)

    axes[1].bar(labels, latencies)
    axes[1].set_ylabel("Latency per frame (ms)")
    axes[1].set_yscale("log")
    axes[1].set_title("Latency comparison")
    for i, v in enumerate(latencies):
        axes[1].text(i, v, f"{v:.3g}", ha="center", va="bottom", fontsize=8)

    plt.tight_layout()
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)


def make_comparison_scheme_table(args, out_csv, out_png):
    rows = [
        ("Pipeline input", "Digitized degraded image", "Optical input projected to programmed OKA"),
        ("Kernel representation", "Digital PDK weight map w(x,y,k)", "Differential responsivity map ΔR(x,y,k)"),
        ("PDK training variable", "w(x,y,k)", "ΔR(x,y,k) = R(x,y,k) - R_ref"),
        ("Power-law photoresponse", "-", "I_photo = R(I_state) P^alpha"),
        ("Signed kernel operation", "Positive/negative digital weights", "Reference subtraction using R_ref"),
        ("Median/reference value", "-", f"{args.ref_mode} reference"),
        ("Programming step", "-", f"{args.pgm_mode}; continuous_ispp assumes program-and-verify to target state"),
        ("Programming energy", "-", "Σ 1/2 C_gate V_PGM^2 N_pulse, counted per single frame; V,N estimated from measured LTD curves"),
        ("Sensing operation", "Sensor exposure before ADC", "Optoelectronic weighted conversion in device"),
        ("ADC", "Raw image ADC before convolution", "ADC after analog current summation"),
        ("Convolution/MAC", "9 digital MACs per output pixel", "Parallel optical weighting and current summation"),
        ("Kernel memory access", "Read PDK weight map", "Stored locally as programmed device state"),
        ("Input memory access", "Read neighboring pixels for 3×3 window", "-"),
        ("Output write", "Write processed digital image to DRAM", "Write ADC output frame to DRAM"),
        ("Read energy", "-", f"{args.read_mode}: V_read t_read Σ(I_dark + I_light) if CDS"),
        ("CDS handling", "-", "Dark read + light read current included when read_mode=cds"),
        ("CPU exposure/readout latency", f"{args.cpu_exposure_ms} ms included; default follows 30 fps CMOS image-sensor baseline", "-"),
        ("Lab scenario", "-", f"read pulse = {args.lab_read_ms} ms per read; CDS uses two reads"),
        ("Fast scenario", "-", f"read pulse = {args.fast_read_us} µs"),
        ("Main advantage", "-", "Reduced digital MAC, raw-frame movement, and PDK-map reads"),
        ("Variation model", "-", f"{args.variation_mode}, CV={args.variation_pct}% fixed responsivity mismatch"),
        ("Main limitation", "-", "Depends on programming accuracy, variation, and readout speed"),
    ]

    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Item", "CPU-based PDK", "OKA-based PDK"])
        w.writerows(rows)

    # Save as PNG table.
    fig_h = max(6, 0.38 * (len(rows) + 1))
    fig, ax = plt.subplots(figsize=(13, fig_h))
    ax.axis("off")

    table_data = [["Item", "CPU-based PDK", "OKA-based PDK"]] + [list(r) for r in rows]
    tbl = ax.table(cellText=table_data, cellLoc="left", loc="center",
                   colWidths=[0.24, 0.34, 0.42])
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(8)
    tbl.scale(1, 1.35)

    for (r, c), cell in tbl.get_celld().items():
        cell.set_linewidth(0.5)
        if r == 0:
            cell.set_text_props(weight="bold")
        if c == 0 and r > 0:
            cell.set_text_props(weight="bold")

    fig.savefig(out_png, dpi=200, bbox_inches="tight")
    plt.close(fig)

    return rows



# =============================================================================
# Main
# =============================================================================

def parse_args():
    p = argparse.ArgumentParser()

    p.add_argument("--pr-data", default="./expdata/photoresponsivity.xlsx")
    p.add_argument("--ltd-data", default="./expdata/depression_data.xlsx")
    p.add_argument("--train-dir", default="./data/BSD300/images/train")
    p.add_argument("--inf-dir", default="./data/BSD300/images/test")
    p.add_argument("--inf-name", default="102061")
    p.add_argument("--img-size", type=int, default=256)

    p.add_argument("--R-fit-mode", choices=["exp", "interp"], default="exp")
    p.add_argument("--alpha-min", type=float, default=0.05)
    p.add_argument("--alpha-max", type=float, default=2.0)

    p.add_argument("--n-train", type=int, default=200)
    p.add_argument("--cpu-iter", type=int, default=2000)
    p.add_argument("--oka-iter", type=int, default=2000)
    p.add_argument("--lr", type=float, default=0.01)
    p.add_argument("--batch", type=int, default=32)

    p.add_argument("--sigma0", type=float, default=0.40)
    p.add_argument("--alpha-psf", type=float, default=2.0)
    p.add_argument("--coma-k", type=float, default=0.60)
    p.add_argument("--alpha-vig", type=float, default=4.0)

    p.add_argument("--P-scale", type=float, default=1000.0)
    p.add_argument("--N-budget", type=int, default=0,
                   help="Max PGM pulse count. 0 means use the full measured LTD pulse range.")
    p.add_argument("--ref-mode", choices=["median", "low_bias"], default="median")
    p.add_argument("--v-assignment", default="10,9.5,9,8.5,8,7.5,7,6.5,6")

    p.add_argument("--V-read", type=float, default=1.0)
    p.add_argument("--read-mode", choices=["cds", "light_only"], default="cds")
    p.add_argument("--include-ref-branch", action="store_true")
    p.add_argument("--lab-read-ms", type=float, default=10.0,
                   help="Read pulse width per read. If read_mode=cds, two read pulses are used.")
    p.add_argument("--read-gap-ms", type=float, default=0.0,
                   help="Optional time gap between dark and light read pulses in CDS mode.")
    p.add_argument("--fast-read-us", type=float, default=1.0,
                   help="Projected fast read pulse width per read. If read_mode=cds, two read pulses are used.")
    p.add_argument("--fast-read-gap-us", type=float, default=0.0,
                   help="Optional time gap between dark and light fast-read pulses in CDS mode.")

    p.add_argument("--cpu-mode", choices=["pdk_map_dram", "pdk_map_sram", "mac_only"], default="pdk_map_dram")
    p.add_argument("--cpu-exposure-ms", type=float, default=33.333,
                   help="CMOS frame period / exposure-readout latency added to CPU pipeline. Default 33.333 ms follows the 30 fps image-sensor baseline used in RedEye. Use 0 for pure digital processing latency.")
    p.add_argument("--C-gate-fF", type=float, default=345.46)

    p.add_argument("--pgm-mode", choices=["continuous_ispp", "discrete"], default="continuous_ispp",
                   help="continuous_ispp: program-and-verify reaches target state; discrete: measured V,N states only.")

    p.add_argument("--variation-mode", choices=["none", "deltaR_gain", "absolute_R"], default="deltaR_gain",
                   help="Fixed device-to-device responsivity variation model.")
    p.add_argument("--variation-pct", type=float, default=10.0,
                   help="Coefficient of variation for responsivity mismatch in percent. Default 10%.")
    p.add_argument("--variation-seed", type=int, default=42,
                   help="Random seed for the fixed variation map.")

    p.add_argument("--res-dir", default="./res/oka35i_deltaR_powerlaw_fast_ispp_variation")
    return p.parse_args()


def main():
    args = parse_args()
    os.makedirs(args.res_dir, exist_ok=True)
    ind_dir = os.path.join(args.res_dir, "individual")
    os.makedirs(ind_dir, exist_ok=True)

    H = W = args.img_size
    v_assignment = parse_float_list(args.v_assignment)

    print("=" * 72)
    print("[1] Fit experimental PR and LTD models")
    print("=" * 72)
    pr_model, _ = PowerLawPhotoModel.fit_from_xlsx(
        args.pr_data,
        R_fit_mode=args.R_fit_mode,
        alpha_bounds=(args.alpha_min, args.alpha_max),
    )
    ltd_model = LTDModel.fit_from_xlsx(args.ltd_data, v_assignment=v_assignment)

    range_info = choose_R_ref_and_range(pr_model, ltd_model, args.N_budget, ref_mode=args.ref_mode)

    print("\n" + "=" * 72)
    print("[2] Load images and train CPU/OKA PDKs")
    print("=" * 72)

    train_imgs = load_train_images(args.train_dir, args.n_train, H)
    if len(train_imgs) == 0:
        raise RuntimeError(f"No training images found in {args.train_dir}")

    degrade_fn = lambda x: degrade_coma_vig(
        x,
        sigma0=args.sigma0,
        alpha_psf=args.alpha_psf,
        coma_k=args.coma_k,
        alpha_vig=args.alpha_vig,
    )

    cpu_model = DirectPDK(H, W)
    train_model(cpu_model, train_imgs, degrade_fn, args.cpu_iter, args.lr, args.batch, "CPU DirectPDK")
    torch.save({"kernels": cpu_model.kernels.detach()}, os.path.join(args.res_dir, "cpu_directpdk.pt"))

    oka_model = DeltaROKA_PDK(
        H, W,
        alpha=pr_model.alpha,
        P_scale=args.P_scale,
        delta_R_limit=range_info["delta_R_limit"],
    )
    train_model(oka_model, train_imgs, degrade_fn, args.oka_iter, args.lr, args.batch, "OKA DeltaR-PDK")
    torch.save(
        {
            "raw_delta": oka_model.raw_delta.detach(),
            "delta_R": oka_model.delta_R().detach(),
            "range_info": range_info,
            "alpha": pr_model.alpha,
            "P_scale": args.P_scale,
        },
        os.path.join(args.res_dir, "oka_deltaR_model.pt")
    )

    print("\n" + "=" * 72)
    print("[3] Map trained DeltaR to programmed states and apply fixed variation")
    print("=" * 72)

    delta_R_map = oka_model.delta_R().detach().cpu().numpy()
    I_target, R_target = deltaR_to_I_target(delta_R_map, pr_model, range_info)
    if args.pgm_mode == "continuous_ispp":
        I_actual, V_map, N_map, err_map = estimate_continuous_ispp_programming(
            I_target, ltd_model, N_budget=args.N_budget, mode="target_exact"
        )
    else:
        I_actual, V_map, N_map, err_map = simulate_discrete_programming(
            I_target, ltd_model, args.N_budget
        )

    R_actual_map, delta_R_actual_map, variation_multiplier, variation_stats = apply_responsivity_variation(
        I_actual,
        pr_model,
        range_info,
        variation_pct=args.variation_pct,
        variation_mode=args.variation_mode,
        variation_seed=args.variation_seed,
    )

    torch.save(
        {
            "delta_R_target": torch.from_numpy(delta_R_map).float(),
            "R_target": torch.from_numpy(R_target).float(),
            "I_target": torch.from_numpy(I_target).float(),
            "I_actual": torch.from_numpy(I_actual).float(),
            "R_actual_map": torch.from_numpy(R_actual_map).float(),
            "delta_R_actual_map": torch.from_numpy(delta_R_actual_map).float(),
            "variation_multiplier": torch.from_numpy(variation_multiplier).float(),
            "variation_stats": variation_stats,
            "V_map": V_map,
            "N_map": N_map,
            "err_map": err_map,
            "range_info": range_info,
        },
        os.path.join(args.res_dir, "oka_programmed_states.pt")
    )

    print("\n" + "=" * 72)
    print("[4] Inference")
    print("=" * 72)

    src = find_image(args.inf_dir, args.inf_name)
    if src is None:
        raise RuntimeError(f"Could not find inference image {args.inf_name} in {args.inf_dir}")

    clean = load_image(src, H)
    with torch.no_grad():
        deg = degrade_fn(clean)
        out_cpu = cpu_model(deg).clamp(0, 1)
        out_oka_ideal = oka_model(deg).clamp(0, 1)
        out_oka = oka_forward_programmed(
            deg,
            I_actual,
            pr_model,
            range_info,
            P_scale=args.P_scale,
            norm_gain=oka_model.norm_gain,
            R_actual_map=R_actual_map,
        ).clamp(0, 1)

    metrics = dict(
        psnr_deg=psnr(clean, deg),
        ssim_deg=ssim_simple(clean, deg),
        psnr_cpu=psnr(clean, out_cpu),
        ssim_cpu=ssim_simple(clean, out_cpu),
        psnr_oka_ideal=psnr(clean, out_oka_ideal),
        ssim_oka_ideal=ssim_simple(clean, out_oka_ideal),
        psnr_oka=psnr(clean, out_oka),
        ssim_oka=ssim_simple(clean, out_oka),
        psnr_cpu_oka=psnr(out_cpu, out_oka),
        ssim_cpu_oka=ssim_simple(out_cpu, out_oka),
    )

    for k, v in metrics.items():
        print(f"  {k}: {v:.6f}")

    save_inference_compare(clean, deg, out_cpu, out_oka, metrics,
                           os.path.join(args.res_dir, "inference_compare.png"))
    save_img(clean, os.path.join(ind_dir, f"{args.inf_name}_input.png"))
    save_img(deg, os.path.join(ind_dir, f"{args.inf_name}_degraded.png"))
    save_img(out_cpu, os.path.join(ind_dir, f"{args.inf_name}_cpu.png"))
    save_img(out_oka, os.path.join(ind_dir, f"{args.inf_name}_oka_programmed.png"))
    save_img(out_oka_ideal, os.path.join(ind_dir, f"{args.inf_name}_oka_ideal_deltaR.png"))

    print("\n" + "=" * 72)
    print("[5] Energy and latency")
    print("=" * 72)

    e_cpu_pJ, t_cpu_ns, br_cpu = cpu_pdk_energy_latency(H, W, mode=args.cpu_mode)
    t_cpu_ns += args.cpu_exposure_ms * 1e6

    e_pgm_total_pJ = oka_programming_energy_pJ(V_map, N_map, args.C_gate_fF)
    # Single-frame comparison: programming energy is counted fully per frame.
    e_pgm_frame_pJ = e_pgm_total_pJ

    e_read_lab_pJ, i_sum_lab = oka_read_energy_pJ(
        deg, I_actual, pr_model, args.P_scale,
        V_read=args.V_read,
        t_read_us=args.lab_read_ms * 1000.0,
        read_mode=args.read_mode,
        include_ref_branch=args.include_ref_branch,
        R_ref=range_info["R_ref"],
        R_actual_map=R_actual_map,
    )
    e_read_fast_pJ, i_sum_fast = oka_read_energy_pJ(
        deg, I_actual, pr_model, args.P_scale,
        V_read=args.V_read,
        t_read_us=args.fast_read_us,
        read_mode=args.read_mode,
        include_ref_branch=args.include_ref_branch,
        R_ref=range_info["R_ref"],
        R_actual_map=R_actual_map,
    )

    n_pix = H * W
    e_oka_adc_pJ = n_pix * ENERGY["adc_pj_sample"]
    # Fairer output-storage assumption:
    # OKA also writes the final ADC output frame to DRAM.
    # OKA still avoids raw-frame storage, 3x3 input-window reads,
    # PDK-map reads, and digital MAC operations.
    e_oka_out_write_pJ = n_pix * 8 * ENERGY["dram_write_pj_bit"]

    e_oka_lab_pJ = e_pgm_frame_pJ + e_read_lab_pJ + e_oka_adc_pJ + e_oka_out_write_pJ
    e_oka_fast_pJ = e_pgm_frame_pJ + e_read_fast_pJ + e_oka_adc_pJ + e_oka_out_write_pJ

    # Latency uses the actual number of read pulses.
    # light_only: one read pulse.
    # cds: dark read + light read, plus an optional gap.
    n_read_pulses = 2 if args.read_mode == "cds" else 1
    lab_read_total_ms = n_read_pulses * args.lab_read_ms
    if n_read_pulses > 1:
        lab_read_total_ms += (n_read_pulses - 1) * args.read_gap_ms

    fast_read_total_us = n_read_pulses * args.fast_read_us
    if n_read_pulses > 1:
        fast_read_total_us += (n_read_pulses - 1) * args.fast_read_gap_us

    t_oka_lab_ns = lab_read_total_ms * 1e6 + H * LATENCY["adc_ns"]
    t_oka_fast_ns = fast_read_total_us * 1000.0 + H * LATENCY["adc_ns"]

    print(f"[CPU] mode={args.cpu_mode}")
    print(f"  energy = {e_cpu_pJ/1e6:.6f} uJ")
    print(f"  latency = {t_cpu_ns/1e6:.6f} ms (including CPU exposure/readout {args.cpu_exposure_ms:.3f} ms)")
    for k, v in br_cpu.items():
        print(f"    {k:<12}: {v/1e6:.6f} uJ ({100*v/e_cpu_pJ:.2f}%)")

    print(f"\n[OKA lab]")
    print(f"  total energy = {e_oka_lab_pJ/1e6:.6f} uJ")
    print(f"  read energy = {e_read_lab_pJ/1e6:.6f} uJ, current sum = {i_sum_lab:.3f} nA")
    print(f"  PGM energy = {e_pgm_total_pJ/1e6:.6f} uJ/frame (single-frame, no reuse amortization)")
    print(f"  output DRAM write = {e_oka_out_write_pJ/1e6:.6f} uJ/frame")
    print(f"  read latency part = {lab_read_total_ms:.6f} ms ({1000.0/lab_read_total_ms:.2f} fps before ADC, if readout-limited)")
    print(f"  total latency = {t_oka_lab_ns/1e6:.6f} ms")
    print(f"  CPU/OKA-lab energy ratio = {e_cpu_pJ/e_oka_lab_pJ:.6f}x")
    print(f"  CPU/OKA-lab latency ratio = {t_cpu_ns/t_oka_lab_ns:.6f}x")

    print(f"\n[OKA fast]")
    print(f"  total energy = {e_oka_fast_pJ/1e6:.6f} uJ")
    print(f"  read energy = {e_read_fast_pJ/1e6:.6f} uJ, current sum = {i_sum_fast:.3f} nA")
    print(f"  PGM energy = {e_pgm_total_pJ/1e6:.6f} uJ/frame (single-frame, no reuse amortization)")
    print(f"  output DRAM write = {e_oka_out_write_pJ/1e6:.6f} uJ/frame")
    print(f"  read latency part = {fast_read_total_us:.6f} us ({1e6/fast_read_total_us:.2f} fps before ADC, if readout-limited)")
    print(f"  total latency = {t_oka_fast_ns/1e6:.6f} ms")
    print(f"  CPU/OKA-fast energy ratio = {e_cpu_pJ/e_oka_fast_pJ:.6f}x")
    print(f"  CPU/OKA-fast latency ratio = {t_cpu_ns/t_oka_fast_ns:.6f}x")

    save_energy_latency_plot(
        e_cpu_pJ/1e6, e_oka_lab_pJ/1e6, e_oka_fast_pJ/1e6,
        t_cpu_ns/1e6, t_oka_lab_ns/1e6, t_oka_fast_ns/1e6,
        os.path.join(args.res_dir, "energy_latency_compare.png")
    )

    scheme_rows = make_comparison_scheme_table(
        args,
        os.path.join(args.res_dir, "comparison_scheme.csv"),
        os.path.join(args.res_dir, "comparison_scheme.png"),
    )


    summary_path = os.path.join(args.res_dir, "35i_summary.csv")
    with open(summary_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["category", "metric", "value"])
        for k, v in metrics.items():
            w.writerow(["quality", k, f"{v:.8f}"])
        w.writerow(["fit", "alpha", f"{pr_model.alpha:.8f}"])
        if pr_model.R_fit_mode == "exp":
            w.writerow(["fit", "R_A", f"{pr_model.A:.12g}"])
            w.writerow(["fit", "R_I0_nA", f"{pr_model.I0:.8f}"])
        for k, v in range_info.items():
            w.writerow(["mapping", k, v])
        w.writerow(["energy", "cpu_mode", args.cpu_mode])
        w.writerow(["energy", "cpu_uJ", f"{e_cpu_pJ/1e6:.8f}"])
        w.writerow(["energy", "oka_lab_uJ", f"{e_oka_lab_pJ/1e6:.8f}"])
        w.writerow(["energy", "oka_fast_uJ", f"{e_oka_fast_pJ/1e6:.8f}"])
        w.writerow(["energy", "oka_output_dram_write_uJ", f"{e_oka_out_write_pJ/1e6:.8f}"])
        w.writerow(["programming", "pgm_mode", args.pgm_mode])
        for vk, vv in variation_stats.items():
            w.writerow(["variation", vk, vv])
        w.writerow(["energy", "ratio_cpu_oka_lab", f"{e_cpu_pJ/e_oka_lab_pJ:.8f}"])
        w.writerow(["energy", "ratio_cpu_oka_fast", f"{e_cpu_pJ/e_oka_fast_pJ:.8f}"])
        w.writerow(["latency", "cpu_ms", f"{t_cpu_ns/1e6:.8f}"])
        w.writerow(["latency", "cpu_exposure_ms", f"{args.cpu_exposure_ms:.8f}"])
        w.writerow(["latency", "oka_lab_ms", f"{t_oka_lab_ns/1e6:.8f}"])
        w.writerow(["latency", "oka_fast_ms", f"{t_oka_fast_ns/1e6:.8f}"])
        w.writerow(["latency", "read_mode", args.read_mode])
        w.writerow(["latency", "n_read_pulses", f"{n_read_pulses}"])
        w.writerow(["latency", "lab_read_total_ms_before_adc", f"{lab_read_total_ms:.8f}"])
        w.writerow(["latency", "lab_read_fps_before_adc", f"{1000.0/lab_read_total_ms:.8f}"])
        w.writerow(["latency", "fast_read_total_us_before_adc", f"{fast_read_total_us:.8f}"])
        w.writerow(["latency", "fast_read_fps_before_adc", f"{1e6/fast_read_total_us:.8f}"])
        w.writerow(["programming", "avg_pulse_per_cell", f"{N_map.mean():.8f}"])
        w.writerow(["programming", "max_pulse", f"{int(N_map.max())}"])
        w.writerow(["programming", "avg_pgm_err_nA", f"{err_map.mean():.8f}"])
        w.writerow(["programming", "max_pgm_err_nA", f"{err_map.max():.8f}"])

    print(f"\n[Saved] {summary_path}")
    print(f"[Saved] {os.path.join(args.res_dir, 'inference_compare.png')}")
    print(f"[Saved] {os.path.join(args.res_dir, 'energy_latency_compare.png')}")
    print(f"[Saved] {os.path.join(args.res_dir, 'comparison_scheme.csv')}")
    print(f"[Saved] {os.path.join(args.res_dir, 'comparison_scheme.png')}")
    print(f"[Done] {os.path.abspath(args.res_dir)}")


if __name__ == "__main__":
    main()
