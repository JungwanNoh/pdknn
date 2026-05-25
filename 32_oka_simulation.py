"""
32_oka_simulation.py
=========================================================
OKA (One-shot Kernel Array) device-aware simulation.

Pipeline (option A: pre-trained PDK weight를 그대로 OKA로 inference):
  1. 27번에서 학습된 DirectPDK weight 로드 (res/direct_pdk/direct_pdk.pt)
  2. Weight → I_state mapping (bias shift로 음수 weight 표현)
  3. I_state → (V_pulse, N_pulse) 역산 (LTD model 사용)
  4. Device-aware forward model로 inference
       I_photo(P, I_state) = α·exp(-I/I₀)·P^(b₀+k·I+1)
       output(x,y) = Σ I_photo(P_neighbor, I_state) - bias·ΣP
  5. CPU 기반 ideal conv와 PSNR/SSIM 비교
  6. Programming + inference energy 계산

Required input data:
  ./Photoresponsivity.xlsx        (PR fitting용)
  ./expdata/depression_data.xlsx       (LTD fitting용)
  res/direct_pdk/direct_pdk.pt    (학습된 PDK weight)

Run:
  python 32_oka_simulation.py
  python 32_oka_simulation.py \\
      --pdk-weight ./res/direct_pdk/direct_pdk.pt \\
      --pr-data ./expdata/photoresponsivity.xlsx \\
      --ltd-data ./expdata/depression_data.xlsx
"""

import argparse, os, glob, random
import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
from scipy.optimize import curve_fit, brentq
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from PIL import Image

SEED = 42
torch.manual_seed(SEED); np.random.seed(SEED); random.seed(SEED)


# ============================================================================
# 1. Device model fitting from measurement data
# ============================================================================

class PhotoresponsivityModel:
    """
    R(P, I_base) = α·exp(-I/I₀) · P^(b₀ + k·I)
    P:       light power (μW/cm²)
    I_base:  PGM state read current (nA)
    R:       photoresponsivity (단위는 측정과 동일)
    """
    def __init__(self, alpha, I0, b0, k, R_max_meas):
        self.alpha = alpha
        self.I0    = I0
        self.b0    = b0
        self.k     = k
        self.R_max_meas = R_max_meas

    def R(self, P, I):
        return self.alpha * np.exp(-I/self.I0) * P**(self.b0 + self.k*I)

    def I_photo(self, P, I_state):
        """단일 cell 광전류"""
        return self.R(P, I_state) * P

    @classmethod
    def fit_from_xlsx(cls, path):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        # row 0: header (Light power, Basecurrent (state), None x4)
        # row 1: ['', '159.6nA', '119.8nA', ...]
        # row 2~: P, R(I_state1), R(I_state2), ...
        I_base_row = rows[1][1:]
        I_base = np.array([
            float(s.replace('nA','').strip())
            for s in I_base_row if s is not None
        ])
        P_list, R_list = [], []
        for r in rows[2:]:
            if r[0] is None: continue
            P_list.append(float(r[0]))
            R_list.append([float(v) for v in r[1:1+len(I_base)]])
        P = np.array(P_list)
        R = np.array(R_list)

        PP, II = np.meshgrid(P, I_base, indexing='ij')
        def model(X, alpha, I0, b0, k):
            Pin, Iin = X
            return alpha*np.exp(-Iin/I0) * Pin**(b0 + k*Iin)
        p, _ = curve_fit(model, (PP.flatten(), II.flatten()), R.flatten(),
                         p0=[0.5, 50, 0.7, 0.002], maxfev=20000)
        rmse = np.sqrt(np.mean((R.flatten() -
                                 model((PP.flatten(), II.flatten()), *p))**2))
        rel  = np.mean(np.abs((R.flatten() -
            model((PP.flatten(), II.flatten()), *p)) /
            R.flatten())) * 100
        print(f"[PR fit] α={p[0]:.4f}, I₀={p[1]:.2f} nA, "
              f"b₀={p[2]:.4f}, k={p[3]:.5f}")
        print(f"         rmse={rmse:.3f}, mean rel err={rel:.2f}%")
        return cls(p[0], p[1], p[2], p[3], R.max()), (P, I_base, R)


class LTDModel:
    """
    각 V_pulse별 I(N) = a1·exp(-N/τ1) + a2·exp(-N/τ2) + I_inf
    Voltage 매핑: 변화량 큰 순 → 큰 V (8→7→6→5→4 V)
    Control(거의 변화 없음) column 자동 검출
    """
    def __init__(self, V_levels, params, I_init_avg, I_min, I_max):
        self.V_levels = V_levels    # array
        self.params = params         # dict V → (a1,t1,a2,t2,Iinf)
        self.I_init_avg = I_init_avg
        self.I_min = I_min
        self.I_max = I_max

    @staticmethod
    def _double_exp(N, a1, t1, a2, t2, Iinf):
        return a1*np.exp(-N/np.abs(t1)) + a2*np.exp(-N/np.abs(t2)) + Iinf

    def I_state(self, V, N):
        """(V, N) → I_state (nA). V_levels 중 가장 가까운 것 사용"""
        if V not in self.params:
            V = self.V_levels[np.argmin(np.abs(self.V_levels - V))]
        return self._double_exp(N, *self.params[V])

    def find_pulse(self, V, target_I):
        """target I_state 만들기 위한 N (V 고정)"""
        f = lambda n: self._double_exp(n, *self.params[V]) - target_I
        n_max = 1000
        if f(1) * f(n_max) > 0:   # 같은 부호면 reachable 안 함
            if abs(f(1)) < abs(f(n_max)):
                return 1
            return n_max
        try:
            n = brentq(f, 1, n_max)
            return max(1, int(round(n)))
        except:
            return 120

    def best_V_N(self, target_I):
        """target I_state에 가장 효율적으로 도달하는 (V, N)"""
        candidates = []
        for V in self.V_levels:
            n = self.find_pulse(V, target_I)
            achieved = self._double_exp(n, *self.params[V])
            err = abs(achieved - target_I)
            candidates.append((V, n, achieved, err))
        # 가장 정확한 것 선택, 동률이면 N 작은 것
        candidates.sort(key=lambda x: (x[3], x[1]))
        return candidates[0]

    @classmethod
    def fit_from_xlsx(cls, path):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        data = np.array([list(r) for r in ws.iter_rows(values_only=True)],
                         dtype=float)
        pulse = data[:, 0].astype(int)
        curves_nA = data[:, 1:] * 1e9    # A → nA

        # control 자동 검출
        ratio = curves_nA[0] / np.maximum(curves_nA[-1], 1e-6)
        control_idx = int(np.argmin(np.abs(ratio - 1.0)))
        delta = curves_nA[0] - curves_nA[-1]
        order = np.argsort(-delta)
        valid_idx = [i for i in order if i != control_idx]
        # 큰 변화 → 큰 V
        V_assignment = [8.0, 7.0, 6.0, 5.0, 4.0][:len(valid_idx)]
        col_to_V = dict(zip(valid_idx, V_assignment))

        params = {}
        for i in valid_idx:
            V = col_to_V[i]
            y = curves_nA[:, i]
            try:
                p, _ = curve_fit(cls._double_exp, pulse, y,
                    p0=[y[0]*0.7, 5, y[0]*0.3, 50, y[-1]],
                    bounds=([0, 0.1, 0, 0.1, 0],
                            [1e3, 1000, 1e3, 1e5, 300]),
                    maxfev=50000)
                params[V] = tuple(p)
            except Exception as e:
                print(f"  [LTD fit] V={V}V failed: {e}")

        V_levels = np.array(sorted(params.keys()))
        I_init_avg = curves_nA[0].mean()
        # reachable I range
        I_max = I_init_avg
        I_min = min(cls._double_exp(120, *params[v])
                    for v in V_levels)
        print(f"[LTD fit] V levels: {sorted(V_levels)}")
        print(f"          I range: {I_min:.2f} ~ {I_max:.1f} nA")
        for V in sorted(V_levels, reverse=True):
            p = params[V]
            print(f"          V={V}V: a1={p[0]:.2f}, τ1={p[1]:.2f}, "
                  f"a2={p[2]:.2f}, τ2={p[3]:.2f}, Iinf={p[4]:.3f}")
        return cls(V_levels, params, I_init_avg, I_min, I_max)


# ============================================================================
# 2. Weight ↔ I_state mapping
# ============================================================================

def map_weights_to_Istate(weights, pr_model, P_ref,
                          I_min=0.3, I_max=250.0,
                          margin=0.05):
    """
    PDK kernel weight 분포 → I_state 매핑 (bias shift로 음수 표현).

    weights: (H, W, 9) torch tensor
    P_ref:   reference light intensity (P_avg)
             이 P에서의 R(P, I) 분포에 weight를 매핑

    return: (I_state map, info dict)
    """
    w = weights.detach().cpu().numpy()
    w_min, w_max = w.min(), w.max()
    print(f"\n[Mapping] weight range: [{w_min:.4f}, {w_max:.4f}]")

    # I_state range 안에서 R 분포
    I_grid = np.linspace(I_min, I_max, 200)
    R_grid = pr_model.R(P_ref, I_grid)
    R_min, R_max = R_grid.min(), R_grid.max()
    print(f"          R range at P={P_ref:.0f}: [{R_min:.3f}, {R_max:.3f}]")

    # bias-shift mapping:
    #   weight + offset → R_target (linear in [R_min, R_max])
    #   여기서 offset은 weight를 [0, w_max-w_min]으로 만드는 값
    #   R_mid는 effective_w=0인 지점 = weight=0 위치

    w_range = (w_max - w_min) * (1 + 2*margin)
    w_center = (w_max + w_min) / 2
    # weight=0 이 R_mid에 대응
    R_mid = (R_min + R_max) / 2

    # weight → R_target (linear)
    R_target = R_mid + (w / w_range) * (R_max - R_min)
    R_target = np.clip(R_target, R_min, R_max)

    # R_target → I_state (역함수, scalar bisection per pixel은 비효율)
    # → I_grid에서 nearest neighbor lookup
    def R_to_I(R_val):
        idx = np.abs(R_grid[:, None, None, None] - R_val[None]).argmin(axis=0)
        return I_grid[idx]
    I_state_map = R_to_I(R_target)

    info = dict(
        w_min=w_min, w_max=w_max, w_range=w_range,
        R_min=R_min, R_max=R_max, R_mid=R_mid,
        P_ref=P_ref, I_min=I_min, I_max=I_max,
    )
    return I_state_map, info


# ============================================================================
# 3. Device-aware forward model
# ============================================================================

def oka_forward(P_image_nA, I_state_map, pr_model, info):
    """
    Device-aware OKA forward.

    Args:
      P_image_nA: (B, 1, H, W) input light intensity (μW/cm² 단위)
      I_state_map: (H, W, 9) numpy I_state values (nA)
      pr_model: PhotoresponsivityModel
      info: dict from map_weights_to_Istate (R_mid 등)

    Returns:
      output: (B, 1, H, W) tensor
    """
    B, C, H, W = P_image_nA.shape
    P_np = P_image_nA.detach().cpu().numpy()
    I_st = I_state_map   # (H, W, 9)

    # neighbor patches: (B, C, 9, H, W)
    P_t = torch.from_numpy(P_np).float()
    patches = F.unfold(P_t, kernel_size=3, padding=1).view(B, C, 9, H*W)
    patches = patches.numpy().reshape(B, C, 9, H, W)
    # I_state for each tap at each output position
    I_st_taps = I_st.transpose(2, 0, 1)   # (9, H, W)

    # I_photo per cell: α·exp(-I/I₀)·P^(b₀+k·I+1)
    α = pr_model.alpha
    I0 = pr_model.I0
    b0 = pr_model.b0
    k  = pr_model.k

    # Broadcast (1, 1, 9, H, W)
    I_b = I_st_taps[None, None]
    exp_term = α * np.exp(-I_b / I0)
    b_eff = b0 + k * I_b + 1.0
    P_safe = np.maximum(patches, 1e-6)
    I_photo = exp_term * P_safe**b_eff

    # bias subtraction: ΣP × R_mid (R(P_ref, I_state where w=0))
    # 여기선 R_mid·P_neighbor 빼주는 게 정확 (bias shift 이전 정의 따라)
    P_ref = info['P_ref']
    R_mid = info['R_mid']
    bias_per_tap = R_mid * patches   # ideal bias term
    output = (I_photo - bias_per_tap).sum(axis=2)   # (B, C, H, W)

    return torch.from_numpy(output).float()


# ============================================================================
# 4. Ideal CPU conv reference
# ============================================================================

def ideal_conv(P_image, weights):
    """이상적 spatially varying conv (CPU 결과 대조용)"""
    B, C, H, W = P_image.shape
    patches = F.unfold(P_image, kernel_size=3, padding=1).view(B, C, 9, H*W)
    k = weights.view(H*W, 9).T.unsqueeze(0).unsqueeze(0)
    return (patches * k).sum(dim=2).view(B, C, H, W)


# ============================================================================
# 4.5  PDK training (옵션 A: 32번 안에서 fresh 재학습)
# ============================================================================

def radial_map(H, W, device='cpu'):
    ys = torch.linspace(-1, 1, H, device=device)
    xs = torch.linspace(-1, 1, W, device=device)
    gy, gx = torch.meshgrid(ys, xs, indexing='ij')
    r = (gy**2 + gx**2).sqrt() / (2**0.5)
    return r, gy, gx


def spatially_varying_conv(x, kernels):
    B, C, H, W = x.shape
    patches = F.unfold(x, kernel_size=3, padding=1).view(B, C, 9, H*W)
    k = kernels.view(H*W, 9).T.unsqueeze(0).unsqueeze(0)
    return (patches * k).sum(dim=2).view(B, C, H, W).clamp(0, 1)


def degrade_coma_vig(x, sigma0=0.40, alpha_psf=2.0,
                     coma_k=0.60, alpha_vig=4.0):
    """27번과 동일한 degradation"""
    r, gy, gx = radial_map(*x.shape[2:], x.device)
    r2 = r**2; rs = r.clamp(1e-6)
    sigma = (sigma0 + alpha_psf*r2).clamp(0.1, 2.0)
    shift_y = coma_k*r2*(gy/rs)
    shift_x = coma_k*r2*(gx/rs)
    coords = torch.tensor([
        [-1.,-1.],[-1.,0.],[-1.,1.],
        [ 0.,-1.],[ 0.,0.],[ 0.,1.],
        [ 1.,-1.],[ 1.,0.],[ 1.,1.]], device=x.device)
    d2 = (coords**2).sum(-1)
    s = sigma.unsqueeze(-1)
    g1 = torch.exp(-d2/(2*s**2)); g1 = g1/g1.sum(-1, keepdim=True)
    sy = shift_y.unsqueeze(-1); sx = shift_x.unsqueeze(-1)
    dy = coords[:,0].unsqueeze(0).unsqueeze(0) - sy
    dx = coords[:,1].unsqueeze(0).unsqueeze(0) - sx
    g2 = torch.exp(-(dy**2+dx**2)/(2*s**2)); g2 = g2/g2.sum(-1, keepdim=True)
    w = (r*0.9).clamp(0, 0.9).unsqueeze(-1)
    k = (1-w)*g1 + w*g2; k = k/k.sum(-1, keepdim=True)
    x = spatially_varying_conv(x, k)
    V = 1.0/(1.0+alpha_vig*r2)
    return (x*V.unsqueeze(0).unsqueeze(0)).clamp(0, 1)


class DirectPDK(nn.Module):
    def __init__(self, H, W):
        super().__init__()
        init = torch.zeros(H, W, 9)
        init[:, :, 4] = 1.0     # delta init
        self.kernels = nn.Parameter(init)

    def forward(self, x):
        return spatially_varying_conv(x, self.kernels)


def loss_fn(out, clean):
    l1 = F.l1_loss(out, clean)
    def gmap(t): return t[:,:,:,1:]-t[:,:,:,:-1], t[:,:,1:,:]-t[:,:,:-1,:]
    ox, oy = gmap(out); cx, cy = gmap(clean)
    return l1 + 0.1*(F.l1_loss(ox, cx) + F.l1_loss(oy, cy))


def load_train_images(data_dir, n, size):
    exts = ('*.png','*.jpg','*.jpeg','*.PNG','*.JPG','*.JPEG')
    paths = []
    for ext in exts:
        paths += glob.glob(os.path.join(data_dir, '**', ext), recursive=True)
        paths += glob.glob(os.path.join(data_dir, ext))
    paths = sorted(set(paths)); random.shuffle(paths)
    imgs = []
    for p in paths:
        try:
            img = Image.open(p).convert('L')
            w, h = img.size; s = min(w, h)
            img = img.crop(((w-s)//2, (h-s)//2, (w+s)//2, (h+s)//2))
            img = img.resize((size, size), Image.BILINEAR)
            t = torch.from_numpy(np.array(img, dtype=np.float32)/255.
                                 ).unsqueeze(0).unsqueeze(0)
            imgs.append(t)
        except: continue
        if len(imgs) == n: break
    print(f"  [Load] {len(imgs)} train images from {data_dir}")
    return imgs


def train_pdk(model, train_imgs, degrade_fn, n_iter, lr, batch=32):
    """27번 train_direct_pdk와 동일한 SGD mini-batch + cosine schedule"""
    opt = torch.optim.Adam(model.parameters(), lr=lr)
    sched = torch.optim.lr_scheduler.CosineAnnealingLR(opt, T_max=n_iter)
    n = len(train_imgs)
    X = torch.cat(train_imgs, dim=0)
    for i in range(n_iter):
        idx = torch.randperm(n)[:batch]
        clean = X[idx]
        deg = degrade_fn(clean)
        opt.zero_grad()
        loss = loss_fn(model(deg), clean)
        loss.backward(); opt.step(); sched.step()
        if (i+1) % 200 == 0:
            print(f"    iter {i+1}/{n_iter}  loss={loss.item():.5f}")


# ============================================================================
# 5. Energy / latency model
# ============================================================================

ENERGY = dict(
    # CPU 기반 (Intel i5-14400F, DDR4-3200 기준)
    cpu_mac_pj       = 100,    # multiply-add per op (보수적, Horowitz 2014)
    dram_read_pj_bit = 5.0,    # DRAM read (Horowitz 2014)
    dram_write_pj_bit= 5.0,
    sram_read_pj_bit = 0.05,   # cache
    adc_pj_sample    = 5.0,    # 10-bit SAR ADC, JSSC 평균
    tia_pj_sample    = 0.5,    # generic TIA
    sense_pj_pix     = 0.5,    # photodiode sensing

    # OKA 추가
    oka_pgm_pulse_pJ_per_V_per_ns = 0.1,  # device 보수적 추정
)

LATENCY = dict(
    cpu_mac_ns       = 1.0,    # ~ 3.5GHz × 3-4 cycles for FMA
    dram_read_ns     = 50.0,
    dram_write_ns    = 50.0,
    sram_read_ns     = 1.0,
    adc_ns           = 100.0,  # 10-bit SAR
    tia_ns           = 10.0,
    sense_ns         = 1000.0, # photodiode response (보수적 1μs)
    # OKA: column-parallel readout 가정
)


def cpu_energy_latency(H, W, k=9, n_bit=8):
    """CPU 기반 한 frame 처리 energy/latency"""
    n_pix = H * W
    # Sensing
    e_sense = n_pix * ENERGY['sense_pj_pix']
    e_tia   = n_pix * ENERGY['tia_pj_sample']
    e_adc   = n_pix * ENERGY['adc_pj_sample']
    # DRAM write (raw image)
    e_dram_w_raw = n_pix * n_bit * ENERGY['dram_write_pj_bit']
    # DRAM read for conv (no cache assumption: read 9 neighbors per pixel)
    e_dram_r_in = n_pix * k * n_bit * ENERGY['dram_read_pj_bit']
    e_dram_r_w  = n_pix * k * n_bit * ENERGY['dram_read_pj_bit']  # weights
    # MAC
    e_mac = n_pix * k * ENERGY['cpu_mac_pj']
    # DRAM write (result)
    e_dram_w_out = n_pix * n_bit * ENERGY['dram_write_pj_bit']

    total_pJ = (e_sense + e_tia + e_adc + e_dram_w_raw +
                e_dram_r_in + e_dram_r_w + e_mac + e_dram_w_out)

    # Latency: column-parallel sensing 가정시 sensing+ADC만 column당 1회
    # CPU MAC는 multi-core 활용 가정 (10 cores × 4-wide SIMD)
    cores_simd = 10 * 4
    t_sense = LATENCY['sense_ns']      # 한 번에 모든 column
    t_tia   = LATENCY['tia_ns']
    t_adc   = H * LATENCY['adc_ns']    # row 단위 sequential
    t_dram_io = (n_pix * (1 + 2*k + 1)) * LATENCY['dram_read_ns'] / cores_simd
    t_mac     = (n_pix * k) * LATENCY['cpu_mac_ns'] / cores_simd
    total_ns  = t_sense + t_tia + t_adc + t_dram_io + t_mac

    return total_pJ, total_ns, dict(
        sense=e_sense, tia=e_tia, adc=e_adc,
        dram_w_raw=e_dram_w_raw, dram_r_in=e_dram_r_in,
        dram_r_w=e_dram_r_w, mac=e_mac, dram_w_out=e_dram_w_out)


def oka_energy_latency(H, W, ltd_model, I_state_map, n_bit=8,
                        N_frames=1000):
    """
    OKA 한 frame 처리 energy/latency.
    Programming은 amortize over N_frames.
    """
    n_cells = H * W * 9

    # ── Programming energy (1회) ─────────────────────────
    e_pgm_total = 0.0
    pulse_per_cell = []
    for ix in range(H):
        for iy in range(W):
            for tap in range(9):
                I_target = I_state_map[ix, iy, tap]
                V, N, _, _ = ltd_model.best_V_N(I_target)
                # programming pulse energy: V × I_avg × t_pulse
                # 보수적: t_pulse=0.1s, I_program ~ 1μA 가정
                # 또는 simpler model: V² × pulse_count × const
                e_per_pulse = (V**2) * 1e-3 * 1e8  # pJ unit (조정 필요)
                e_pgm_total += e_per_pulse * N
                pulse_per_cell.append(N)
    pulse_avg = np.mean(pulse_per_cell)

    # Per-frame inference
    n_pix = H * W
    # Sensing (광이 들어오면 자동 MAC)
    e_sense = n_pix * ENERGY['sense_pj_pix']
    # TIA: column-parallel (W TIAs)
    e_tia = W * ENERGY['tia_pj_sample']
    # ADC: column-parallel
    e_adc = W * ENERGY['adc_pj_sample'] * H  # row-by-row
    # SRAM write (작은 buffer)
    e_sram_w = n_pix * n_bit * ENERGY['sram_read_pj_bit']

    e_inference = e_sense + e_tia + e_adc + e_sram_w

    # Amortized total
    e_amortized = e_inference + e_pgm_total / N_frames

    # Latency
    t_sense = LATENCY['sense_ns']         # 모든 column 동시
    t_tia   = LATENCY['tia_ns']
    t_adc   = H * LATENCY['adc_ns']        # row sequential
    t_sram  = LATENCY['sram_read_ns']
    total_ns = t_sense + t_tia + t_adc + t_sram

    return e_amortized, total_ns, dict(
        pgm_total=e_pgm_total, pgm_amortized=e_pgm_total/N_frames,
        sense=e_sense, tia=e_tia, adc=e_adc, sram_w=e_sram_w,
        pulse_avg=pulse_avg, inference=e_inference)


# ============================================================================
# 6. Image utilities
# ============================================================================

def psnr(a, b):
    mse = F.mse_loss(a.float(), b.float()).item()
    return 99.9 if mse < 1e-10 else 10*np.log10(1.0/mse)

def ssim_simple(a, b):
    a, b = a.float(), b.float()
    mu_a = F.avg_pool2d(a, 11, stride=1, padding=5)
    mu_b = F.avg_pool2d(b, 11, stride=1, padding=5)
    mu_a2, mu_b2, mu_ab = mu_a**2, mu_b**2, mu_a*mu_b
    sa2 = F.avg_pool2d(a**2, 11, stride=1, padding=5) - mu_a2
    sb2 = F.avg_pool2d(b**2, 11, stride=1, padding=5) - mu_b2
    sab = F.avg_pool2d(a*b, 11, stride=1, padding=5) - mu_ab
    c1, c2 = 0.01**2, 0.03**2
    return ((2*mu_ab+c1)*(2*sab+c2) /
            ((mu_a2+mu_b2+c1)*(sa2+sb2+c2))).mean().item()


def load_image(path, size=256):
    img = Image.open(path).convert('L')
    w, h = img.size; s = min(w, h)
    img = img.crop(((w-s)//2, (h-s)//2, (w+s)//2, (h+s)//2))
    img = img.resize((size, size), Image.BILINEAR)
    arr = np.array(img, dtype=np.float32)/255.
    return torch.from_numpy(arr).unsqueeze(0).unsqueeze(0)


def find_image(data_dir, name):
    exts = (".png",".jpg",".jpeg",".PNG",".JPG",".JPEG")
    for ext in exts:
        c = os.path.join(data_dir, name+ext)
        if os.path.exists(c): return c
        hits = glob.glob(os.path.join(data_dir, "**", name+ext),
                         recursive=True)
        if hits: return hits[0]
    return None


# ============================================================================
# 7. Main
# ============================================================================

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--pr-data",    default="./expdata/photoresponsivity.xlsx")
    p.add_argument("--ltd-data",   default="./expdata/depression_data.xlsx")
    p.add_argument("--train-dir",  default="./data/BSD300/images/train")
    p.add_argument("--inf-dir",    default="./data/BSD300/images/test")
    p.add_argument("--inf-name",   default="102061")
    p.add_argument("--img-size",   type=int, default=256)
    # PDK training
    p.add_argument("--n-train",    type=int, default=200)
    p.add_argument("--pdk-iter",   type=int, default=2000)
    p.add_argument("--pdk-lr",     type=float, default=0.01)
    p.add_argument("--batch",      type=int, default=32)
    # Degradation (27번과 동일 기본값)
    p.add_argument("--sigma0",     type=float, default=0.40)
    p.add_argument("--alpha-psf",  type=float, default=2.0)
    p.add_argument("--coma-k",     type=float, default=0.60)
    p.add_argument("--alpha-vig",  type=float, default=4.0)
    # OKA mapping
    p.add_argument("--P-ref",      type=float, default=600.0,
                   help="reference light intensity (μW/cm²) for mapping")
    p.add_argument("--P-scale",    type=float, default=1000.0,
                   help="image grayscale [0,1] → light power scaling")
    p.add_argument("--I-min",      type=float, default=0.3)
    p.add_argument("--I-max",      type=float, default=250.0)
    p.add_argument("--n-frames",   type=int, default=1000)
    p.add_argument("--res-dir",    default="./res/oka_simulation")
    return p.parse_args()


def main():
    args = parse_args()
    os.makedirs(args.res_dir, exist_ok=True)
    H = W = args.img_size

    # ── 1. Device model fitting ─────────────────────────
    print("="*60)
    print("[Step 1] Device model fitting from measurement data")
    print("="*60)
    pr_model, _ = PhotoresponsivityModel.fit_from_xlsx(args.pr_data)
    ltd_model = LTDModel.fit_from_xlsx(args.ltd_data)

    # ── 2. Train PDK fresh ──────────────────────────────
    print("\n" + "="*60)
    print("[Step 2] Train DirectPDK from scratch (within 32)")
    print("="*60)
    train_imgs = load_train_images(args.train_dir, args.n_train, H)
    if len(train_imgs) == 0:
        print(f"  [Error] no train images in {args.train_dir}")
        return

    degrade = lambda x: degrade_coma_vig(
        x, args.sigma0, args.alpha_psf, args.coma_k, args.alpha_vig)

    pdk = DirectPDK(H, W)
    print(f"  PDK params: {H}×{W}×9 = {H*W*9:,}")
    train_pdk(pdk, train_imgs, degrade,
               n_iter=args.pdk_iter, lr=args.pdk_lr, batch=args.batch)

    weights = pdk.kernels.detach()
    torch.save({'kernels': weights},
               os.path.join(args.res_dir, 'pdk_kernels.pt'))
    print(f"  saved: {os.path.join(args.res_dir, 'pdk_kernels.pt')}")

    # ── 3. Map weights → I_state ────────────────────────
    print("\n" + "="*60)
    print("[Step 3] Map PDK weights → device I_state")
    print("="*60)
    I_state_map, info = map_weights_to_Istate(
        weights, pr_model, args.P_ref,
        I_min=args.I_min, I_max=args.I_max)
    print(f"  I_state range: [{I_state_map.min():.2f}, "
          f"{I_state_map.max():.2f}] nA")
    print(f"  R_mid (bias): {info['R_mid']:.3f}")

    # ── 4. Inference comparison ─────────────────────────
    print("\n" + "="*60)
    print("[Step 4] Inference: Input / Degraded / CPU / OKA")
    print("="*60)

    src = find_image(args.inf_dir, args.inf_name)
    if src is None:
        print(f"  [Error] {args.inf_name} not found in {args.inf_dir}")
        return
    clean = load_image(src, H)        # [0, 1]
    print(f"  Test image: {args.inf_name}")

    # 4-stage pipeline:
    #   clean → degrade → CPU(PDK)  : ideal digital correction
    #   clean → degrade → OKA       : device-aware analog correction
    with torch.no_grad():
        deg = degrade(clean)
        # CPU: PDK kernel을 [0,1] domain에서 ideal conv
        out_cpu = ideal_conv(deg, weights).clamp(0, 1)
        # OKA: device-aware (P_image scale 적용)
        P_deg = deg * args.P_scale         # μW/cm²
        out_oka_raw = oka_forward(P_deg, I_state_map, pr_model, info)

    print(f"  Light intensity range (degraded): "
          f"[{P_deg.min():.1f}, {P_deg.max():.1f}] μW/cm²")

    # OKA 출력 정규화 (clean 범위로 맞춤)
    def normalize(x, ref=None):
        x_min = x.min().item()
        x_max = x.max().item()
        x_n = (x - x_min) / max(x_max - x_min, 1e-9)
        return x_n
    out_oka = normalize(out_oka_raw)

    # 4가지 비교 PSNR/SSIM (vs clean)
    p_deg = psnr(clean, deg);     s_deg = ssim_simple(clean, deg)
    p_cpu = psnr(clean, out_cpu); s_cpu = ssim_simple(clean, out_cpu)
    p_oka = psnr(clean, out_oka); s_oka = ssim_simple(clean, out_oka)
    p_cpu_oka = psnr(out_cpu, out_oka)
    s_cpu_oka = ssim_simple(out_cpu, out_oka)

    print(f"  vs clean:")
    print(f"    Degraded : PSNR={p_deg:.2f}, SSIM={s_deg:.4f}")
    print(f"    CPU(PDK) : PSNR={p_cpu:.2f}, SSIM={s_cpu:.4f}")
    print(f"    OKA      : PSNR={p_oka:.2f}, SSIM={s_oka:.4f}")
    print(f"  CPU vs OKA: PSNR={p_cpu_oka:.2f}, SSIM={s_cpu_oka:.4f}")

    # ─ 4-panel 비교 그림 ──────────────────────────────
    fig, axes = plt.subplots(1, 4, figsize=(17, 4.5))
    panels = [
        ('Input (clean)',                  clean,    None,      None),
        (f'Degraded\nPSNR={p_deg:.1f}',    deg,      p_deg,     s_deg),
        (f'CPU (PDK ideal)\nPSNR={p_cpu:.1f}', out_cpu, p_cpu, s_cpu),
        (f'OKA (device-aware)\nPSNR={p_oka:.1f}', out_oka, p_oka, s_oka),
    ]
    for ax, (title, img, _, _) in zip(axes, panels):
        ax.imshow(img.squeeze().numpy(), cmap='gray', vmin=0, vmax=1)
        ax.set_title(title, fontsize=10, fontweight='bold')
        ax.axis('off')
    plt.tight_layout()
    fig.savefig(os.path.join(args.res_dir, 'inference_compare.png'),
                dpi=150, bbox_inches='tight')
    plt.close(fig)

    # ─ 개별 figure 저장 ──────────────────────────────
    individual_dir = os.path.join(args.res_dir, 'individual')
    os.makedirs(individual_dir, exist_ok=True)
    for name, img in [
        ('input',    clean),
        ('degraded', deg),
        ('cpu',      out_cpu),
        ('oka',      out_oka),
    ]:
        arr = np.clip(img.squeeze().numpy(), 0, 1)
        f, a = plt.subplots(1, 1,
            figsize=(arr.shape[1]/100, arr.shape[0]/100), dpi=300)
        a.imshow(arr, cmap='gray', vmin=0, vmax=1)
        a.axis('off')
        plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
        f.savefig(os.path.join(individual_dir,
                                f'{args.inf_name}_{name}.png'),
                  dpi=300, bbox_inches='tight', pad_inches=0)
        plt.close(f)
    print(f"  [Saved] individual: {individual_dir}/")

    # ── 5. Energy/Latency comparison ────────────────────
    print("\n" + "="*60)
    print("[Step 5] Energy / Latency comparison")
    print("="*60)
    e_cpu, t_cpu, br_cpu = cpu_energy_latency(H, W)
    e_oka, t_oka, br_oka = oka_energy_latency(
        H, W, ltd_model, I_state_map,
        N_frames=args.n_frames)

    print(f"\n  CPU per-frame:")
    print(f"    Energy:  {e_cpu/1e6:>10.3f} μJ")
    print(f"    Latency: {t_cpu/1e6:>10.3f} ms")
    print(f"    Breakdown:")
    for k, v in br_cpu.items():
        print(f"      {k:<14}: {v/1e6:>8.3f} μJ ({100*v/e_cpu:>5.1f}%)")

    print(f"\n  OKA per-frame (amortized over {args.n_frames} frames):")
    print(f"    Energy:  {e_oka/1e6:>10.3f} μJ")
    print(f"    Latency: {t_oka/1e3:>10.3f} μs")
    print(f"    Programming total: {br_oka['pgm_total']/1e6:.2f} μJ "
          f"(avg {br_oka['pulse_avg']:.1f} pulses/cell)")
    print(f"    Inference only:    {br_oka['inference']/1e6:.3f} μJ")

    print(f"\n  Comparison:")
    print(f"    Energy   ratio (CPU/OKA): {e_cpu/e_oka:>10.1f}×")
    print(f"    Latency  ratio (CPU/OKA): {t_cpu/t_oka:>10.1f}×")

    # CSV summary
    import csv
    with open(os.path.join(args.res_dir, '32_summary.csv'), 'w') as f:
        w = csv.writer(f)
        w.writerow(['metric', 'CPU', 'OKA', 'ratio (CPU/OKA)'])
        w.writerow(['Energy_uJ', f'{e_cpu/1e6:.3f}',
                    f'{e_oka/1e6:.3f}', f'{e_cpu/e_oka:.1f}'])
        w.writerow(['Latency_us', f'{t_cpu/1e3:.3f}',
                    f'{t_oka/1e3:.3f}', f'{t_cpu/t_oka:.1f}'])
        w.writerow(['PSNR_CPUvsOKA_dB', '-',
                    f'{p_cpu_oka:.2f}', '-'])
        w.writerow(['SSIM_CPUvsOKA', '-',
                    f'{s_cpu_oka:.4f}', '-'])
        w.writerow(['Pulses_per_cell_avg', '-',
                    f'{br_oka["pulse_avg"]:.1f}', '-'])
    print(f"\n[Saved] {os.path.join(args.res_dir, '32_summary.csv')}")
    print(f"[Saved] {os.path.join(args.res_dir, 'inference_compare.png')}")
    print(f"\n[Done] {os.path.abspath(args.res_dir)}")


if __name__ == '__main__':
    main()