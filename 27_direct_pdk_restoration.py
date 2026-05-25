"""
27_direct_pdk_restoration.py
==========================================
DirectPDK vs Global kernel 이미지 복원 비교

23번(ComaPDK, 6 params)과의 차이:
  - DirectPDK: H×W×9 파라미터 직접 학습 (no functional form 가정)
  - SGD mini-batch 방식 (on-the-fly degradation)
  - 더 유연하지만 파라미터 수 많음

비교:
  Global  : 9 params (spatially invariant)
  DirectPDK: H×W×9 params (spatially varying, direct)

Degradation: Coma + Vignetting (23번과 동일)
Dataset:     BSDS300 train/test

출력:
  res/direct_pdk/
    figures/raw_XX.png
    figures/degraded_XX.png
    figures/global_XX.png
    figures/pdk_XX.png
    figures/kernel_map.png
    figures/metrics_table.png
    27_summary.png

Run:
  python 27_direct_pdk_restoration.py
  python 27_direct_pdk_restoration.py \\
      --train-dir ./data/BSD300/images/train \\
      --inf-dir   ./data/BSD300/images/test  \\
      --n-train 20 --img-size 256 --n-iter 2000
"""

import argparse, os, glob, random
import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from PIL import Image

SEED = 42
torch.manual_seed(SEED); np.random.seed(SEED); random.seed(SEED)


# ============================================================================
# Utilities
# ============================================================================

def radial_map(H, W, device="cpu"):
    ys = torch.linspace(-1, 1, H, device=device)
    xs = torch.linspace(-1, 1, W, device=device)
    gy, gx = torch.meshgrid(ys, xs, indexing="ij")
    r = (gy**2 + gx**2).sqrt() / (2**0.5)
    return r, gy, gx

def spatially_varying_conv(x, kernels):
    B, C, H, W = x.shape
    patches = F.unfold(x, kernel_size=3, padding=1).view(B, C, 9, H*W)
    k = kernels.view(H*W, 9).T.unsqueeze(0).unsqueeze(0)
    return (patches * k).sum(dim=2).view(B, C, H, W).clamp(0, 1)

def global_conv(x, kernel_9):
    k = kernel_9.view(1, 1, 3, 3)
    return F.conv2d(x, k.expand(x.shape[1], 1, 3, 3),
                    padding=1, groups=x.shape[1]).clamp(0, 1)

def loss_fn(out, clean):
    l1 = F.l1_loss(out, clean)
    def gmap(t): return t[:,:,:,1:]-t[:,:,:,:-1], t[:,:,1:,:]-t[:,:,:-1,:]
    ox, oy = gmap(out); cx, cy = gmap(clean)
    return l1 + 0.1*(F.l1_loss(ox,cx)+F.l1_loss(oy,cy))

def psnr(a, b):
    mse = F.mse_loss(a.float(), b.float()).item()
    return 99.9 if mse < 1e-10 else 10*np.log10(1.0/mse)

def ssim_simple(a, b):
    a = a.float(); b = b.float()
    mu_a = F.avg_pool2d(a,11,stride=1,padding=5)
    mu_b = F.avg_pool2d(b,11,stride=1,padding=5)
    mu_a2=mu_a**2; mu_b2=mu_b**2; mu_ab=mu_a*mu_b
    sig_a2=F.avg_pool2d(a**2,11,stride=1,padding=5)-mu_a2
    sig_b2=F.avg_pool2d(b**2,11,stride=1,padding=5)-mu_b2
    sig_ab=F.avg_pool2d(a*b, 11,stride=1,padding=5)-mu_ab
    c1,c2=0.01**2,0.03**2
    return ((2*mu_ab+c1)*(2*sig_ab+c2)/
            ((mu_a2+mu_b2+c1)*(sig_a2+sig_b2+c2))).mean().item()

def to_np(t):
    return t.squeeze().detach().cpu().float().numpy()


# ============================================================================
# Degradation (23번과 동일)
# ============================================================================

def degrade_coma_vig(x, sigma0=0.40, alpha_psf=2.0,
                     coma_k=0.60, alpha_vig=4.0):
    r, gy, gx = radial_map(*x.shape[2:], x.device)
    r2 = r**2; rs = r.clamp(1e-6)
    sigma = (sigma0 + alpha_psf*r2).clamp(0.1, 2.0)
    shift_y = coma_k*r2*(gy/rs)
    shift_x = coma_k*r2*(gx/rs)
    coords = torch.tensor([
        [-1.,-1.],[-1.,0.],[-1.,1.],
        [ 0.,-1.],[ 0.,0.],[ 0.,1.],
        [ 1.,-1.],[ 1.,0.],[ 1.,1.]], device=x.device)
    d2 = (coords**2).sum(-1)
    s  = sigma.unsqueeze(-1)
    g1 = torch.exp(-d2/(2*s**2)); g1=g1/g1.sum(-1,keepdim=True)
    sy = shift_y.unsqueeze(-1); sx = shift_x.unsqueeze(-1)
    dy = coords[:,0].unsqueeze(0).unsqueeze(0)-sy
    dx = coords[:,1].unsqueeze(0).unsqueeze(0)-sx
    g2 = torch.exp(-(dy**2+dx**2)/(2*s**2)); g2=g2/g2.sum(-1,keepdim=True)
    w  = (r*0.9).clamp(0,0.9).unsqueeze(-1)
    k  = (1-w)*g1+w*g2; k=k/k.sum(-1,keepdim=True)
    x  = spatially_varying_conv(x, k)
    V  = 1.0/(1.0+alpha_vig*r2)
    return (x*V.unsqueeze(0).unsqueeze(0)).clamp(0,1)


# ============================================================================
# Models
# ============================================================================

class GlobalKernel(nn.Module):
    """9 params: spatially invariant 3×3"""
    def __init__(self):
        super().__init__()
        init = torch.zeros(9); init[4] = 3.0
        self.kernel_logits = nn.Parameter(init)

    def forward(self, x):
        k = torch.softmax(self.kernel_logits, dim=0)
        return global_conv(x, k)

    def get_kernel_3x3(self):
        return torch.softmax(self.kernel_logits, dim=0
                             ).detach().cpu().numpy().reshape(3,3)


class DirectPDK(nn.Module):
    """
    H×W×9 파라미터 직접 학습.
    - delta 초기화: center=1, 나머지=0
    - 어떤 함수 형태도 가정하지 않음
    - SGD로 학습 시 각 위치가 독립적으로 최적화
    """
    def __init__(self, H, W):
        super().__init__()
        init = torch.zeros(H, W, 9)
        init[:, :, 4] = 1.0          # delta: correction 없음 상태
        self.kernels = nn.Parameter(init)

    def forward(self, x):
        return spatially_varying_conv(x, self.kernels)

    def get_kernels(self):
        return self.kernels.detach().cpu().numpy()


# ============================================================================
# Data loading (23번과 동일)
# ============================================================================

def load_images(data_dir, n, size):
    exts = ("*.png","*.jpg","*.jpeg","*.PNG","*.JPG","*.JPEG")
    paths = []
    for ext in exts:
        paths += glob.glob(os.path.join(data_dir,"**",ext), recursive=True)
        paths += glob.glob(os.path.join(data_dir, ext))
    paths = sorted(set(paths))
    random.shuffle(paths)
    imgs = []
    for p in paths:
        try:
            img = Image.open(p).convert("L")
            w,h = img.size; s=min(w,h)
            img = img.crop(((w-s)//2,(h-s)//2,(w+s)//2,(h+s)//2))
            img = img.resize((size,size), Image.BILINEAR)
            t = torch.from_numpy(
                np.array(img,dtype=np.float32)/255.
            ).unsqueeze(0).unsqueeze(0)
            imgs.append(t)
        except: continue
        if len(imgs)==n: break
    print(f"  [Load] {len(imgs)} images from {data_dir}")
    return imgs

def load_images_by_names(data_dir, names, size):
    exts = (".png",".jpg",".jpeg",".PNG",".JPG",".JPEG")
    imgs = []
    for name in names:
        found = None
        for ext in exts:
            c = os.path.join(data_dir, name+ext)
            if os.path.exists(c): found=c; break
        if found is None:
            for ext in exts:
                hits = glob.glob(os.path.join(data_dir,"**",name+ext),recursive=True)
                if hits: found=hits[0]; break
        if found is None:
            print(f"  [Warning] {name} not found"); continue
        try:
            img = Image.open(found).convert("L")
            w,h=img.size; s=min(w,h)
            img=img.crop(((w-s)//2,(h-s)//2,(w+s)//2,(h+s)//2))
            img=img.resize((size,size),Image.BILINEAR)
            t=torch.from_numpy(np.array(img,dtype=np.float32)/255.
                               ).unsqueeze(0).unsqueeze(0)
            imgs.append(t)
            print(f"  [Load] {found}")
        except Exception as e: print(f"  [Error] {e}")
    return imgs


# ============================================================================
# Training
# ============================================================================

def train_global(model, train_imgs, degrade_fn, n_iter, lr):
    """Global: 파라미터 적으므로 dataset 전체 평균 방식"""
    opt   = torch.optim.Adam(model.parameters(), lr=lr)
    sched = torch.optim.lr_scheduler.CosineAnnealingLR(opt, T_max=n_iter)
    pairs = [(img, degrade_fn(img)) for img in train_imgs]
    for i in range(n_iter):
        opt.zero_grad()
        total = sum(loss_fn(model(deg), clean) for clean,deg in pairs)
        total = total / len(pairs)
        total.backward(); opt.step(); sched.step()
        if (i+1) % 200 == 0:
            print(f"    iter {i+1}/{n_iter}  loss={total.item():.5f}")


def train_direct_pdk(model, train_imgs, degrade_fn, n_iter, lr, batch=4):
    """
    DirectPDK SGD mini-batch 학습.
    매 iter마다 랜덤 batch 이미지 샘플링 → on-the-fly degradation
    """
    opt   = torch.optim.Adam(model.parameters(), lr=lr)
    sched = torch.optim.lr_scheduler.CosineAnnealingLR(opt, T_max=n_iter)
    n = len(train_imgs)
    X = torch.cat(train_imgs, dim=0)

    for i in range(n_iter):
        idx   = torch.randperm(n)[:batch]
        clean = X[idx]
        deg   = degrade_fn(clean)
        opt.zero_grad()
        loss = loss_fn(model(deg), clean)
        loss.backward(); opt.step(); sched.step()
        if (i+1) % 400 == 0:
            print(f"    iter {i+1}/{n_iter}  loss={loss.item():.5f}")


# ============================================================================
# Saving
# ============================================================================

def save_img(arr, path):
    arr = np.clip(arr, 0, 1)
    fig, ax = plt.subplots(1,1,
        figsize=(arr.shape[1]/100, arr.shape[0]/100), dpi=300)
    ax.imshow(arr, cmap="gray", vmin=0, vmax=1)
    ax.axis("off")
    plt.subplots_adjust(left=0,right=1,top=1,bottom=0)
    fig.savefig(path, dpi=300, bbox_inches="tight", pad_inches=0)
    plt.close(fig)


def save_kernel_map(kernels_hwk, path, n_sample=5):
    H, W, _ = kernels_hwk.shape
    # 가장자리/꼭짓점 제외: 전체 범위의 10~90% 내부만 샘플
    margin = max(int(H * 0.10), 4)
    positions = [(iy, ix)
                 for iy in np.linspace(margin, H-1-margin, n_sample, dtype=int)
                 for ix in np.linspace(margin, W-1-margin, n_sample, dtype=int)]
    vmax = max(np.abs(kernels_hwk).max(), 0.01)

    fig = plt.figure(figsize=(n_sample*1.8+0.8, n_sample*1.8))
    from matplotlib.gridspec import GridSpec
    gs = GridSpec(n_sample, n_sample+1, figure=fig,
                  width_ratios=[1]*n_sample+[0.08],
                  hspace=0.35, wspace=0.15)
    im = None
    for idx, (iy,ix) in enumerate(positions):
        r,c = idx//n_sample, idx%n_sample
        ax = fig.add_subplot(gs[r,c])
        k  = kernels_hwk[iy,ix].reshape(3,3)
        im = ax.imshow(k, cmap="RdBu_r",
                       vmin=-vmax, vmax=vmax,
                       interpolation="nearest", aspect="equal")
        ax.set_xticks(np.arange(-0.5,3,1),minor=True)
        ax.set_yticks(np.arange(-0.5,3,1),minor=True)
        ax.tick_params(which="minor",length=0)
        ax.grid(which="minor",color="black",linewidth=0.8)
        ax.set_xticks([]); ax.set_yticks([])
        ax.set_title(f"({ix},{iy})",fontsize=7,pad=3)
    cb_ax = fig.add_subplot(gs[:,n_sample])
    cb = fig.colorbar(im, cax=cb_ax)
    cb.set_label("kernel weight",fontsize=8)
    cb.ax.tick_params(labelsize=7)
    fig.suptitle("DirectPDK kernel map (sampled positions)",
                 fontsize=10, y=1.01)
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  [Saved] {path}")


def save_metrics_data(results, path_csv, path_xlsx=None):
    """
    PSNR/SSIM 수치를 CSV로 저장. xlsx 옵션도 지원.
    각 행: image, psnr_deg, psnr_glb, psnr_pdk, ssim_deg, ssim_glb, ssim_pdk
    마지막 행: Mean
    """
    import csv
    headers = ["image",
               "psnr_deg", "psnr_glb", "psnr_pdk",
               "ssim_deg", "ssim_glb", "ssim_pdk"]
    rows = []
    for i, r in enumerate(results):
        rows.append([f"img_{i+1}",
                     f"{r['psnr_deg']:.4f}",
                     f"{r['psnr_glb']:.4f}",
                     f"{r['psnr_pdk']:.4f}",
                     f"{r['ssim_deg']:.6f}",
                     f"{r['ssim_glb']:.6f}",
                     f"{r['ssim_pdk']:.6f}"])
    means = [np.mean([r[k] for r in results])
             for k in ["psnr_deg","psnr_glb","psnr_pdk",
                       "ssim_deg","ssim_glb","ssim_pdk"]]
    rows.append(["Mean"] + [f"{v:.4f}" if i<3 else f"{v:.6f}"
                            for i,v in enumerate(means)])

    with open(path_csv, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers); w.writerows(rows)
    print(f"  [Saved] {path_csv}")

    if path_xlsx is not None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = Workbook(); ws = wb.active; ws.title = "metrics"
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", start_color="2C2C2A")
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[cell.column_letter].width = 14
            for i, row in enumerate(rows, 2):
                for c, val in enumerate(row, 1):
                    cell = ws.cell(row=i, column=c, value=val)
                    cell.alignment = Alignment(horizontal="center")
                    if c == 4 or c == 7:   # PDK 열 강조
                        cell.font = Font(bold=True, color="085041")
                    if i == len(rows)+1:    # Mean 행 강조
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill("solid", start_color="F1EFE8")
            wb.save(path_xlsx)
            print(f"  [Saved] {path_xlsx}")
        except ImportError:
            print("  [Warning] openpyxl not found, skipping xlsx export")


def save_metrics_histogram(results, path):
    """
    PSNR / SSIM histogram (각 condition별로 그룹화).
    가로축: 이미지 인덱스, 세로축: 값
    Degraded / Global / PDK 3개 막대 묶음
    """
    n = len(results)
    x = np.arange(n)
    width = 0.27

    fig, axes = plt.subplots(1, 2, figsize=(max(8, n*1.2), 5))

    # PSNR
    ax = axes[0]
    ax.bar(x - width, [r['psnr_deg'] for r in results],
           width, label='Degraded', color='#E24B4A')
    ax.bar(x,         [r['psnr_glb'] for r in results],
           width, label='Global',   color='#378ADD')
    ax.bar(x + width, [r['psnr_pdk'] for r in results],
           width, label='PDK',      color='#EF9F27')
    ax.set_xticks(x)
    ax.set_xticklabels([f"img {i+1}" for i in range(n)], fontsize=9)
    ax.set_ylabel("PSNR (dB)")
    ax.set_title("PSNR per image", fontweight="bold")
    ax.legend(loc='lower right', fontsize=9)
    ax.grid(axis='y', alpha=0.3)

    # SSIM
    ax = axes[1]
    ax.bar(x - width, [r['ssim_deg'] for r in results],
           width, label='Degraded', color='#E24B4A')
    ax.bar(x,         [r['ssim_glb'] for r in results],
           width, label='Global',   color='#378ADD')
    ax.bar(x + width, [r['ssim_pdk'] for r in results],
           width, label='PDK',      color='#EF9F27')
    ax.set_xticks(x)
    ax.set_xticklabels([f"img {i+1}" for i in range(n)], fontsize=9)
    ax.set_ylabel("SSIM")
    ax.set_ylim(0, 1.05)
    ax.set_title("SSIM per image", fontweight="bold")
    ax.legend(loc='lower right', fontsize=9)
    ax.grid(axis='y', alpha=0.3)

    plt.suptitle("Image restoration quality — Degraded vs Global vs PDK",
                 fontsize=12, y=1.00)
    plt.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  [Saved] {path}")


def save_metrics_table(results, path):
    headers = ["Image","Degraded","Global","DirectPDK"]
    rows_psnr, rows_ssim = [], []
    for i,r in enumerate(results):
        rows_psnr.append([f"img {i+1}",
                          f"{r['psnr_deg']:.2f}",
                          f"{r['psnr_glb']:.2f}",
                          f"{r['psnr_pdk']:.2f}"])
        rows_ssim.append([f"img {i+1}",
                          f"{r['ssim_deg']:.4f}",
                          f"{r['ssim_glb']:.4f}",
                          f"{r['ssim_pdk']:.4f}"])
    def mean_col(rows,col):
        return np.mean([float(r[col]) for r in rows])
    for rows in [rows_psnr, rows_ssim]:
        rows.append(["Mean",
                     f"{mean_col(rows,1):.2f}",
                     f"{mean_col(rows,2):.2f}",
                     f"{mean_col(rows,3):.2f}"])

    fig, axes = plt.subplots(1,2,figsize=(10, 0.5+0.5*len(rows_psnr)))
    for ax,rows,title in [(axes[0],rows_psnr,"PSNR (dB)"),
                          (axes[1],rows_ssim,"SSIM")]:
        ax.axis("off")
        tbl = ax.table(cellText=rows, colLabels=headers,
                       loc="center", cellLoc="center")
        tbl.auto_set_font_size(False); tbl.set_fontsize(9); tbl.scale(1,1.5)
        for j in range(len(headers)):
            tbl[0,j].set_facecolor("#2C2C2A")
            tbl[0,j].set_text_props(color="white",fontweight="bold")
        for i in range(1,len(rows)+1):
            tbl[i,3].set_facecolor("#E1F5EE")
            tbl[i,3].set_text_props(color="#085041",fontweight="bold")
        for j in range(len(headers)):
            tbl[len(rows),j].set_facecolor("#F1EFE8")
            tbl[len(rows),j].set_text_props(fontweight="bold")
        ax.set_title(title,fontsize=11,pad=10)
    plt.suptitle("Coma + Vignetting correction: Global vs DirectPDK",
                 fontsize=12, y=1.02)
    plt.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  [Saved] {path}")


def save_summary(results, path):
    n = len(results)
    fig, axes = plt.subplots(n, 4, figsize=(14, 3.6*n))
    if n==1: axes=axes[np.newaxis,:]
    titles = ["Clean","Degraded","Global kernel","DirectPDK"]
    for c,t in enumerate(titles):
        axes[0,c].set_title(t,fontsize=11,fontweight="bold",pad=6)
    for r,res in enumerate(results):
        for c,key in enumerate(["raw","deg","glb","pdk"]):
            axes[r,c].imshow(to_np(res[key]),cmap="gray",vmin=0,vmax=1)
            axes[r,c].axis("off")
        axes[r,1].text(0.5,-0.04,
            f"PSNR={res['psnr_deg']:.1f}  SSIM={res['ssim_deg']:.3f}",
            transform=axes[r,1].transAxes,
            ha="center",fontsize=8,color="dimgray")
        axes[r,2].text(0.5,-0.04,
            f"PSNR={res['psnr_glb']:.1f}  SSIM={res['ssim_glb']:.3f}",
            transform=axes[r,2].transAxes,
            ha="center",fontsize=8,color="steelblue")
        axes[r,3].text(0.5,-0.04,
            f"PSNR={res['psnr_pdk']:.1f}  SSIM={res['ssim_pdk']:.3f}",
            transform=axes[r,3].transAxes,
            ha="center",fontsize=8,color="#085041")
    plt.tight_layout(rect=[0,0.03,1,1])
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  [Saved] {path}")


# ============================================================================
# Args
# ============================================================================

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--train-dir", type=str,
                   default="./data/BSD300/images/train")
    p.add_argument("--inf-dir",   type=str,
                   default="./data/BSD300/images/test")
    p.add_argument("--n-train",   type=int, default=200)
    p.add_argument("--img-size",  type=int, default=256)
    p.add_argument("--n-iter",    type=int, default=2000,
                   help="PDK SGD iterations (more needed vs 23번)")
    p.add_argument("--glb-iter",  type=int, default=800)
    p.add_argument("--lr",        type=float, default=0.01)
    p.add_argument("--batch",     type=int,   default=32,
                   help="SGD mini-batch size for DirectPDK")
    p.add_argument("--sigma0",    type=float, default=0.40)
    p.add_argument("--alpha-psf", type=float, default=2.0)
    p.add_argument("--coma-k",    type=float, default=0.60)
    p.add_argument("--alpha-vig", type=float, default=4.0)
    p.add_argument("--res-dir",   type=str,
                   default="./res/direct_pdk")
    return p.parse_args()


# ============================================================================
# Main
# ============================================================================

def main():
    args = parse_args()
    fig_dir = os.path.join(args.res_dir, "figures")
    os.makedirs(fig_dir, exist_ok=True)
    H = W = args.img_size

    degrade = lambda x: degrade_coma_vig(
        x, args.sigma0, args.alpha_psf, args.coma_k, args.alpha_vig)

    INF_NAMES = ["102061","143090","103070","145086"]

    print("[Load] Train images...")
    train_imgs = load_images(args.train_dir, args.n_train, H)
    print("[Load] Inference images...")
    inf_imgs   = load_images_by_names(args.inf_dir, INF_NAMES, H)

    # ── Global kernel (full-batch, 파라미터 9개) ──────────────
    print(f"\n[Train] Global kernel ({args.glb_iter} iter, full-batch)...")
    glb_model = GlobalKernel()
    train_global(glb_model, train_imgs, degrade,
                 n_iter=args.glb_iter, lr=0.02)

    # ── DirectPDK (SGD, H×W×9 params) ────────────────────────
    n_params = H * W * 9
    print(f"\n[Train] DirectPDK ({args.n_iter} iter, SGD batch={args.batch})")
    print(f"  params: {H}×{W}×9 = {n_params:,}")
    pdk_model = DirectPDK(H, W)
    train_direct_pdk(pdk_model, train_imgs, degrade,
                     n_iter=args.n_iter, lr=args.lr, batch=args.batch)

    # ── Save models ───────────────────────────────────────────
    torch.save(pdk_model.state_dict(),
               os.path.join(args.res_dir, "direct_pdk.pt"))

    # ── Inference ─────────────────────────────────────────────
    print("\n[Inference] unseen images...")
    results = []
    for i, clean in enumerate(inf_imgs):
        deg = degrade(clean)
        with torch.no_grad():
            glb = glb_model(deg)
            pdk = pdk_model(deg)

        p_deg=psnr(clean,deg); s_deg=ssim_simple(clean,deg)
        p_glb=psnr(clean,glb); s_glb=ssim_simple(clean,glb)
        p_pdk=psnr(clean,pdk); s_pdk=ssim_simple(clean,pdk)

        print(f"  img {i+1}:  Deg={p_deg:.2f}  "
              f"Glb={p_glb:.2f}  PDK={p_pdk:.2f} dB")

        tag = f"{i+1:02d}"
        save_img(to_np(clean), os.path.join(fig_dir, f"raw_{tag}.png"))
        save_img(to_np(deg),   os.path.join(fig_dir, f"degraded_{tag}.png"))
        save_img(to_np(glb),   os.path.join(fig_dir, f"global_{tag}.png"))
        save_img(to_np(pdk),   os.path.join(fig_dir, f"pdk_{tag}.png"))

        results.append(dict(
            raw=clean, deg=deg, glb=glb, pdk=pdk,
            psnr_deg=p_deg, psnr_glb=p_glb, psnr_pdk=p_pdk,
            ssim_deg=s_deg, ssim_glb=s_glb, ssim_pdk=s_pdk))

    # ── Kernel map ────────────────────────────────────────────
    print("\n[Kernel] Saving DirectPDK kernel map...")
    kernels = pdk_model.get_kernels()
    save_kernel_map(kernels, os.path.join(fig_dir, "kernel_map.png"))

    glb_k = glb_model.get_kernel_3x3()
    print(f"  Global kernel:\n{np.round(glb_k,4)}")

    # 주요 위치 수치 출력
    sample_pts = [(0,0,"top-left"),(H//2,W//2,"center"),(H-1,W-1,"bot-right")]
    print("  DirectPDK sample kernels:")
    for iy,ix,name in sample_pts:
        k = kernels[iy,ix].reshape(3,3)
        print(f"    [{name}]:\n{np.round(k,4)}")

    # ── Metrics + Summary ─────────────────────────────────────
    save_metrics_table(results, os.path.join(fig_dir, "metrics_table.png"))
    save_metrics_histogram(results,
        os.path.join(args.res_dir, "27_metrics_histogram.png"))
    save_metrics_data(results,
        os.path.join(args.res_dir, "27_metrics.csv"),
        os.path.join(args.res_dir, "27_metrics.xlsx"))
    save_summary(results, os.path.join(args.res_dir, "27_summary.png"))

    print("\n" + "="*55)
    print("[Result] Mean metrics (unseen images):")
    for key, label in [("psnr_deg","Degraded"),
                        ("psnr_glb","Global  "),
                        ("psnr_pdk","DirectPDK")]:
        vals = [r[key] for r in results]
        print(f"  {label}: PSNR={np.mean(vals):.2f} dB")
    print("="*55)
    print(f"\n[Done] {os.path.abspath(args.res_dir)}")
    print(f"  DirectPDK params: {n_params:,}  vs  Global params: 9")


if __name__ == "__main__":
    main()