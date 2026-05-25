"""
26_classification_pdk.py
==========================================
단순 FC 분류기로 degradation의 영향 측정

목적:
  CNN을 쓰면 degradation에 robust해져서 차이가 없음
  → 단순 FC (flatten → linear) 만 사용
  → spatial structure에 의존 불가
  → degradation이 분류 성능에 직접 영향

실험 구성:
  Dataset: Fashion-MNIST (10-class, 28×28)
  Degradation: Coma + Vignetting (23번과 동일 파라미터)
  Correction: Global kernel / PDK (dataset-based)
  Classifier: FC (flatten → 512 → 10), no conv

  3가지 비교:
    A. Clean → FC              (oracle)
    B. Degraded → FC           (no correction)
    C. Degraded → Global → FC  (global correction)
    D. Degraded → PDK → FC     (PDK correction)

  모두 동일한 FC 가중치 사용 (clean image로 학습)
  → correction 품질만 변수

출력:
  26_classification_result.png  -- accuracy bar chart
  26_confusion.png              -- confusion matrix 4종
  터미널                        -- accuracy 수치

Run:
  python 26_classification_pdk.py
  python 26_classification_pdk.py --n-train 500 --n-test 200 --n-iter 600
"""

import argparse, os, random
import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

SEED = 42
torch.manual_seed(SEED); np.random.seed(SEED); random.seed(SEED)


# ============================================================================
# Utilities
# ============================================================================

def radial_map(H, W, device="cpu"):
    ys = torch.linspace(-1, 1, H, device=device)
    xs = torch.linspace(-1, 1, W, device=device)
    gy, gx = torch.meshgrid(ys, xs, indexing="ij")
    r = (gy**2 + gx**2).sqrt() / (2**0.5)
    return r, gy, gx

def spatially_varying_conv(x, kernels):
    B, C, H, W = x.shape
    patches = F.unfold(x, kernel_size=3, padding=1).view(B, C, 9, H*W)
    k = kernels.view(H*W, 9).T.unsqueeze(0).unsqueeze(0)
    return (patches * k).sum(dim=2).view(B, C, H, W).clamp(0, 1)

def global_conv(x, kernel_9):
    k = kernel_9.view(1, 1, 3, 3)
    return F.conv2d(x, k.expand(x.shape[1], 1, 3, 3),
                    padding=1, groups=x.shape[1]).clamp(0, 1)

def loss_fn(out, clean):
    l1 = F.l1_loss(out, clean)
    def gmap(t): return t[:,:,:,1:]-t[:,:,:,:-1], t[:,:,1:,:]-t[:,:,:-1,:]
    ox,oy = gmap(out); cx,cy = gmap(clean)
    return l1 + 0.1*(F.l1_loss(ox,cx)+F.l1_loss(oy,cy))


# ============================================================================
# Degradation (23번과 동일)
# ============================================================================

def degrade_coma_vig(x, sigma0=0.20, alpha_psf=1.5,
                     coma_k=0.40, alpha_vig=3.0):
    """
    Fashion-MNIST 28×28 grayscale 기준 degradation.
    강도 설정: 육안으로 확인 가능하되 완전히 망가지지 않는 수준
    """
    r, gy, gx = radial_map(*x.shape[2:], x.device)
    r2 = r**2; rs = r.clamp(1e-6)
    sigma = (sigma0 + alpha_psf * r2).clamp(0.1, 2.0)
    shift_y = coma_k * r2 * (gy / rs)
    shift_x = coma_k * r2 * (gx / rs)
    coords = torch.tensor([
        [-1.,-1.],[-1.,0.],[-1.,1.],
        [ 0.,-1.],[ 0.,0.],[ 0.,1.],
        [ 1.,-1.],[ 1.,0.],[ 1.,1.]], device=x.device)
    d2 = (coords**2).sum(-1)
    s = sigma.unsqueeze(-1)
    g1 = torch.exp(-d2/(2*s**2)); g1 = g1/g1.sum(-1, keepdim=True)
    sy = shift_y.unsqueeze(-1); sx = shift_x.unsqueeze(-1)
    dy = coords[:,0].unsqueeze(0).unsqueeze(0) - sy
    dx = coords[:,1].unsqueeze(0).unsqueeze(0) - sx
    g2 = torch.exp(-(dy**2+dx**2)/(2*s**2)); g2 = g2/g2.sum(-1, keepdim=True)
    w = (r*0.9).clamp(0, 0.9).unsqueeze(-1)
    k = (1-w)*g1 + w*g2; k = k/k.sum(-1, keepdim=True)
    x = spatially_varying_conv(x, k)
    V = 1.0/(1.0 + alpha_vig*r2)
    return (x * V.unsqueeze(0).unsqueeze(0)).clamp(0, 1)


# ============================================================================
# PDK models
# ============================================================================

class GlobalKernel(nn.Module):
    def __init__(self):
        super().__init__()
        init = torch.zeros(9); init[4] = 3.0
        self.kernel_logits = nn.Parameter(init)

    def forward(self, x):
        k = torch.softmax(self.kernel_logits, dim=0)
        return global_conv(x, k)


class DirectPDK(nn.Module):
    """
    Pure position-dependent kernel: H×W×9 직접 학습
    - 파라미터 수: H × W × 9  (32×32: 9,216개)
    - 6개 파라미터로 계산하는 ComaPDK와 달리
      각 픽셀 위치의 3×3 kernel weight를 직접 최적화
    - delta kernel로 초기화 (보정 없음 상태)
    """
    def __init__(self, H, W):
        super().__init__()
        # delta 초기화: center(w5)=1, 나머지=0
        init = torch.zeros(H, W, 9)
        init[:, :, 4] = 1.0
        self.kernels = nn.Parameter(init)

    def forward(self, x):
        return spatially_varying_conv(x, self.kernels)



# ============================================================================
# Classifiers
# ============================================================================

class FCClassifier(nn.Module):
    """
    Flatten → FC512 → ReLU → Dropout → FC10
    Conv 없음 → spatial feature 의존 불가
    → degradation에 직접 취약
    """
    def __init__(self, in_dim, n_classes=10):
        super().__init__()
        self.fc1 = nn.Linear(in_dim, 512)
        self.fc2 = nn.Linear(512, n_classes)
        self.drop = nn.Dropout(0.1)

    def forward(self, x):
        x = x.view(x.size(0), -1)
        x = F.relu(self.fc1(x))
        x = self.drop(x)
        return self.fc2(x)


class MLPClassifier(nn.Module):
    """
    Flatten → FC1024 → BN → ReLU → Dropout
           → FC512  → BN → ReLU → Dropout
           → FC256  → BN → ReLU
           → FC10
    Conv 없음 → FC보다 더 많은 파라미터
    → degradation에 FC만큼 직접 취약
    """
    def __init__(self, in_dim, n_classes=10):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(in_dim, 1024),
            nn.BatchNorm1d(1024),
            nn.ReLU(),
            nn.Dropout(0.1),
            nn.Linear(1024, 512),
            nn.BatchNorm1d(512),
            nn.ReLU(),
            nn.Dropout(0.1),
            nn.Linear(512, 256),
            nn.BatchNorm1d(256),
            nn.ReLU(),
            nn.Linear(256, n_classes),
        )

    def forward(self, x):
        return self.net(x.view(x.size(0), -1))


# ============================================================================
# Data: Fashion-MNIST
# ============================================================================

def load_fashion_mnist(n_train, n_test, img_size=28):
    try:
        import torchvision
        import torchvision.transforms as T
        tf = T.Compose([
            T.Resize(img_size),
            T.ToTensor(),          # already grayscale (1ch)
        ])
        train_ds = torchvision.datasets.FashionMNIST(
            root='/tmp/fashion_mnist', train=True,  download=True, transform=tf)
        test_ds  = torchvision.datasets.FashionMNIST(
            root='/tmp/fashion_mnist', train=False, download=True, transform=tf)

        idx_tr = list(range(len(train_ds))); random.shuffle(idx_tr)
        idx_te = list(range(len(test_ds)));  random.shuffle(idx_te)

        X_tr = torch.stack([train_ds[i][0] for i in idx_tr[:n_train]])
        y_tr = torch.tensor([train_ds[i][1] for i in idx_tr[:n_train]])
        X_te = torch.stack([test_ds[i][0]  for i in idx_te[:n_test]])
        y_te = torch.tensor([test_ds[i][1] for i in idx_te[:n_test]])

        print(f"  [FashionMNIST] train={len(X_tr)}, test={len(X_te)}, size={img_size}×{img_size}")
        return X_tr, y_tr, X_te, y_te, 10

    except Exception as e:
        print(f"  [Warning] FashionMNIST load failed: {e}")
        print("  [Fallback] Using synthetic data")
        return _synthetic_data(n_train, n_test, img_size)


def _synthetic_data(n_train, n_test, img_size=28):
    """Fashion-MNIST 없을 때: synthetic 10-class"""
    n_classes = 10
    def make(n):
        X, y = [], []
        for i in range(n):
            c = i % n_classes
            img = torch.zeros(1, img_size, img_size)
            cx = img_size // 2
            cy = img_size // 2
            r  = img_size // 5
            for dy in range(-r, r+1):
                for dx in range(-r, r+1):
                    if dx*dx+dy*dy < r*r:
                        iy = min(max(cy+dy, 0), img_size-1)
                        ix = min(max(cx+dx+c-5, 0), img_size-1)
                        img[0, iy, ix] = 0.5 + 0.5*(c/n_classes)
            img += torch.randn_like(img)*0.05
            img = img.clamp(0,1)
            X.append(img); y.append(c)
        return torch.stack(X), torch.tensor(y)
    X_tr, y_tr = make(n_train)
    X_te, y_te = make(n_test)
    print(f"  [Synthetic] train={n_train}, test={n_test}, classes={n_classes}")
    return X_tr, y_tr, X_te, y_te, n_classes



# ============================================================================
# Training / Evaluation
# ============================================================================

def train_classifier(model, X_tr, y_tr, n_epochs=30, lr=1e-3, batch=64,
                     eval_sets=None):
    """
    eval_sets: dict of {label: (X_te, y_te)} — epoch마다 accuracy 기록
    반환: dict of {label: [acc per epoch]}
    """
    opt = torch.optim.Adam(model.parameters(), lr=lr)

    # ── LR warmup + cosine annealing ──────────────────────────
    # epoch 1~warmup_epochs: lr = lr*0.01 → lr (linear 증가)
    # 이후: cosine 감쇠
    # → 초반 급상승 방지, 점진적 학습 곡선
    warmup_epochs = max(1, int(n_epochs * 0.3))   # 30%
    warmup = torch.optim.lr_scheduler.LinearLR(
        opt, start_factor=0.01, end_factor=1.0,
        total_iters=warmup_epochs)
    cosine = torch.optim.lr_scheduler.CosineAnnealingLR(
        opt, T_max=max(1, n_epochs - warmup_epochs))
    sched = torch.optim.lr_scheduler.SequentialLR(
        opt, schedulers=[warmup, cosine], milestones=[warmup_epochs])
    n = len(X_tr)
    history = {k: [] for k in (eval_sets or {})}

    # epoch 0: 학습 전 초기 정확도
    if eval_sets:
        model.eval()
        with torch.no_grad():
            for k, (Xv, yv) in eval_sets.items():
                pred = torch.cat([
                    model(Xv[i:i+128]).argmax(1)
                    for i in range(0, len(Xv), 128)])
                history[k].append((pred == yv).float().mean().item())

    for ep in range(n_epochs):
        model.train()
        idx = torch.randperm(n)
        total_loss = 0.0
        for i in range(0, n, batch):
            b_idx = idx[i:i+batch]
            xb = X_tr[b_idx]; yb = y_tr[b_idx]
            opt.zero_grad()
            loss = F.cross_entropy(model(xb), yb)
            loss.backward(); opt.step()
            total_loss += loss.item()

        sched.step()

        if eval_sets:
            model.eval()
            with torch.no_grad():
                for k, (Xv, yv) in eval_sets.items():
                    pred = torch.cat([
                        model(Xv[i:i+128]).argmax(1)
                        for i in range(0, len(Xv), 128)])
                    history[k].append((pred == yv).float().mean().item())

        if (ep+1) % 10 == 0:
            accs = {k: f'{v[-1]*100:.1f}%' for k,v in history.items()} if history else {}
            print(f"    epoch {ep+1}/{n_epochs}  loss={total_loss:.3f}  {accs}")

    return history

def evaluate(model, X, y, batch=128):
    model.eval()
    all_pred = []
    with torch.no_grad():
        for i in range(0, len(X), batch):
            xb = X[i:i+batch]
            pred = model(xb).argmax(dim=1)
            all_pred.append(pred)
    pred = torch.cat(all_pred)
    acc = (pred == y).float().mean().item()
    return acc, pred.numpy()

def confusion_matrix(y_true, y_pred, n_classes):
    cm = np.zeros((n_classes, n_classes), dtype=int)
    for t, p in zip(y_true, y_pred):
        cm[t, p] += 1
    return cm

def train_pdk(model, X_clean, degrade_fn, n_iter=800, lr=0.01, batch=32):
    """
    PDK correction SGD 학습.
    매 iter마다 랜덤 mini-batch 하나만 → 빠르고 일반적인 SGD
    degrade_fn은 batch마다 on-the-fly 적용
    """
    opt   = torch.optim.Adam(model.parameters(), lr=lr)
    sched = torch.optim.lr_scheduler.CosineAnnealingLR(opt, T_max=n_iter)
    n = len(X_clean)

    for it in range(n_iter):
        # 랜덤 mini-batch 샘플링
        idx   = torch.randperm(n)[:batch]
        clean = X_clean[idx]
        deg   = degrade_fn(clean)

        opt.zero_grad()
        loss = loss_fn(model(deg), clean)
        loss.backward()
        opt.step()
        sched.step()

        if (it+1) % 200 == 0:
            print(f"    iter {it+1}/{n_iter}  loss={loss.item():.5f}")


# ============================================================================
# Visualization
# ============================================================================

def save_epoch_curves(history_fc, history_cnn, n_classes, res_dir):
    """FC vs MLP epoch별 accuracy 곡선 — 2개 subplot"""
    key_meta = {
        'clean': ('Clean (oracle)',    '#1D9E75', '-',  2.2),
        'deg':   ('Degraded',          '#E24B4A', '-',  1.8),
        'glb':   ('Global correction', '#378ADD', '--', 1.8),
        'pdk':   ('PDK correction',    '#EF9F27', '-',  2.2),
    }
    n_epochs = len(next(iter(history_fc.values())))
    epochs   = list(range(0, n_epochs))

    fig, axes = plt.subplots(1, 2, figsize=(14, 6), sharey=False)

    for ax, history, title in [
        (axes[0], history_fc,  'FC  (no conv, flatten → 512 → 10)'),
        (axes[1], history_cnn, 'MLP  (3 hidden layers, 1024→512→256)'),
    ]:
        for key, (label, color, ls, lw) in key_meta.items():
            if key not in history: continue
            acc_pct = [a*100 for a in history[key]]
            ax.plot(epochs, acc_pct,
                    label=f'{label}  ({acc_pct[-1]:.1f}%)',
                    color=color, linestyle=ls, linewidth=lw)
            ax.annotate(f'{acc_pct[-1]:.1f}%',
                        xy=(epochs[-1], acc_pct[-1]),
                        xytext=(5, 0), textcoords='offset points',
                        va='center', fontsize=9,
                        color=color, fontweight='500')

        ax.set_xlabel('Epoch', fontsize=11)
        ax.set_ylabel('Test Accuracy (%)', fontsize=11)
        ax.set_title(title, fontsize=11)
        ax.legend(fontsize=9, framealpha=0.9)
        ax.set_xlim(0, n_epochs - 1)
        ax.set_ylim(0, 100)
        ax.set_yticks(range(0, 101, 20))
        ax.grid(True, alpha=0.25)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

    fig.suptitle(
        f'Classification accuracy per epoch  ({n_classes}-class Fashion-MNIST)\n'
        f'Degradation: Coma + Vignetting  |  No conv — degradation hits directly',
        fontsize=12)
    plt.tight_layout()
    path = os.path.join(res_dir, '26_epoch_accuracy.png')
    fig.savefig(path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f'[Saved] {path}')


def save_confusion(results, n_classes, res_dir, tag=''):
    keys    = ['clean','deg','glb','pdk']
    titles  = ['Clean (oracle)', 'Degraded', 'Global correction', 'PDK correction']
    colors  = ['#1D9E75','#E24B4A','#378ADD','#EF9F27']

    fig, axes = plt.subplots(1, 4, figsize=(16, 4))
    for ax, key, title, col in zip(axes, keys, titles, colors):
        cm = results[key]['cm']
        cm_norm = cm.astype(float) / cm.sum(axis=1, keepdims=True).clip(1)
        im = ax.imshow(cm_norm, cmap='Blues', vmin=0, vmax=1)
        ax.set_title(f'{title}\n{results[key]["acc"]*100:.1f}%',
                     fontsize=9, color=col, fontweight='bold')
        ax.set_xlabel('Predicted', fontsize=8)
        ax.set_ylabel('True', fontsize=8)
        ax.tick_params(labelsize=7)
        plt.colorbar(im, ax=ax, shrink=0.8)

    plt.suptitle(f'Confusion matrices [{tag}] (row-normalized)', fontsize=12)
    plt.tight_layout()
    fname = f'26_confusion_{tag.lower()}.png' if tag else '26_confusion.png'
    path = os.path.join(res_dir, fname)
    fig.savefig(path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f'[Saved] {path}')


def save_sample(X_clean, X_deg, X_glb, X_pdk, y, n_classes, res_dir):
    """각 condition별 샘플 이미지 6장"""
    n_show = min(6, len(X_clean))
    fig, axes = plt.subplots(4, n_show, figsize=(n_show*2, 9))
    row_labels = ['Clean', 'Degraded', 'Global', 'PDK']
    for r, (imgs, lbl) in enumerate(zip(
            [X_clean, X_deg, X_glb, X_pdk], row_labels)):
        for c in range(n_show):
            ax = axes[r, c]
            ax.imshow(imgs[c].squeeze().numpy(), cmap='gray', vmin=0, vmax=1)
            ax.axis('off')
            if r == 0:
                ax.set_title(f'class {y[c].item()}', fontsize=8)
        axes[r, 0].set_ylabel(lbl, fontsize=9, rotation=0,
                               labelpad=40, va='center')
    plt.suptitle('Sample images per condition', fontsize=11)
    plt.tight_layout()
    path = os.path.join(res_dir, '26_samples.png')
    fig.savefig(path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f'[Saved] {path}')
def save_per_label_images(X_clean, X_deg, X_glb, X_pdk, y, n_classes, res_dir):
    """
    라벨별로 raw / degraded / global / pdk 이미지 저장.
    각 class에서 첫 번째로 발견된 이미지 1장씩.
    출력: res_dir/per_label/class_XX_{condition}.png
    """
    label_dir = os.path.join(res_dir, 'per_label')
    os.makedirs(label_dir, exist_ok=True)

    class_idx = {}
    for i, label in enumerate(y.numpy()):
        if label not in class_idx:
            class_idx[label] = i
        if len(class_idx) == n_classes:
            break

    def save_single(arr, path):
        arr = np.clip(arr, 0, 1)
        fig, ax = plt.subplots(1, 1,
            figsize=(arr.shape[1]/100, arr.shape[0]/100), dpi=300)
        ax.imshow(arr, cmap='gray', vmin=0, vmax=1)
        ax.axis('off')
        plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
        fig.savefig(path, dpi=300, bbox_inches='tight', pad_inches=0)
        plt.close(fig)

    conditions = [
        ('raw',      X_clean),
        ('degraded', X_deg),
        ('global',   X_glb),
        ('pdk',      X_pdk),
    ]

    for cls in sorted(class_idx.keys()):
        idx = class_idx[cls]
        for cond_name, X in conditions:
            arr = X[idx].squeeze().numpy()
            path = os.path.join(label_dir, f'class_{cls:02d}_{cond_name}.png')
            save_single(arr, path)

    n_cols = 4
    col_titles = ['Clean', 'Degraded', 'Global', 'PDK']
    fig, axes = plt.subplots(n_classes, n_cols,
                             figsize=(n_cols * 2, n_classes * 2))
    for c, title in enumerate(col_titles):
        axes[0, c].set_title(title, fontsize=10, fontweight='bold', pad=4)
    for cls in sorted(class_idx.keys()):
        idx = class_idx[cls]
        axes[cls, 0].set_ylabel(f'class {cls}', fontsize=8,
                                rotation=0, labelpad=32, va='center')
        for c, (_, X) in enumerate(conditions):
            axes[cls, c].imshow(X[idx].squeeze().numpy(),
                                cmap='gray', vmin=0, vmax=1)
            axes[cls, c].axis('off')

    plt.suptitle('Per-label: raw / degraded / global / PDK', fontsize=12)
    plt.tight_layout()
    path = os.path.join(res_dir, '26_per_label_overview.png')
    fig.savefig(path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f'[Saved] {path}')
    print(f'[Saved] {label_dir}/ (individual images)')


def save_epoch_xlsx(history_fc, history_cnn, res_dir):
    """epoch별 accuracy를 xlsx로 저장 (FC, MLP 각각 sheet)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  [Warning] openpyxl not found, skipping xlsx export")
        return

    wb = Workbook()
    key_labels = {
        'clean': 'Clean (oracle)',
        'deg':   'Degraded',
        'glb':   'Global correction',
        'pdk':   'PDK correction',
    }
    colors = {
        'clean': '1D9E75',
        'deg':   'E24B4A',
        'glb':   '378ADD',
        'pdk':   'EF9F27',
    }

    for sheet_name, history in [('FC', history_fc), ('MLP', history_cnn)]:
        ws = wb.active if sheet_name == 'FC' else wb.create_sheet(sheet_name)
        ws.title = sheet_name

        keys = list(history.keys())
        headers = ['Epoch'] + [key_labels.get(k, k) for k in keys]

        # header row
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(name='Arial', bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', start_color='2C2C2A')
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = 20

        # epoch 0 ~ n
        n_epochs = len(next(iter(history.values())))
        for ep in range(n_epochs):
            row = ep + 2
            ws.cell(row=row, column=1, value=ep)
            for col, k in enumerate(keys, 2):
                val = round(history[k][ep] * 100, 2)
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = Font(name='Arial', color=colors.get(k, '000000'))
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = '0.00"%"'

    path = os.path.join(res_dir, '26_epoch_accuracy.xlsx')
    wb.save(path)
    print(f'[Saved] {path}')


def save_per_label_images(X_clean, X_deg, X_glb, X_pdk, y, n_classes, res_dir):
    """
    라벨별로 raw / degraded / global / pdk 이미지 저장.
    각 class에서 첫 번째로 발견된 이미지 1장씩.
    출력: res_dir/per_label/class_XX_{condition}.png
    """
    label_dir = os.path.join(res_dir, 'per_label')
    os.makedirs(label_dir, exist_ok=True)

    # 각 class별 대표 이미지 index 찾기
    class_idx = {}
    for i, label in enumerate(y.numpy()):
        if label not in class_idx:
            class_idx[label] = i
        if len(class_idx) == n_classes:
            break

    def save_single(arr, path):
        arr = np.clip(arr, 0, 1)
        fig, ax = plt.subplots(1, 1,
            figsize=(arr.shape[1]/100, arr.shape[0]/100), dpi=300)
        ax.imshow(arr, cmap='gray', vmin=0, vmax=1)
        ax.axis('off')
        plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
        fig.savefig(path, dpi=300, bbox_inches='tight', pad_inches=0)
        plt.close(fig)

    conditions = [
        ('raw',      X_clean),
        ('degraded', X_deg),
        ('global',   X_glb),
        ('pdk',      X_pdk),
    ]

    for cls in sorted(class_idx.keys()):
        idx = class_idx[cls]
        for cond_name, X in conditions:
            arr = X[idx].squeeze().numpy()
            path = os.path.join(label_dir, f'class_{cls:02d}_{cond_name}.png')
            save_single(arr, path)

    # 전체 overview: n_classes행 × 4열
    n_cols = 4
    col_titles = ['Clean', 'Degraded', 'Global', 'PDK']
    fig, axes = plt.subplots(n_classes, n_cols,
                             figsize=(n_cols * 2, n_classes * 2))
    for c, title in enumerate(col_titles):
        axes[0, c].set_title(title, fontsize=10, fontweight='bold', pad=4)
    for cls in sorted(class_idx.keys()):
        idx = class_idx[cls]
        axes[cls, 0].set_ylabel(f'class {cls}', fontsize=8,
                                rotation=0, labelpad=32, va='center')
        for c, (_, X) in enumerate(conditions):
            axes[cls, c].imshow(X[idx].squeeze().numpy(),
                                cmap='gray', vmin=0, vmax=1)
            axes[cls, c].axis('off')

    plt.suptitle('Per-label: raw / degraded / global / PDK', fontsize=12)
    plt.tight_layout()
    path = os.path.join(res_dir, '26_per_label_overview.png')
    fig.savefig(path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f'[Saved] {path}')
    print(f'[Saved] {label_dir}/ (individual images)')



# ============================================================================

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument('--n-train',    type=int, default=2000)
    p.add_argument('--n-test',     type=int, default=5000)
    p.add_argument('--img-size',   type=int, default=28)
    p.add_argument('--cls-epochs', type=int, default=40,
                   help='FC classifier training epochs')
    p.add_argument('--pdk-iter',   type=int, default=400,
                   help='PDK correction training iterations')
    p.add_argument('--sigma0',     type=float, default=0.20)
    p.add_argument('--alpha-psf',  type=float, default=1.5)
    p.add_argument('--coma-k',     type=float, default=0.40)
    p.add_argument('--alpha-vig',  type=float, default=3.0)
    p.add_argument('--res-dir',    type=str,   default='./res/classification_fashion')
    return p.parse_args()


def main():
    args = parse_args()
    os.makedirs(args.res_dir, exist_ok=True)
    H = W = args.img_size

    degrade = lambda x: degrade_coma_vig(
        x, args.sigma0, args.alpha_psf, args.coma_k, args.alpha_vig)

    # ── 데이터 로드 ──────────────────────────────────────────
    print('[Load] Fashion-MNIST...')
    X_tr, y_tr, X_te, y_te, n_classes = load_fashion_mnist(
        args.n_train, args.n_test, H)

    in_dim = H * W  # grayscale, for FC

    # ── Degraded / corrected 이미지 생성 ─────────────────────
    print('[Degrade] Applying coma+vignetting...')
    with torch.no_grad():
        X_tr_deg = degrade(X_tr)
        X_te_deg = degrade(X_te)

    # ── PDK correction 학습 ───────────────────────────────────
    print('\n[Train] Global kernel correction...')
    glb_model = GlobalKernel()
    train_pdk(glb_model, X_tr, degrade, n_iter=args.pdk_iter, lr=0.02)

    print('\n[Train] DirectPDK correction (H×W×9 direct params, SGD)...')
    pdk_model = DirectPDK(H, W)
    train_pdk(pdk_model, X_tr, degrade, n_iter=args.pdk_iter, lr=0.01, batch=32)

    # ── Corrected 이미지 생성 ─────────────────────────────────
    print('\n[Correct] Applying corrections...')
    with torch.no_grad():
        X_tr_glb = glb_model(X_tr_deg)
        X_te_glb = glb_model(X_te_deg)
        X_tr_pdk = pdk_model(X_tr_deg)
        X_te_pdk = pdk_model(X_te_deg)

    def psnr(a, b):
        mse = F.mse_loss(a,b).item()
        return 99.9 if mse<1e-10 else 10*np.log10(1/mse)
    print(f'  Degraded PSNR: {psnr(X_te, X_te_deg):.2f} dB')
    print(f'  Global   PSNR: {psnr(X_te, X_te_glb):.2f} dB')
    print(f'  PDK      PSNR: {psnr(X_te, X_te_pdk):.2f} dB')

    # ── 텐서 동일성 진단 ─────────────────────────────────────
    print('\n[Diagnostic] Tensor difference check:')
    def mse(a, b): return F.mse_loss(a.float(), b.float()).item()
    print(f'  MSE(clean, deg):     {mse(X_te, X_te_deg):.6f}')
    print(f'  MSE(clean, glb):     {mse(X_te, X_te_glb):.6f}')
    print(f'  MSE(clean, pdk):     {mse(X_te, X_te_pdk):.6f}')
    print(f'  MSE(deg,   glb):     {mse(X_te_deg, X_te_glb):.6f}')
    print(f'  MSE(deg,   pdk):     {mse(X_te_deg, X_te_pdk):.6f}')
    print(f'  same object? clean==deg: {X_te.data_ptr()==X_te_deg.data_ptr()}')
    print(f'  same object? deg==glb:   {X_te_deg.data_ptr()==X_te_glb.data_ptr()}')
    print(f'  same object? deg==pdk:   {X_te_deg.data_ptr()==X_te_pdk.data_ptr()}')

    # ── Global kernel 저장 ────────────────────────────────────
    with torch.no_grad():
        glb_k = torch.softmax(glb_model.kernel_logits, dim=0
                              ).numpy().reshape(3, 3)

    print(f'\n  Global kernel (3×3):\n{np.round(glb_k, 4)}')

    fig, ax = plt.subplots(1, 1, figsize=(3, 3))
    im = ax.imshow(glb_k, cmap='RdBu_r',
                   vmin=-abs(glb_k).max(), vmax=abs(glb_k).max(),
                   interpolation='nearest')
    for i in range(3):
        for j in range(3):
            ax.text(j, i, f'{glb_k[i,j]:.3f}',
                    ha='center', va='center', fontsize=9)
    plt.colorbar(im, ax=ax, shrink=0.8)
    ax.set_xticks([]); ax.set_yticks([])
    ax.set_title('Global kernel (learned)', fontsize=10)
    plt.tight_layout()
    glb_k_path = os.path.join(args.res_dir, '26_global_kernel.png')
    fig.savefig(glb_k_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f'  [Saved] {glb_k_path}')

    eval_sets = {
        'clean': (X_te,     y_te),
        'deg':   (X_te_deg, y_te),
        'glb':   (X_te_glb, y_te),
        'pdk':   (X_te_pdk, y_te),
    }

    # ── FC Classifier 학습 ────────────────────────────────────
    print('\n[Train] FC Classifier on clean images...')
    fc_model = FCClassifier(in_dim, n_classes)
    history_fc = train_classifier(
        fc_model, X_tr, y_tr,
        n_epochs=args.cls_epochs, lr=1e-3,
        eval_sets=eval_sets)

    # ── TinyCNN Classifier 학습 ───────────────────────────────
    print('\n[Train] MLP (1024→512→256) on clean images...')
    cnn_model = MLPClassifier(in_dim, n_classes)
    history_cnn = train_classifier(
        cnn_model, X_tr, y_tr,
        n_epochs=args.cls_epochs, lr=1e-3,
        eval_sets=eval_sets)

    # ── Evaluation ────────────────────────────────────────────
    print('\n[Evaluate — FC]')
    results_fc = {}
    for key, X_te_var in [
        ('clean', X_te), ('deg', X_te_deg),
        ('glb', X_te_glb), ('pdk', X_te_pdk),
    ]:
        acc, pred = evaluate(fc_model, X_te_var, y_te)
        cm = confusion_matrix(y_te.numpy(), pred, n_classes)
        results_fc[key] = {'acc': acc, 'cm': cm}
        label = {'clean':'Clean','deg':'Degraded','glb':'Global','pdk':'PDK'}[key]
        print(f'  {label:<12}: {acc*100:.1f}%')

    print('\n[Evaluate — MLP]')
    results_cnn = {}
    for key, X_te_var in [
        ('clean', X_te), ('deg', X_te_deg),
        ('glb', X_te_glb), ('pdk', X_te_pdk),
    ]:
        acc, pred = evaluate(cnn_model, X_te_var, y_te)
        cm = confusion_matrix(y_te.numpy(), pred, n_classes)
        results_cnn[key] = {'acc': acc, 'cm': cm}
        label = {'clean':'Clean','deg':'Degraded','glb':'Global','pdk':'PDK'}[key]
        print(f'  {label:<12}: {acc*100:.1f}%')

    print('\n' + '='*55)
    print(f'{"":20} {"FC":>10} {"MLP":>10}')
    print('-'*55)
    for key, lbl in [('clean','Clean (oracle)'),('deg','Degraded'),
                     ('glb','Global'),('pdk','PDK')]:
        print(f'  {lbl:<18} {results_fc[key]["acc"]*100:>9.1f}%'
              f' {results_cnn[key]["acc"]*100:>9.1f}%')
    print('='*55)

    # ── 시각화 ────────────────────────────────────────────────
    print('\n[Save]')
    save_epoch_curves(history_fc, history_cnn, n_classes, args.res_dir)
    save_epoch_xlsx(history_fc, history_cnn, args.res_dir)
    save_confusion(results_fc,  n_classes, args.res_dir, tag='FC')
    save_confusion(results_cnn, n_classes, args.res_dir, tag='MLP')
    save_sample(X_te, X_te_deg, X_te_glb, X_te_pdk, y_te,
                n_classes, args.res_dir)
    save_per_label_images(X_te, X_te_deg, X_te_glb, X_te_pdk, y_te,
                          n_classes, args.res_dir)

    print(f'\n[Done] {os.path.abspath(args.res_dir)}')


if __name__ == '__main__':
    main()