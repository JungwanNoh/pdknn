"""
33_oka_device_aware.py
======================================================================
OKA (One-shot Kernel Array) device-aware simulation.

전체 파이프라인:
  ┌─────────────────────────────────────────────────────────────────┐
  │  [Step 1] Device 측정 데이터 → 모델 fit                          │
  │    PR  (photoresponsivity)  : R(I_state) = α·exp(-I_state/I₀)    │
  │    LTD (long-term depression): I(N|V) = a₁e^(-N/τ₁)+a₂e^(-N/τ₂)+I∞│
  │                                                                   │
  │  [Step 2] DirectPDK 학습 (CPU baseline)                           │
  │    학습 변수: kernels (H, W, 9) free real values                  │
  │    Forward : ideal linear conv  out = Σ w_k · P_k                 │
  │    Loss    : L1(out, clean) + 0.1·L1(∇out, ∇clean)               │
  │                                                                   │
  │  [Step 3] DirectPDK weight → OKA I_state (post-hoc mapping)       │
  │    Weight-aware bias: scale_pos = scale_neg 되도록 R_mid 자동결정│
  │    R_target = R_mid + w/scale  →  I_state = R⁻¹(R_target)         │
  │                                                                   │
  │  [Step 4] PGM 시뮬레이션 (continuous V, 1 pulse)                  │
  │    각 cell의 target I_state → (V_continuous, N=1) 결정            │
  │    → I_actual (실제 도달 I_state, 보간 가정 하 거의 정확)         │
  │                                                                   │
  │  [Step 5] Inference: 4-panel 비교                                 │
  │    Input | Degraded | CPU(DirectPDK) | OKA(I_actual + R_mid bias) │
  │                                                                   │
  │  [Step 6] Energy / Latency 비교 (CPU vs OKA)                      │
  │    CPU: sense+ADC+DRAM I/O + MAC                                  │
  │    OKA: PGM 1회 amortize + sense(=MAC)+ADC                        │
  └─────────────────────────────────────────────────────────────────┘

입력 데이터:
  ./expdata/photoresponsivity.xlsx     (PR 측정: 4 light × 5 PGM state)
  ./expdata/depression_data.xlsx       (LTD 측정: 121 pulse × 6 V)
  ./data/BSD300/images/train           (DirectPDK 학습용)
  ./data/BSD300/images/test/...        (inference test)

실행:
  python 33_oka_device_aware.py
  python 33_oka_device_aware.py --N-budget 1 --pgm-mode continuous_V
"""

import argparse, os, glob, random
import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
from scipy.optimize import curve_fit, brentq
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from PIL import Image

SEED = 42
torch.manual_seed(SEED); np.random.seed(SEED); random.seed(SEED)


# ╔══════════════════════════════════════════════════════════════════╗
# ║                                                                  ║
# ║   SECTION 1.  Device measurement → models                        ║
# ║                                                                  ║
# ║   PR  (photoresponsivity): single-cell 광응답                    ║
# ║   LTD (long-term depression): pulse-by-pulse conductance 변화   ║
# ║                                                                  ║
# ╚══════════════════════════════════════════════════════════════════╝

class PhotoresponsivityModel:
    """
    PGM-state responsivity model (P-independent).

      R(I_state) = α · exp(-I_state / I₀)
      I_photo    = R(I_state) · P                ← γ=1, classical R = I/P

    P:        light power (μW/cm²)
    I_state:  PGM state read current (nA)
    R:        photoresponsivity [nA per μW/cm²]
    """
    def __init__(self, alpha, I0, b0, k, R_max_meas):
        self.alpha = alpha
        self.I0    = I0
        # b0, k: legacy slots (이전 모델 P^(b₀+k·I) 흔적). 새 모델 사용 안 함.
        self.b0    = b0
        self.k     = k
        self.R_max_meas = R_max_meas

    def R(self, I):
        """PGM-state responsivity. I_state만의 함수."""
        return self.alpha * np.exp(-np.asarray(I) / self.I0)

    def I_photo(self, P, I_state):
        """단일 cell 광전류 = R(I) · P."""
        return self.R(I_state) * P

    @classmethod
    def fit_from_xlsx(cls, path):
        """
        실측 PR 데이터로부터 (α, I₀) fit.

        xlsx 구조:
          row 0: 헤더 (Light power, Basecurrent (state), ...)
          row 1: ['', '159.6nA', '119.8nA', ...]  ← I_state per state
          row 2~: P [μW/cm²], I_photo[state1], I_photo[state2], ...  [nA]
        """
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))

        # I_state header (5개 상태)
        I_base_row = rows[1][1:]
        I_base = np.array([
            float(s.replace('nA','').strip())
            for s in I_base_row if s is not None
        ])

        # 본 데이터
        P_list, Iphoto_list = [], []
        for r in rows[2:]:
            if r[0] is None: continue
            P_list.append(float(r[0]))
            Iphoto_list.append([float(v) for v in r[1:1+len(I_base)]])
        P = np.array(P_list)
        I_photo_data = np.array(Iphoto_list)   # (N_P, N_state) [nA]

        # I_photo = α·exp(-I_state/I₀) · P    (γ=1)
        PP, II = np.meshgrid(P, I_base, indexing='ij')
        def model(X, alpha, I0):
            Pin, Iin = X
            return alpha * np.exp(-Iin/I0) * Pin
        p, _ = curve_fit(model, (PP.flatten(), II.flatten()),
                         I_photo_data.flatten(),
                         p0=[0.06, 150.0],
                         bounds=([0.0, 1.0], [10.0, 1000.0]),
                         maxfev=20000)

        # Fit 정확도 평가
        pred = model((PP.flatten(), II.flatten()), *p)
        rmse = np.sqrt(np.mean((I_photo_data.flatten() - pred)**2))
        rel  = np.mean(np.abs((I_photo_data.flatten() - pred) /
                       np.clip(I_photo_data.flatten(), 1e-6, None))) * 100

        print(f"[PR fit] α={p[0]:.4f}, I₀={p[1]:.2f} nA  (γ=1, P-independent R)")
        print(f"         R(I) = α·exp(-I/I₀),   I_photo = R · P")
        print(f"         rmse={rmse:.3f} nA, mean rel err={rel:.2f}%")

        # b0, k legacy slot은 0으로 채움 (외부 코드 호환)
        return cls(p[0], p[1], 0.0, 0.0, I_photo_data.max()), \
               (P, I_base, I_photo_data)


class LTDModel:
    """
    Long-Term Depression model (펄스 수에 따른 conductance 감소).

    각 V_pulse 별로:
      I(N | V) = a₁·exp(-N/τ₁) + a₂·exp(-N/τ₂) + I_inf

    Convention (중요):
      - User 시점 N=0  → 펄스 안 가함 (초기값)
      - User 시점 N=1  → 1 펄스 적용
      - Fit 내부 인덱싱: fit_N = user_N + 1
        (측정 데이터의 첫 row가 N=0 인지 N=1 인지에 따른 offset)
    """

    def __init__(self, V_levels, params, I_init_avg, I_min, I_max):
        self.V_levels   = V_levels      # 측정된 V array
        self.params     = params         # dict V → (a1, τ1, a2, τ2, I_inf)
        self.I_init_avg = I_init_avg     # 초기 conductance 평균
        self.I_min      = I_min          # 도달 가능한 최저 I
        self.I_max      = I_max          # 도달 가능한 최고 I (= I_init_avg)

    @staticmethod
    def _double_exp(N, a1, t1, a2, t2, Iinf):
        """Fit 내부용. user_N + 1 로 호출해야 일관."""
        return a1*np.exp(-N/np.abs(t1)) + a2*np.exp(-N/np.abs(t2)) + Iinf

    # ------------------------------------------------------------------
    # 1-pulse continuous-V (현재 메인으로 사용)
    # ------------------------------------------------------------------
    def best_continuous_V_1pulse(self, target_I, V_range=None):
        """
        연속 V로 1펄스 가했을 때 target_I에 도달하는 V를 보간으로 찾음.

        측정된 V level (4, 5, 6, 7, 8 V) 사이를 piecewise-linear 보간.
        no-pulse는 V=0 으로 표현 (I = I_init_avg).

        Returns:
          (V_continuous, N, achieved, err)
          target이 도달 가능 범위 안이면 err≈0.

        주의: V_intermediate 에서의 LTD 거동이 진짜 linear interp 따라가는지는
              측정 데이터 없음. 시뮬레이션 가정.
        """
        # (V, I_after_1pulse) 테이블 구축
        pts = [(0.0, float(self.I_init_avg))]   # V=0 → no pulse → init
        for V in self.V_levels:
            fit_n = 2     # user N=1 → fit n=2 (offset)
            I = float(self._double_exp(fit_n, *self.params[V]))
            pts.append((float(V), I))
        pts.sort(key=lambda x: x[0])
        Vs = np.array([p[0] for p in pts])
        Is = np.array([p[1] for p in pts])

        # Boundary cases
        I_high = Is[0]    # V=0, init (가장 큼)
        I_low  = Is[-1]   # V_max, 가장 낮음
        if target_I >= I_high:
            return 0.0, 0, I_high, abs(I_high - target_I)
        if target_I <= I_low:
            return float(Vs[-1]), 1, I_low, abs(I_low - target_I)

        # 내부 보간: target I → V (piecewise linear inverse)
        order = np.argsort(Is)
        V_target = float(np.interp(target_I, Is[order], Vs[order]))
        I_achieved = float(np.interp(V_target, Vs, Is))
        err = abs(I_achieved - target_I)
        N_used = 0 if V_target < 1e-6 else 1
        return V_target, N_used, I_achieved, err

    # ------------------------------------------------------------------
    # Discrete V × multi-pulse (대체 모드, 비교용)
    # ------------------------------------------------------------------
    def find_pulse(self, V, target_I):
        """target I_state 만들기 위한 N (V 고정). Legacy."""
        f = lambda n: self._double_exp(n, *self.params[V]) - target_I
        n_max = 1000
        if f(1) * f(n_max) > 0:
            return 1 if abs(f(1)) < abs(f(n_max)) else n_max
        try:
            n = brentq(f, 1, n_max)
            return max(1, int(round(n)))
        except:
            return 120

    def best_V_N(self, target_I, N_max=None):
        """
        target I_state에 가장 효율적으로 도달하는 (V, N).

        N_max 지정 시: discrete V × N ∈ [1, N_max] 안에서 best.
                       N=0 (no-pulse) 후보 포함.
        N_max=None: legacy mode (find_pulse 사용).
        """
        candidates = []
        if N_max is not None:
            # N=0 후보: 펄스 안 가함 → 초기값 유지
            I_init = float(self.I_init_avg)
            candidates.append((0.0, 0, I_init, abs(I_init - target_I)))
            # 측정된 V × user_N ∈ [1, N_max]
            for V in self.V_levels:
                for user_n in range(1, int(N_max) + 1):
                    fit_n = user_n + 1
                    achieved = self._double_exp(fit_n, *self.params[V])
                    err = abs(achieved - target_I)
                    candidates.append((V, user_n, achieved, err))
        else:
            # Legacy: find_pulse는 fit-internal N 반환
            for V in self.V_levels:
                n = self.find_pulse(V, target_I)
                achieved = self._double_exp(n, *self.params[V])
                err = abs(achieved - target_I)
                candidates.append((V, n, achieved, err))
        candidates.sort(key=lambda x: (x[3], x[1]))   # err, then N
        return candidates[0]

    # ------------------------------------------------------------------
    # Fitting
    # ------------------------------------------------------------------
    @classmethod
    def fit_from_xlsx(cls, path):
        """
        측정 LTD 데이터로 (V, a₁, τ₁, a₂, τ₂, I_inf) fit.

        xlsx 구조:
          col 0: pulse number
          col 1~: 각 V 별 measured current (control 포함)
        """
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        data = np.array([list(r) for r in ws.iter_rows(values_only=True)],
                         dtype=float)
        pulse = data[:, 0].astype(int)
        curves_nA = data[:, 1:] * 1e9    # A → nA

        # Control (변화 거의 없음) 컬럼 자동 검출
        ratio = curves_nA[0] / np.maximum(curves_nA[-1], 1e-6)
        control_idx = int(np.argmin(np.abs(ratio - 1.0)))
        delta = curves_nA[0] - curves_nA[-1]
        order = np.argsort(-delta)
        valid_idx = [i for i in order if i != control_idx]
        # 변화량 큰 순 → 큰 V 할당 (8→7→6→5→4)
        V_assignment = [10.0, 9.5, 9.0, 8.5, 8.0, 7.5, 7.0, 6.5, 6.0][:len(valid_idx)]
        col_to_V = dict(zip(valid_idx, V_assignment))

        # V별 double-exp fit
        params = {}
        for i in valid_idx:
            V = col_to_V[i]
            y = curves_nA[:, i]
            try:
                p, _ = curve_fit(cls._double_exp, pulse, y,
                    p0=[y[0]*0.7, 5, y[0]*0.3, 50, y[-1]],
                    bounds=([0, 0.1, 0, 0.1, 0],
                            [1e3, 1000, 1e3, 1e5, 300]),
                    maxfev=50000)
                params[V] = tuple(p)
            except Exception as e:
                print(f"  [LTD fit] V={V}V failed: {e}")

        V_levels = np.array(sorted(params.keys()))
        I_init_avg = curves_nA[0].mean()
        I_max = I_init_avg
        I_min = min(cls._double_exp(120, *params[v]) for v in V_levels)

        print(f"[LTD fit] V levels: {sorted(V_levels)}")
        print(f"          I range: {I_min:.2f} ~ {I_max:.1f} nA")
        for V in sorted(V_levels, reverse=True):
            p = params[V]
            print(f"          V={V}V: a1={p[0]:.2f}, τ1={p[1]:.2f}, "
                  f"a2={p[2]:.2f}, τ2={p[3]:.2f}, Iinf={p[4]:.3f}")
        return cls(V_levels, params, I_init_avg, I_min, I_max)


# ╔══════════════════════════════════════════════════════════════════╗
# ║                                                                  ║
# ║   SECTION 2.  Image processing primitives                        ║
# ║                                                                  ║
# ║   - radial_map: 정규화 좌표/반경 (degradation 위치 dependent)    ║
# ║   - spatially_varying_conv: 위치별 다른 kernel을 갖는 conv       ║
# ║   - degrade_coma_vig: 합성 광학 열화 (coma + vignetting)         ║
# ║                                                                  ║
# ╚══════════════════════════════════════════════════════════════════╝

def radial_map(H, W, device='cpu'):
    """이미지 좌표를 [-1, 1]로 정규화, 반경 r ∈ [0, 1] 반환."""
    ys = torch.linspace(-1, 1, H, device=device)
    xs = torch.linspace(-1, 1, W, device=device)
    gy, gx = torch.meshgrid(ys, xs, indexing='ij')
    r = (gy**2 + gx**2).sqrt() / (2**0.5)
    return r, gy, gx


def spatially_varying_conv(x, kernels):
    """
    각 픽셀 위치에서 서로 다른 3×3 kernel을 적용하는 conv.

    x:       (B, C, H, W)   입력 이미지
    kernels: (H, W, 9)      위치별 kernel weight
    """
    B, C, H, W = x.shape
    patches = F.unfold(x, kernel_size=3, padding=1).view(B, C, 9, H*W)
    k = kernels.view(H*W, 9).T.unsqueeze(0).unsqueeze(0)
    return (patches * k).sum(dim=2).view(B, C, H, W).clamp(0, 1)


def degrade_coma_vig(x, sigma0=0.40, alpha_psf=2.0,
                     coma_k=0.60, alpha_vig=4.0):
    """
    합성 광학 열화: spatially varying coma blur + vignetting.

    중심부는 미세 blur, 가장자리로 갈수록:
      - PSF σ가 커짐 (alpha_psf·r²)
      - PSF가 방사형으로 shift (coma_k·r²)
      - 밝기 감소 (vignetting: 1/(1+α·r²))
    """
    r, gy, gx = radial_map(*x.shape[2:], x.device)
    r2 = r**2; rs = r.clamp(1e-6)
    sigma = (sigma0 + alpha_psf*r2).clamp(0.1, 2.0)
    shift_y = coma_k*r2*(gy/rs)
    shift_x = coma_k*r2*(gx/rs)

    coords = torch.tensor([
        [-1.,-1.],[-1.,0.],[-1.,1.],
        [ 0.,-1.],[ 0.,0.],[ 0.,1.],
        [ 1.,-1.],[ 1.,0.],[ 1.,1.]], device=x.device)
    d2 = (coords**2).sum(-1)
    s = sigma.unsqueeze(-1)

    # Center kernel (no shift)
    g1 = torch.exp(-d2/(2*s**2)); g1 = g1/g1.sum(-1, keepdim=True)
    # Shifted kernel (coma)
    sy = shift_y.unsqueeze(-1); sx = shift_x.unsqueeze(-1)
    dy = coords[:,0].unsqueeze(0).unsqueeze(0) - sy
    dx = coords[:,1].unsqueeze(0).unsqueeze(0) - sx
    g2 = torch.exp(-(dy**2+dx**2)/(2*s**2))
    g2 = g2/g2.sum(-1, keepdim=True)

    # Mix: 중심은 g1, 가장자리는 g2
    w = (r*0.9).clamp(0, 0.9).unsqueeze(-1)
    k = (1-w)*g1 + w*g2
    k = k/k.sum(-1, keepdim=True)

    x = spatially_varying_conv(x, k)
    # Vignetting
    V = 1.0/(1.0+alpha_vig*r2)
    return (x*V.unsqueeze(0).unsqueeze(0)).clamp(0, 1)


# ╔══════════════════════════════════════════════════════════════════╗
# ║                                                                  ║
# ║   SECTION 3.  DirectPDK (CPU ideal baseline)                     ║
# ║                                                                  ║
# ║   Free real-valued kernel weights를 학습.                        ║
# ║   Device 제약 없는 ideal upper-bound.                            ║
# ║                                                                  ║
# ╚══════════════════════════════════════════════════════════════════╝

class DirectPDK(nn.Module):
    """
    CPU baseline PDK: 자유 실수값 weight per pixel/tap.

    학습 변수:  kernels (H, W, 9)   real, unconstrained
    Forward :   ideal linear conv  out = Σ w_k · P_k
    역할    :   digital domain의 이상적 PDK (device 제약 X)
    """
    def __init__(self, H, W):
        super().__init__()
        init = torch.zeros(H, W, 9)
        init[:, :, 4] = 1.0     # delta init (center tap = 1, 나머지 0)
        self.kernels = nn.Parameter(init)

    def forward(self, x):
        return spatially_varying_conv(x, self.kernels)


def loss_fn(out, clean):
    """L1 + 0.1 × edge gradient loss (edge 보존 강화)."""
    l1 = F.l1_loss(out, clean)
    def gmap(t):
        return t[:,:,:,1:]-t[:,:,:,:-1], t[:,:,1:,:]-t[:,:,:-1,:]
    ox, oy = gmap(out)
    cx, cy = gmap(clean)
    return l1 + 0.1*(F.l1_loss(ox, cx) + F.l1_loss(oy, cy))


def train_pdk(model, train_imgs, degrade_fn, n_iter, lr, batch=32):
    """Adam + Cosine schedule, mini-batch SGD."""
    opt = torch.optim.Adam(model.parameters(), lr=lr)
    sched = torch.optim.lr_scheduler.CosineAnnealingLR(opt, T_max=n_iter)
    n = len(train_imgs)
    X = torch.cat(train_imgs, dim=0)
    for i in range(n_iter):
        idx = torch.randperm(n)[:batch]
        clean = X[idx]
        deg   = degrade_fn(clean)
        opt.zero_grad()
        loss = loss_fn(model(deg), clean)
        loss.backward()
        opt.step()
        sched.step()
        if (i+1) % 200 == 0:
            print(f"    iter {i+1}/{n_iter}  loss={loss.item():.5f}")


def load_train_images(data_dir, n, size):
    """디렉토리에서 grayscale 이미지 n장 로드, center-crop + resize."""
    exts = ('*.png','*.jpg','*.jpeg','*.PNG','*.JPG','*.JPEG')
    paths = []
    for ext in exts:
        paths += glob.glob(os.path.join(data_dir, '**', ext), recursive=True)
        paths += glob.glob(os.path.join(data_dir, ext))
    paths = sorted(set(paths)); random.shuffle(paths)
    imgs = []
    for p in paths:
        try:
            img = Image.open(p).convert('L')
            w, h = img.size; s = min(w, h)
            img = img.crop(((w-s)//2, (h-s)//2, (w+s)//2, (h+s)//2))
            img = img.resize((size, size), Image.BILINEAR)
            t = torch.from_numpy(np.array(img, dtype=np.float32)/255.
                                 ).unsqueeze(0).unsqueeze(0)
            imgs.append(t)
        except: continue
        if len(imgs) == n: break
    print(f"  [Load] {len(imgs)} train images from {data_dir}")
    return imgs


# ╔══════════════════════════════════════════════════════════════════╗
# ║                                                                  ║
# ║   SECTION 4.  Weight → I_state mapping (post-hoc)                ║
# ║                                                                  ║
# ║   학습된 DirectPDK weight를 OKA cell의 PGM state로 변환.        ║
# ║   Weight-aware bias로 PGM cost 자동 최소화.                      ║
# ║                                                                  ║
# ╚══════════════════════════════════════════════════════════════════╝

def direct_to_oka_mapping(weights, pr_model, ltd_model,
                          N_budget=5, V_max=8.0, I_max_native=None):
    """
    학습된 DirectPDK weight 분포를 OKA의 I_state map으로 사후 매핑.

    핵심 아이디어:
      1. 1-pulse 도달 가능한 가장 낮은 I_state (V_max에서) → I_min
         R range = [R(I_max_native), R(I_min)]
      2. weight asymmetry (max_pos vs max_neg) 분석
      3. Weight-aware bias 자동 결정:
         R_mid = (w_pos · R_min + w_neg · R_max) / (w_pos + w_neg)
         → scale_pos = scale_neg (양쪽 균형)
         → R_mid이 weight 비대칭 따라 high-I 쪽으로 자동 시프트
         → 대부분 cell이 I 큰 영역 (cheap PGM)
      4. weight → R_target = R_mid + w/scale → I_state = R⁻¹(R_target)

    Returns:
      I_state_map (H, W, 9):  cell별 target I_state
      scale (float):          weight → R 변환 계수
      info (dict):            R_mid, R_min, R_max 등 메타데이터
    """
    if I_max_native is None:
        I_max_native = float(ltd_model.I_init_avg)

    # ─── (1) 1-pulse 도달 가능한 최저 I_state ──────────────────
    if V_max not in ltd_model.params:
        raise ValueError(f"V={V_max} not in LTD fit ({list(ltd_model.params)})")
    params_V = ltd_model.params[V_max]
    fit_n = N_budget + 1   # user N=1 → fit n=2
    I_min = float(ltd_model._double_exp(fit_n, *params_V))

    # ─── (2) R range ──────────────────────────────────────────
    R_min = float(pr_model.R(I_max_native))      # I_init에서 (가장 낮은 R)
    R_max = float(pr_model.R(I_min))              # 1-pulse 도달점에서 (가장 높은 R)

    # ─── (3) weight 분포 분석 ─────────────────────────────────
    w = weights.detach().cpu().numpy()
    w_pos = max(0.0, float(w.max()))
    w_neg = max(0.0, float(-w.min()))

    if R_max - R_min < 1e-9:
        raise ValueError(
            f"R range degenerate (R_max={R_max:.6f}, R_min={R_min:.6f}). "
            f"Increase N_budget (current={N_budget})."
        )

    # ─── (4) Weight-aware bias 자동 결정 ──────────────────────
    # scale_pos = scale_neg 되도록 R_mid 결정
    if (w_pos + w_neg) > 1e-9:
        R_mid = (w_pos * R_min + w_neg * R_max) / (w_pos + w_neg)
    else:
        R_mid = (R_min + R_max) / 2

    R_pos = R_max - R_mid     # positive weight side
    R_neg = R_mid - R_min     # negative weight side
    scale_pos = w_pos / R_pos if R_pos > 1e-9 else 0.0
    scale_neg = w_neg / R_neg if R_neg > 1e-9 else 0.0
    scale = max(scale_pos, scale_neg, 1e-9)

    # I_mid = R⁻¹(R_mid): bias 셀이 위치할 I_state
    I_mid = float(-pr_model.I0 * np.log(R_mid / pr_model.alpha))

    print(f"\n[Mapping] DirectPDK weight → OKA I_state (post-hoc, weight-aware bias)")
    print(f"          N_budget={N_budget} pulses @ V_max={V_max}V")
    print(f"          → I_min={I_min:.2f} nA (1-pulse 도달점)")
    print(f"          I range : [{I_min:.2f}, {I_max_native:.2f}] nA")
    print(f"          weight  : [{w.min():.3f}, {w.max():.3f}] "
          f"(asymmetry pos/neg = {w_pos/(w_neg+1e-9):.2f})")
    print(f"          R range : [{R_min:.4f}, {R_max:.4f}]")
    print(f"          R_mid   : {R_mid:.4f} → I_mid={I_mid:.2f} nA  "
          f"(weight-aware bias)")
    print(f"          R pos/neg headroom: {R_pos:.4f} / {R_neg:.4f}")
    print(f"          scale   : pos={scale_pos:.2f}, neg={scale_neg:.2f} → {scale:.2f}")

    # ─── (5) weight → R_target → I_state 매핑 ─────────────────
    R_target = R_mid + w / scale
    R_target = np.clip(R_target, R_min + 1e-7, R_max - 1e-7)
    # R = α·exp(-I/I₀)  →  I = -I₀·ln(R/α)
    I_state_map = -pr_model.I0 * np.log(R_target / pr_model.alpha)
    I_state_map = np.clip(I_state_map, I_min, I_max_native)

    # 매핑 결과 검증 출력 — endpoint가 R_min/R_max에 정확히 닿는지 확인용
    w_min_idx = np.unravel_index(w.argmin(), w.shape)
    w_max_idx = np.unravel_index(w.argmax(), w.shape)
    R_at_wmin = float(R_target[w_min_idx])
    R_at_wmax = float(R_target[w_max_idx])
    I_at_wmin = float(I_state_map[w_min_idx])
    I_at_wmax = float(I_state_map[w_max_idx])

    print(f"          ── endpoint mapping ──")
    print(f"          weight {w.min():+.3f} → R={R_at_wmin:.4f} (R_min={R_min:.4f}) "
          f"→ I_state={I_at_wmin:.2f} nA  (≈ I_init, no/weak pulse)")
    print(f"          weight {w.max():+.3f} → R={R_at_wmax:.4f} (R_max={R_max:.4f}) "
          f"→ I_state={I_at_wmax:.2f} nA  (V_max, strongest pulse)")
    print(f"          weight 0.000 → R={R_mid:.4f} → I_state={I_mid:.2f} nA  (bias)")
    print(f"          mapped R range : [{R_target.min():.4f}, {R_target.max():.4f}]")
    print(f"          mapped I range : [{I_state_map.min():.2f}, {I_state_map.max():.2f}] nA")

    info = dict(
        I_min=I_min, I_max=I_max_native, I_mid=I_mid,
        R_min=R_min, R_max=R_max, R_mid=R_mid,
        scale=float(scale),
        N_budget=int(N_budget), V_max=float(V_max),
        P_ref=600.0,   # legacy slot
    )
    return I_state_map, float(scale), info


# ╔══════════════════════════════════════════════════════════════════╗
# ║                                                                  ║
# ║   SECTION 5.  Programming simulation                             ║
# ║                                                                  ║
# ║   Target I_state → 실제 PGM 결과 (V, N, achieved I).             ║
# ║   Continuous V (1 pulse) 또는 discrete V (multi-pulse) 모드.     ║
# ║                                                                  ║
# ╚══════════════════════════════════════════════════════════════════╝

def simulate_programming(I_target_map, ltd_model, N_budget=5,
                          mode='continuous_V'):
    """
    각 cell의 target I_state를 실제 PGM 결과로 변환.

    mode:
      'continuous_V'         : 연속 V 1펄스 (default)
                               R 값 연속, err≈0 (보간 가정)
      'discrete_multipulse'  : 측정된 V level × N≤N_budget만 사용
                               R 값 discrete, err 큼

    Returns:
      I_actual:  (H, W, 9)  실제 도달 I_state
      V_map:     (H, W, 9)  사용된 V (continuous 모드면 실수)
      N_map:     (H, W, 9)  사용된 펄스 수
      err_map:   (H, W, 9)  |achieved - target|
    """
    H, W, K = I_target_map.shape
    I_actual = np.zeros_like(I_target_map, dtype=np.float64)
    V_map    = np.zeros_like(I_target_map, dtype=np.float64)
    N_map    = np.zeros_like(I_target_map, dtype=np.int32)
    err_map  = np.zeros_like(I_target_map, dtype=np.float64)

    for ix in range(H):
        for iy in range(W):
            for k in range(K):
                target = float(I_target_map[ix, iy, k])
                if mode == 'continuous_V':
                    V, N, achieved, err = ltd_model.best_continuous_V_1pulse(
                        target)
                else:
                    V, N, achieved, err = ltd_model.best_V_N(
                        target, N_max=N_budget)
                I_actual[ix, iy, k] = achieved
                V_map[ix, iy, k]    = V
                N_map[ix, iy, k]    = N
                err_map[ix, iy, k]  = err

    n_total = H * W * K
    n_zero  = int((N_map == 0).sum())
    print(f"\n[Programming] mode={mode}, simulated {n_total} cells")
    print(f"   pulses/cell : avg={N_map.mean():.2f}, "
          f"max={int(N_map.max())}, no-pulse={n_zero} ({100*n_zero/n_total:.1f}%)")
    if mode == 'continuous_V':
        V_active = V_map[N_map > 0]
        if len(V_active) > 0:
            print(f"   V (active)  : min={V_active.min():.2f}, "
                  f"max={V_active.max():.2f}, avg={V_active.mean():.2f} V")
    print(f"   pgm error   : avg={err_map.mean():.3f} nA, "
          f"max={err_map.max():.2f}")
    print(f"   I_actual range: [{I_actual.min():.2f}, "
          f"{I_actual.max():.2f}] nA")
    return I_actual, V_map, N_map, err_map


# ╔══════════════════════════════════════════════════════════════════╗
# ║                                                                  ║
# ║   SECTION 6.  OKA forward (inference)                            ║
# ║                                                                  ║
# ║   Programmed I_state로부터 OKA 출력 계산.                        ║
# ║   out = Σ_k (R(I_state_k) - R_mid) · P_k                         ║
# ║                                                                  ║
# ╚══════════════════════════════════════════════════════════════════╝

def oka_forward_torch(P_image, I_state, pr_model, info, P_scale,
                       output_norm=None):
    """
    OKA forward (PyTorch, autograd 가능).

    Args:
      P_image:   (B, 1, H, W) tensor in [0, 1] (image scale)
      I_state:   (H, W, 9)    tensor of I_state values (nA)
      pr_model:  PhotoresponsivityModel
      info:      dict with R_mid (weight-aware bias), I_min, I_max
      P_scale:   image [0,1] → light power (μW/cm²) 변환
      output_norm: None | float
                   None: raw output (단위 전류 기반)
                   float: 학습된 fixed scale로 나누기 (학습 안정화용)

    Returns:
      out: (B, 1, H, W) tensor

    Math:
      P_uW = P_image · P_scale
      I_photo  = R(I_state) · P_uW    where R(I) = α·exp(-I/I₀)
      bias     = R_mid · P_uW
      out_raw  = Σ_k (I_photo_k - bias_k)
               = Σ_k (R_state_k - R_mid) · P_uW_k
               = Σ_k effective_w_k · P_uW_k
    """
    B, C, H, W = P_image.shape
    P_uW = P_image * P_scale
    patches = F.unfold(P_uW, kernel_size=3, padding=1).view(B, C, 9, H, W)
    I_taps = I_state.permute(2, 0, 1).unsqueeze(0).unsqueeze(0)

    α  = pr_model.alpha
    I0 = pr_model.I0

    # I_photo = R(I_state) · P
    R_state = α * torch.exp(-I_taps / I0)        # (1,1,9,H,W)
    P_safe  = patches.clamp(min=1e-3)
    I_photo = R_state * P_safe

    # bias = R_mid · P  (effective_w = R_state − R_mid)
    # info['R_mid'] 우선 (weight-aware bias). 없으면 arithmetic fallback.
    if 'R_mid' in info:
        R_mid = float(info['R_mid'])
    else:
        I_mid = (info['I_min'] + info['I_max']) / 2.0
        R_mid = float(α * np.exp(-I_mid / I0))
    bias_per_tap = R_mid * P_safe

    output = (I_photo - bias_per_tap).sum(dim=2)   # (B, C, H, W)

    # 학습 시 stable scaling 옵션
    if output_norm is not None:
        output = output / output_norm
    return output


# ╔══════════════════════════════════════════════════════════════════╗
# ║                                                                  ║
# ║   SECTION 7.  Energy / Latency models                            ║
# ║                                                                  ║
# ║   CPU baseline: sense → ADC → DRAM I/O → MAC → DRAM write        ║
# ║   OKA: PGM 1회 (amortized) + sense(=MAC) → ADC → SRAM            ║
# ║                                                                  ║
# ║   가정값들 (출처):                                               ║
# ║   - DRAM 5 pJ/bit   (Horowitz ISSCC 2014)                        ║
# ║   - SRAM 0.05 pJ/bit                                             ║
# ║   - 10-bit SAR ADC ~5 pJ/conv (Murmann survey)                   ║
# ║   - CPU MAC 100 pJ (보수적, system-level)                        ║
# ║   - PGM: ½·C_gate·V² (capacitive, gate leakage <1nA → 무시)      ║
# ║         C_gate = 345 fF (Al2O3 29nm + Nb2O5 28nm series)         ║
# ║                                                                  ║
# ╚══════════════════════════════════════════════════════════════════╝

ENERGY = dict(
    cpu_mac_pj       = 100,    # CPU multiply-add per op (system level)
    dram_read_pj_bit = 5.0,    # DDR4 (Horowitz 2014)
    dram_write_pj_bit= 5.0,
    sram_read_pj_bit = 0.05,   # on-chip cache
    adc_pj_sample    = 5.0,    # 10-bit SAR ADC (JSSC 평균)
    tia_pj_sample    = 0.5,    # generic TIA
    sense_pj_pix     = 0.5,    # photodiode sensing
)

LATENCY = dict(
    cpu_mac_ns       = 1.0,    # ~3.5 GHz × 3-4 cycles for FMA
    dram_read_ns     = 50.0,
    dram_write_ns    = 50.0,
    sram_read_ns     = 1.0,
    adc_ns           = 100.0,  # 10-bit SAR
    tia_ns           = 10.0,
    sense_ns         = 1000.0, # photodiode response (보수적 1 μs)
)


def cpu_energy_latency(H, W, k=9, n_bit=8):
    """
    CPU 기반 한 프레임 처리 energy/latency.

    Per-pixel pipeline:
      sense → ADC → DRAM write (raw)
      → DRAM read 9 neighbors + 9 weights (no cache 가정)
      → 9× MAC → DRAM write (result)
    """
    n_pix = H * W
    e_sense    = n_pix * ENERGY['sense_pj_pix']
    e_tia      = n_pix * ENERGY['tia_pj_sample']
    e_adc      = n_pix * ENERGY['adc_pj_sample']
    e_dram_w_raw = n_pix * n_bit * ENERGY['dram_write_pj_bit']
    e_dram_r_in  = n_pix * k * n_bit * ENERGY['dram_read_pj_bit']  # 9 neighbors
    e_dram_r_w   = n_pix * k * n_bit * ENERGY['dram_read_pj_bit']  # 9 weights
    e_mac        = n_pix * k * ENERGY['cpu_mac_pj']
    e_dram_w_out = n_pix * n_bit * ENERGY['dram_write_pj_bit']

    total_pJ = (e_sense + e_tia + e_adc + e_dram_w_raw +
                e_dram_r_in + e_dram_r_w + e_mac + e_dram_w_out)

    # Latency: column-parallel sensing + multi-core SIMD MAC 가정
    cores_simd = 10 * 4
    t_sense   = LATENCY['sense_ns']
    t_tia     = LATENCY['tia_ns']
    t_adc     = H * LATENCY['adc_ns']    # row-sequential
    t_dram_io = (n_pix * (1 + 2*k + 1)) * LATENCY['dram_read_ns'] / cores_simd
    t_mac     = (n_pix * k) * LATENCY['cpu_mac_ns'] / cores_simd
    total_ns  = t_sense + t_tia + t_adc + t_dram_io + t_mac

    return total_pJ, total_ns, dict(
        sense=e_sense, tia=e_tia, adc=e_adc,
        dram_w_raw=e_dram_w_raw, dram_r_in=e_dram_r_in,
        dram_r_w=e_dram_r_w, mac=e_mac, dram_w_out=e_dram_w_out)


def oka_energy_latency(H, W, V_map, N_map, n_bit=8,
                        N_frames=1,
                        C_gate_fF=345.46):
    """
    OKA 한 프레임 처리 energy/latency.

    Programming은 1회로 보고 N_frames에 amortize.

    V_map, N_map: simulate_programming에서 미리 계산된 cell별 (V, N)

    PGM energy model — capacitive charging:
      E_per_pulse_per_cell = ½ · C_gate · V²      [J]
      E_total = Σ_cells ½·C_gate · V_cell² · N_cell

    이 device는 게이트 전류가 측정 한계(< 1 nA) 미만이라
    V × I × t 모델 대신 capacitive ½CV² 사용.
    참고: Shrivastava et al. (arXiv:1902.09417) "≤ 3 fJ/pulse" at 180 nm,
          FN tunneling current is "very low" — capacitive 항이 dominant.

    C_gate 계산 (이 device geometry):
      Active area: 5 × 30 = 150 μm²
      Stack: Al₂O₃ (29 nm, k=9) ‖ Nb₂O₅ (28 nm, k=45) 직렬
      → C_gate ≈ 345 fF/cell

    가정값 (CLI):
      C_gate_fF: 345.46 (default, geometry 기반 계산값)
    """
    # ─── Programming energy (1회) ──────────────────────────────
    # E_per_pulse = ½ · C_gate · V²
    # C_fF[fF] × V²[V²] = (1e-15 F)(V²) = 1e-15 J = 0.001 pJ
    e_per_cell = 0.5 * C_gate_fF * V_map**2 * N_map * 1e-3   # [pJ]
    e_pgm_total = float(e_per_cell.sum())
    pulse_avg = float(N_map.mean())
    pulse_max = int(N_map.max())
    n_zero    = int((N_map == 0).sum())

    # ─── Per-frame inference ───────────────────────────────────
    n_pix = H * W
    e_sense  = n_pix * ENERGY['sense_pj_pix']
    e_tia    = W * ENERGY['tia_pj_sample']            # column-parallel
    e_adc    = W * ENERGY['adc_pj_sample'] * H        # row-sequential
    e_sram_w = n_pix * n_bit * ENERGY['sram_read_pj_bit']
    e_inference = e_sense + e_tia + e_adc + e_sram_w

    # ─── Amortized total ───────────────────────────────────────
    e_amortized = e_inference + e_pgm_total / N_frames

    # Latency
    t_sense = LATENCY['sense_ns']
    t_tia   = LATENCY['tia_ns']
    t_adc   = H * LATENCY['adc_ns']
    t_sram  = LATENCY['sram_read_ns']
    total_ns = t_sense + t_tia + t_adc + t_sram

    return e_amortized, total_ns, dict(
        pgm_total=e_pgm_total, pgm_amortized=e_pgm_total/N_frames,
        sense=e_sense, tia=e_tia, adc=e_adc, sram_w=e_sram_w,
        pulse_avg=pulse_avg, pulse_max=pulse_max,
        n_zero=n_zero, inference=e_inference)


# ╔══════════════════════════════════════════════════════════════════╗
# ║                                                                  ║
# ║   SECTION 8.  Image utilities                                    ║
# ║                                                                  ║
# ╚══════════════════════════════════════════════════════════════════╝

def psnr(a, b):
    mse = F.mse_loss(a.float(), b.float()).item()
    return 99.9 if mse < 1e-10 else 10*np.log10(1.0/mse)


def ssim_simple(a, b):
    a, b = a.float(), b.float()
    mu_a = F.avg_pool2d(a, 11, stride=1, padding=5)
    mu_b = F.avg_pool2d(b, 11, stride=1, padding=5)
    mu_a2, mu_b2, mu_ab = mu_a**2, mu_b**2, mu_a*mu_b
    sa2 = F.avg_pool2d(a**2, 11, stride=1, padding=5) - mu_a2
    sb2 = F.avg_pool2d(b**2, 11, stride=1, padding=5) - mu_b2
    sab = F.avg_pool2d(a*b, 11, stride=1, padding=5) - mu_ab
    c1, c2 = 0.01**2, 0.03**2
    return ((2*mu_ab+c1)*(2*sab+c2) /
            ((mu_a2+mu_b2+c1)*(sa2+sb2+c2))).mean().item()


def load_image(path, size=256):
    img = Image.open(path).convert('L')
    w, h = img.size; s = min(w, h)
    img = img.crop(((w-s)//2, (h-s)//2, (w+s)//2, (h+s)//2))
    img = img.resize((size, size), Image.BILINEAR)
    arr = np.array(img, dtype=np.float32)/255.
    return torch.from_numpy(arr).unsqueeze(0).unsqueeze(0)


def find_image(data_dir, name):
    exts = (".png",".jpg",".jpeg",".PNG",".JPG",".JPEG")
    for ext in exts:
        c = os.path.join(data_dir, name+ext)
        if os.path.exists(c): return c
        hits = glob.glob(os.path.join(data_dir, "**", name+ext),
                         recursive=True)
        if hits: return hits[0]
    return None


# ╔══════════════════════════════════════════════════════════════════╗
# ║                                                                  ║
# ║   SECTION 9.  Main pipeline                                      ║
# ║                                                                  ║
# ╚══════════════════════════════════════════════════════════════════╝

def parse_args():
    p = argparse.ArgumentParser()
    # 입력 데이터
    p.add_argument("--pr-data",    default="./expdata/photoresponsivity.xlsx")
    p.add_argument("--ltd-data",   default="./expdata/depression_data.xlsx")
    p.add_argument("--train-dir",  default="./data/BSD300/images/train")
    p.add_argument("--inf-dir",    default="./data/BSD300/images/test")
    p.add_argument("--inf-name",   default="102061")
    p.add_argument("--img-size",   type=int, default=256)
    # PDK 학습
    p.add_argument("--n-train",    type=int, default=200)
    p.add_argument("--pdk-iter",   type=int, default=2000)
    p.add_argument("--pdk-lr",     type=float, default=0.01)
    p.add_argument("--batch",      type=int, default=32)
    # Degradation (27번과 동일 기본값)
    p.add_argument("--sigma0",     type=float, default=0.40)
    p.add_argument("--alpha-psf",  type=float, default=2.0)
    p.add_argument("--coma-k",     type=float, default=0.60)
    p.add_argument("--alpha-vig",  type=float, default=4.0)
    # OKA mapping & PGM
    p.add_argument("--P-ref",      type=float, default=600.0,
                   help="reference light intensity (μW/cm²) for mapping")
    p.add_argument("--P-scale",    type=float, default=1000.0,
                   help="image grayscale [0,1] → light power 변환")
    p.add_argument("--pgm-mode",   default='continuous_V',
                   choices=['continuous_V', 'discrete_multipulse'],
                   help="continuous_V (default): 연속 V 1펄스. "
                        "discrete_multipulse: V_levels × N_budget.")
    p.add_argument("--N-budget",   type=int, default=5,
                   help="discrete_multipulse 모드에서 cell당 허용 펄스 수.")
    p.add_argument("--V-max",      type=float, default=10.0,
                   help="가장 강한 PGM voltage (LTD 측정값 중)")
    # PGM energy 가정 (capacitive model)
    p.add_argument("--C-gate-fF",  type=float, default=345.46,
                   help="Cell gate capacitance [fF]. "
                        "Default 345.46 = (5×30 μm² active area, "
                        "Al₂O₃ 29nm/k=9 ‖ Nb₂O₅ 28nm/k=45 series).")
    # 출력
    p.add_argument("--n-frames",   type=int, default=1,
                   help="programming amortization 기준 프레임 수")
    p.add_argument("--res-dir",    default="./res/oka_device_aware")
    return p.parse_args()


def main():
    args = parse_args()
    os.makedirs(args.res_dir, exist_ok=True)
    H = W = args.img_size

    # ┌─────────────────────────────────────────────────────────────┐
    # │  Step 1. Device measurement → models                         │
    # └─────────────────────────────────────────────────────────────┘
    print("="*60)
    print("[Step 1] Device model fitting from measurement data")
    print("="*60)
    pr_model, _ = PhotoresponsivityModel.fit_from_xlsx(args.pr_data)
    ltd_model   = LTDModel.fit_from_xlsx(args.ltd_data)

    # ┌─────────────────────────────────────────────────────────────┐
    # │  Step 2. DirectPDK 학습 + OKA 사후 매핑                      │
    # └─────────────────────────────────────────────────────────────┘
    print("\n" + "="*60)
    print("[Step 2] Train DirectPDK + post-hoc OKA mapping")
    print("="*60)
    train_imgs = load_train_images(args.train_dir, args.n_train, H)
    if len(train_imgs) == 0:
        print(f"  [Error] no train images in {args.train_dir}")
        return

    degrade = lambda x: degrade_coma_vig(
        x, args.sigma0, args.alpha_psf, args.coma_k, args.alpha_vig)

    # ─── 2-A. CPU baseline (DirectPDK) 학습 ────────────────────
    print("\n  [2-A] CPU PDK: ideal linear conv")
    pdk_cpu = DirectPDK(H, W)
    train_pdk(pdk_cpu, train_imgs, degrade,
              n_iter=args.pdk_iter, lr=args.pdk_lr, batch=args.batch)
    weights_cpu = pdk_cpu.kernels.detach()
    torch.save({'kernels': weights_cpu},
               os.path.join(args.res_dir, 'pdk_cpu.pt'))
    print(f"    saved: pdk_cpu.pt  (weight range "
          f"[{weights_cpu.min():.3f}, {weights_cpu.max():.3f}])")

    # ─── 2-B. OKA 사후 매핑 (학습 X) ───────────────────────────
    print("\n  [2-B] OKA: post-hoc mapping from DirectPDK")
    I_target_np, scale_oka, info = direct_to_oka_mapping(
        weights_cpu, pr_model, ltd_model,
        N_budget=args.N_budget, V_max=args.V_max,
        I_max_native=ltd_model.I_init_avg)

    # PGM 시뮬레이션 → 실제 도달 I_state
    I_actual_np, V_map, N_map, err_map = simulate_programming(
        I_target_np, ltd_model, N_budget=args.N_budget,
        mode=args.pgm_mode)

    I_target = torch.from_numpy(I_target_np).float()
    I_actual = torch.from_numpy(I_actual_np).float()
    torch.save({
        'I_target': I_target, 'I_actual': I_actual,
        'V_map': V_map, 'N_map': N_map, 'err_map': err_map,
        'scale': scale_oka, 'info': info,
    }, os.path.join(args.res_dir, 'pdk_oka.pt'))
    print(f"    saved: pdk_oka.pt (target + actual + pgm log)")

    # ┌─────────────────────────────────────────────────────────────┐
    # │  Step 3. (skipped — 두 PDK 모두 직접 만들어짐)              │
    # └─────────────────────────────────────────────────────────────┘
    print("\n" + "="*60)
    print("[Step 3] (skipped — DirectPDK trained, OKA mapped from it)")
    print("="*60)

    # ┌─────────────────────────────────────────────────────────────┐
    # │  Step 4. Inference: 4-panel 비교                             │
    # │    Input | Degraded | CPU(DirectPDK) | OKA(programmed)      │
    # └─────────────────────────────────────────────────────────────┘
    print("\n" + "="*60)
    print("[Step 4] Inference: Input / Degraded / CPU / OKA")
    print("="*60)

    src = find_image(args.inf_dir, args.inf_name)
    if src is None:
        print(f"  [Error] {args.inf_name} not found in {args.inf_dir}")
        return
    clean = load_image(src, H)
    print(f"  Test image: {args.inf_name}")

    with torch.no_grad():
        deg = degrade(clean)
        # CPU: DirectPDK linear conv
        out_cpu = pdk_cpu(deg).clamp(0, 1)
        # OKA (programmed): 실제 PGM된 I_state로 forward (device 한계 반영)
        oka_raw = oka_forward_torch(deg, I_actual, pr_model, info,
                                     P_scale=args.P_scale, output_norm=None)
        out_oka = (oka_raw * scale_oka / args.P_scale).clamp(0, 1)
        # OKA (target): mapping 직후 I_state로 forward (math reconstruction floor)
        oka_raw_ideal = oka_forward_torch(deg, I_target, pr_model, info,
                                           P_scale=args.P_scale, output_norm=None)
        out_oka_ideal = (oka_raw_ideal * scale_oka / args.P_scale).clamp(0, 1)

    # PSNR/SSIM
    p_deg       = psnr(clean, deg);          s_deg = ssim_simple(clean, deg)
    p_cpu       = psnr(clean, out_cpu);      s_cpu = ssim_simple(clean, out_cpu)
    p_oka       = psnr(clean, out_oka);      s_oka = ssim_simple(clean, out_oka)
    p_oka_ideal = psnr(clean, out_oka_ideal)
    s_oka_ideal = ssim_simple(clean, out_oka_ideal)
    p_cpu_oka   = psnr(out_cpu, out_oka)
    s_cpu_oka   = ssim_simple(out_cpu, out_oka)

    print(f"  vs clean:")
    print(f"    Degraded         : PSNR={p_deg:.2f}, SSIM={s_deg:.4f}")
    print(f"    CPU (ideal)      : PSNR={p_cpu:.2f}, SSIM={s_cpu:.4f}")
    print(f"    OKA (target)     : PSNR={p_oka_ideal:.2f}, SSIM={s_oka_ideal:.4f}  "
          f"← math reconstruction floor")
    print(f"    OKA (programmed) : PSNR={p_oka:.2f}, SSIM={s_oka:.4f}  "
          f"← actual device output")
    print(f"  Drop from PGM error: {p_oka_ideal - p_oka:.2f} dB")
    print(f"  CPU vs OKA(prog)   : PSNR={p_cpu_oka:.2f}, SSIM={s_cpu_oka:.4f}")

    # 4-panel 합본
    fig, axes = plt.subplots(1, 4, figsize=(17, 4.5))
    panels = [
        ('Input (clean)',                          clean),
        (f'Degraded\nPSNR={p_deg:.1f}',           deg),
        (f'CPU (ideal PDK)\nPSNR={p_cpu:.1f}',    out_cpu),
        (f'OKA (device-aware)\nPSNR={p_oka:.1f}', out_oka),
    ]
    for ax, (title, img) in zip(axes, panels):
        ax.imshow(img.squeeze().numpy(), cmap='gray', vmin=0, vmax=1)
        ax.set_title(title, fontsize=10, fontweight='bold')
        ax.axis('off')
    plt.tight_layout()
    fig.savefig(os.path.join(args.res_dir, 'inference_compare.png'),
                dpi=150, bbox_inches='tight')
    plt.close(fig)

    # 개별 figure 저장 (300 dpi)
    individual_dir = os.path.join(args.res_dir, 'individual')
    os.makedirs(individual_dir, exist_ok=True)
    for name, img in [
        ('input',    clean),
        ('degraded', deg),
        ('cpu',      out_cpu),
        ('oka',      out_oka),
    ]:
        arr = np.clip(img.squeeze().numpy(), 0, 1)
        f, a = plt.subplots(1, 1,
            figsize=(arr.shape[1]/100, arr.shape[0]/100), dpi=300)
        a.imshow(arr, cmap='gray', vmin=0, vmax=1)
        a.axis('off')
        plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
        f.savefig(os.path.join(individual_dir,
                                f'{args.inf_name}_{name}.png'),
                  dpi=300, bbox_inches='tight', pad_inches=0)
        plt.close(f)
    print(f"  [Saved] individual: {individual_dir}/")

    # ┌─────────────────────────────────────────────────────────────┐
    # │  Step 5. Energy / Latency (CPU vs OKA)                       │
    # └─────────────────────────────────────────────────────────────┘
    print("\n" + "="*60)
    print("[Step 5] Energy / Latency comparison")
    print("="*60)

    e_cpu, t_cpu, br_cpu = cpu_energy_latency(H, W)
    e_oka, t_oka, br_oka = oka_energy_latency(
        H, W, V_map, N_map,
        N_frames=args.n_frames,
        C_gate_fF=args.C_gate_fF)

    print(f"\n  CPU per-frame:")
    print(f"    Energy:  {e_cpu/1e6:>10.3f} μJ")
    print(f"    Latency: {t_cpu/1e6:>10.3f} ms")
    print(f"    Breakdown:")
    for k, v in br_cpu.items():
        print(f"      {k:<14}: {v/1e6:>8.3f} μJ ({100*v/e_cpu:>5.1f}%)")

    print(f"\n  OKA per-frame (amortized over {args.n_frames} frames):")
    print(f"    Energy:  {e_oka/1e6:>10.3f} μJ")
    print(f"    Latency: {t_oka/1e3:>10.3f} μs")
    print(f"    Programming total: {br_oka['pgm_total']/1e6:.2f} μJ")
    print(f"      pulses/cell: avg={br_oka['pulse_avg']:.2f}, "
          f"max={br_oka['pulse_max']}, "
          f"no-pulse={br_oka['n_zero']} ({100*br_oka['n_zero']/(H*W*9):.1f}%)")
    print(f"      pgm error  : avg={float(err_map.mean()):.3f} nA, "
          f"max={float(err_map.max()):.2f}")
    print(f"      energy model: ½·C_gate·V², C_gate={args.C_gate_fF:.1f} fF "
          f"(gate leakage <1nA, capacitive dominant)")
    print(f"    Inference only:    {br_oka['inference']/1e6:.3f} μJ")

    print(f"\n  Comparison:")
    print(f"    Energy   ratio (CPU/OKA): {e_cpu/e_oka:>10.1f}×")
    print(f"    Latency  ratio (CPU/OKA): {t_cpu/t_oka:>10.1f}×")

    # CSV summary
    import csv
    with open(os.path.join(args.res_dir, '33_summary.csv'), 'w') as f:
        w = csv.writer(f)
        w.writerow(['metric', 'CPU', 'OKA', 'ratio (CPU/OKA)'])
        w.writerow(['Energy_uJ', f'{e_cpu/1e6:.3f}',
                    f'{e_oka/1e6:.3f}', f'{e_cpu/e_oka:.1f}'])
        w.writerow(['Latency_us', f'{t_cpu/1e3:.3f}',
                    f'{t_oka/1e3:.3f}', f'{t_cpu/t_oka:.1f}'])
        w.writerow(['PSNR_CPUvsOKA_dB', '-',
                    f'{p_cpu_oka:.2f}', '-'])
        w.writerow(['SSIM_CPUvsOKA', '-',
                    f'{s_cpu_oka:.4f}', '-'])
        w.writerow(['Pulses_per_cell_avg', '-',
                    f'{br_oka["pulse_avg"]:.2f}', '-'])
    print(f"\n[Saved] {os.path.join(args.res_dir, '33_summary.csv')}")
    print(f"[Saved] {os.path.join(args.res_dir, 'inference_compare.png')}")
    print(f"\n[Done] {os.path.abspath(args.res_dir)}")


if __name__ == '__main__':
    main()